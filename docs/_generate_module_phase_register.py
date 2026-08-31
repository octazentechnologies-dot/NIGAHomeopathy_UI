#!/usr/bin/env python3
"""
Module + Phase aligned task register.

Edit on M00–M19 module sheets (layer Status + employee assignment).
01_MASTER, track sheets and Dashboard read those cells live via INDIRECT/INDEX.

Each sub-task tracks Frontend / Backend / Database / API / Mobile / Integration:
  Scope  = Yes/No (is this function in this sub-task?)
  Status = N/A | Not Started | In Progress | Blocked | Done | Deferred
  Assigned = employee from 00_Employees
Overall Status is calculated from the six in-scope layer statuses.
"""

from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName

SRC = "/Users/OctazenWork/NIGA Project/NigaHomeopathy-UI/docs/Homeocentrum_Task_Register_Web_UI_Mobile_QA.xlsx"
OUT = "/Users/OctazenWork/NIGA Project/NigaHomeopathy-UI/docs/Homeocentrum_Module_Phase_Aligned_Task_Register.xlsx"

THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILLS = {
    "header": PatternFill("solid", fgColor="1B3A4B"),
    "web": PatternFill("solid", fgColor="1A5276"),
    "ui": PatternFill("solid", fgColor="B9770E"),
    "mobile": PatternFill("solid", fgColor="117A65"),
    "qa": PatternFill("solid", fgColor="6C3483"),
    "exist": PatternFill("solid", fgColor="D5F5E3"),
    "improve": PatternFill("solid", fgColor="FCF3CF"),
    "modify": PatternFill("solid", fgColor="FDEBD0"),
    "new": PatternFill("solid", fgColor="D6EAF8"),
    "newint": PatternFill("solid", fgColor="D5D8DC"),
    "cfg": PatternFill("solid", fgColor="E8DAEF"),
    "mig": PatternFill("solid", fgColor="FADBD8"),
    "client": PatternFill("solid", fgColor="F5B7B1"),
    "reg": PatternFill("solid", fgColor="D5F5E3"),
    "alt": PatternFill("solid", fgColor="F8F9F9"),
    "white": PatternFill("solid", fgColor="FFFFFF"),
    "gold": PatternFill("solid", fgColor="FEF9E7"),
    "pre": PatternFill("solid", fgColor="2C3E50"),
    "modhdr": PatternFill("solid", fgColor="1B3A4B"),
    "link": PatternFill("solid", fgColor="D4E6F1"),
    "edit": PatternFill("solid", fgColor="F9E79F"),
    "done": PatternFill("solid", fgColor="ABEBC6"),
    "wip": PatternFill("solid", fgColor="F9E79F"),
    "blocked": PatternFill("solid", fgColor="F5B7B1"),
    "deferred": PatternFill("solid", fgColor="D5D8DC"),
    "na": PatternFill("solid", fgColor="E5E8E8"),
    "yes": PatternFill("solid", fgColor="D5F5E3"),
    "no": PatternFill("solid", fgColor="F4F6F7"),
}
for i, hx in enumerate(
    ["7B241C", "1A5276", "B9770E", "117A65", "6C3483", "1F618D", "0E6655",
     "922B21", "4A235A", "1B4F72", "145A32", "6E2C00", "1A5276", "4A235A",
     "0E6251", "7B241C", "1F618D", "117A65", "6C3483", "34495E"]
):
    FILLS[f"phase{i}"] = PatternFill("solid", fgColor=hx)

FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_WHITE_MD = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="1B3A4B", size=18)
FONT_H = Font(name="Calibri", bold=True, color="1B3A4B", size=13)
FONT_N = Font(name="Calibri", size=10, color="1C2833")
FONT_S = Font(name="Calibri", size=9, color="5D6D7E")
FONT_BOLD = Font(name="Calibri", bold=True, size=10, color="1C2833")
FONT_LINK = Font(name="Calibri", size=9, color="1A5276", underline="single")

TYPE_FILL = {
    "Existing": FILLS["exist"],
    "Existing Improvement": FILLS["improve"],
    "Existing Modification": FILLS["modify"],
    "New": FILLS["new"],
    "New Integration": FILLS["newint"],
    "Configuration": FILLS["cfg"],
    "Data Migration": FILLS["mig"],
    "Client Decision": FILLS["client"],
    "Regression": FILLS["reg"],
}
TRACK_FILL = {
    "Web": FILLS["web"],
    "UI": FILLS["ui"],
    "Mobile": FILLS["mobile"],
    "QA": FILLS["qa"],
}
PHASE_FILL = {str(i): FILLS[f"phase{i}"] for i in range(20)}
PHASE_FILL["PRE"] = FILLS["pre"]

LAYER_STATUS_VALUES = "N/A,Not Started,In Progress,Blocked,Done,Deferred"
OVERALL_STATUS_VALUES = "N/A,Not Started,In Progress,Blocked,Done,Deferred"
WT_VALUES = "Existing,Existing Improvement,Existing Modification,New,New Integration,Configuration,Data Migration,Client Decision,Regression"
STATUS_VALUES_LIST = ["Not Started", "In Progress", "Blocked", "Done", "Deferred"]

LAYERS = ["Frontend", "Backend", "Database", "API", "Mobile", "Integration"]

DEV_MODS = [
    ("M00", "PRE — Client gates"),
    ("M01", "Foundation & Security"),
    ("M02", "Admin clinical masters"),
    ("M03", "Doctor clinical workspace"),
    ("M04", "Doctor daily operations"),
    ("M05", "Appointments & scheduling"),
    ("M06", "Reception portal"),
    ("M07", "Payment system"),
    ("M08", "Account & finance"),
    ("M09", "Trust & verification"),
    ("M10", "Patient website"),
    ("M11", "Digital prescription (eRx)"),
    ("M12", "Telemedicine"),
    ("M13", "Communications"),
    ("M14", "Support & help desk"),
    ("M15", "HomeoMeds"),
    ("M16", "Patient continuity APIs"),
    ("M17", "Patient mobile app"),
    ("M18", "Doctor mobile app"),
    ("M19", "Reports, NFR & delivery gate"),
]
DEV_NAME = {c: n for c, n in DEV_MODS}

AREA_TO_DEV = {
    "Finance": "M00",
    "Finance / Reception": "M00",
    "Platform": "M01",
    "Security": "M01",
    "Security / Trust": "M01",
    "Repertory": "M02",
    "Materia Medica": "M02",
    "Diagnosis": "M02",
    "Adverse Effect": "M02",
    "Clinical Questions": "M02",
    "3D Body": "M02",
    "Business Management": "M02",
    "Patient Board": "M03",
    "Audio AI": "M03",
    "Repertorization": "M03",
    "Anatomy": "M03",
    "Doctor Dashboard": "M04",
    "Patients": "M04",
    "Staff": "M04",
    "Doctor Profile": "M04",
    "Appointments": "M05",
    "Scheduling": "M05",
    "Booking": "M05",
    "Reception": "M06",
    "Payments": "M07",
    "Account": "M08",
    "Patient money": "M08",
    "Trust": "M09",
    "Patient Website": "M10",
    "Legal": "M10",
    "Auth": "M10",
    "eRx": "M11",
    "Telemedicine": "M12",
    "Notifications": "M13",
    "Support": "M14",
    "HomeoMeds": "M15",
    "Patient identity": "M16",
    "Patient records": "M16",
    "Continuity": "M16",
    "Privacy": "M16",
    "Patient Mobile": "M17",
    "Doctor Mobile": "M18",
    "Reports": "M19",
    "Admin": "M19",
    "QA": "M19",
    "NFR": "M19",
}

SHEET_FOR_DEV = {
    "M00": "M00_PRE_Client_Gates",
    "M01": "M01_Foundation_Security",
    "M02": "M02_Admin_Clinical",
    "M03": "M03_Clinical_Workspace",
    "M04": "M04_Doctor_Daily_Ops",
    "M05": "M05_Appointments",
    "M06": "M06_Reception",
    "M07": "M07_Payments",
    "M08": "M08_Account_Finance",
    "M09": "M09_Trust",
    "M10": "M10_Patient_Website",
    "M11": "M11_Digital_eRx",
    "M12": "M12_Telemedicine",
    "M13": "M13_Communications",
    "M14": "M14_Support",
    "M15": "M15_HomeoMeds",
    "M16": "M16_Patient_APIs",
    "M17": "M17_Patient_Mobile",
    "M18": "M18_Doctor_Mobile",
    "M19": "M19_Reports_NFR_Gate",
}

EMPLOYEES = [
    ("UNASSIGNED", "Unassigned", "Any", "Any"),
    ("E01", "Gourav Nikam", "Frontend", "UI"),
    ("E02", "", "Frontend", "UI"),
    ("E03", "Gourav Nikam", "Backend", "Web"),
    ("E04", "", "Backend", "Web"),
    ("E05", "Gourav Nikam", "Database", "Web"),
    ("E06", "Gourav Nikam", "API", "Web"),
    ("E07", "", "API", "Web"),
    ("E08", "", "Mobile", "Mobile"),
    ("E09", "", "Mobile", "Mobile"),
    ("E10", "", "Integration", "Web"),
    ("E11", "", "QA", "QA"),
    ("E12", "", "QA", "QA"),
    ("E13", "Gourav Nikam", "Tech Lead", "Any"),
]

HEADERS = [
    "Module Seq",
    "Original Seq",
    "Dev Module Code",
    "Dev Module",
    "Track",
    "Phase",
    "Phase Name",
    "WBS",
    "Main Task ID",
    "Main Task",
    "Sub Task ID",
    "Sub Task",
    "Work Type",
    "Source Module / Area",
    "PDF Section",
    "PDF Feature (exact wording)",
    "PDF Delivery Text",
    "Application Surface",
    "Layer",
    "Assigned To (lead)",
]
for layer in LAYERS:
    HEADERS.extend([layer, f"{layer} Status", f"{layer} Assigned"])
HEADERS.extend(["Layers Remaining", "Overall Status", "Notes", "Open Module Sheet"])

COL = {name: i + 1 for i, name in enumerate(HEADERS)}
SUB_COL = COL["Sub Task ID"]
CODE_COL = COL["Dev Module Code"]
LEAD_COL = COL["Assigned To (lead)"]
REMAIN_COL = COL["Layers Remaining"]
OVERALL_COL = COL["Overall Status"]
NOTES_COL = COL["Notes"]
OPEN_COL = COL["Open Module Sheet"]
NCOLS = len(HEADERS)
EMP_HELPER_COL = NCOLS + 1  # AQ — live copy of 00_Employees column B
EMP_LIST_FIRST = 3
EMP_LIST_LAST = 100
EMP_RANGE = f"'00_Employees'!$B${EMP_LIST_FIRST}:$B${EMP_LIST_LAST}"

APPS_SCRIPT = r'''/**
 * Homeocentrum — Assigned To dropdowns follow 00_Employees column B.
 * One-time setup in Google Sheets:
 *   Extensions → Apps Script → paste this file → Save → Run setupAssignedDropdowns → Allow.
 * After that, type a new name in 00_Employees column B and it appears on every module sheet.
 */
function onOpen() {
  SpreadsheetApp.getUi()
    .createMenu("Homeocentrum")
    .addItem("Connect Assigned dropdowns to Employees", "setupAssignedDropdowns")
    .addToUi();
}

function setupAssignedDropdowns() {
  var ss = SpreadsheetApp.getActive();
  var emp = ss.getSheetByName("00_Employees");
  if (!emp) {
    SpreadsheetApp.getUi().alert("Sheet 00_Employees was not found.");
    return;
  }
  var nameRange = emp.getRange("B3:B100");
  var rule = SpreadsheetApp.newDataValidation()
    .requireValueInRange(nameRange, true)
    .setAllowInvalid(true)
    .setHelpText("Pick a name from 00_Employees column B. New names appear automatically.")
    .build();

  var sheetsDone = 0;
  ss.getSheets().forEach(function (sh) {
    if (!/^M\d{2}_/.test(sh.getName())) return;
    var lastCol = Math.max(sh.getLastColumn(), 42);
    var headers = sh.getRange(2, 1, 1, lastCol).getValues()[0];
    var lastRow = Math.max(sh.getLastRow(), 3);
    var height = lastRow - 2;
    if (height < 1) return;
    headers.forEach(function (h, i) {
      var title = String(h || "");
      if (title === "Assigned To (lead)" || / Assigned$/.test(title)) {
        sh.getRange(3, i + 1, height, 1).setDataValidation(rule);
      }
    });
    sheetsDone++;
  });
  SpreadsheetApp.getUi().alert(
    "Done. Assigned To on " + sheetsDone +
    " module sheets now follows 00_Employees column B.\n\n" +
    "Type a new name in column B — it appears in the dropdown. You do not need to run this again."
  );
}
'''


def tcol(name):
    """Excel table structured reference. Extra brackets when the header has spaces."""
    if any(ch.isspace() for ch in name) or any(ch in name for ch in "()"):
        return f"T_MASTER[[{name}]]"
    return f"T_MASTER[{name}]"

DATA_WIDTHS = (
    [11, 12, 14, 32, 10, 8, 34, 10, 14, 46, 16, 52, 20, 20, 12, 28, 36, 22, 14, 22]
    + [10, 14, 20] * 6
    + [12, 16, 22, 20]
)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def phase_key(ph):
    s = str(ph)
    if s == "PRE":
        return -1
    try:
        return int(s)
    except ValueError:
        return 99


def track_key(tr):
    return {"Web": 0, "UI": 1, "Mobile": 2, "QA": 3}.get(tr, 9)


def load_rows():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    rows = []
    for sheet in ("01_Web_Sequential", "02_UI_Sequential", "03_Mobile_Sequential", "04_QA_Sequential"):
        ws = wb[sheet]
        h = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        for row in ws.iter_rows(min_row=3, values_only=True):
            d = dict(zip(h, row))
            if not d.get("Sub Task ID"):
                continue
            area = d.get("Module / Area")
            code = AREA_TO_DEV.get(area)
            if not code:
                raise SystemExit(f"Unmapped module area: {area!r}")
            if str(d.get("Phase")) == "PRE":
                code = "M00"
            d["Dev Module Code"] = code
            d["Dev Module"] = DEV_NAME[code]
            d["Original Seq"] = int(d.get("Original Seq") or 0)
            for layer in LAYERS:
                d[layer] = d.get(layer) or "No"
            rows.append(d)
    wb.close()
    if len(rows) != 1000:
        raise SystemExit(f"Expected 1000 rows, got {len(rows)}")
    return rows


def sort_module_phase(rows):
    return sorted(
        rows,
        key=lambda d: (
            d["Dev Module Code"],
            phase_key(d.get("Phase")),
            d["Original Seq"],
            track_key(d.get("Track")),
        ),
    )


def sheet_lookup(view_row):
    c = get_column_letter(CODE_COL)
    return f"VLOOKUP(${c}{view_row},ModMap!$A$2:$B$21,2,FALSE)"


def pull_from_module(col_idx, view_row):
    """Live value from the module sheet that owns this Sub Task ID."""
    letter = get_column_letter(col_idx)
    sub = get_column_letter(SUB_COL)
    sh = sheet_lookup(view_row)
    return (
        f'=IFERROR(INDEX(INDIRECT("\'"&{sh}&"\'!{letter}:{letter}"),'
        f'MATCH(${sub}{view_row},INDIRECT("\'"&{sh}&"\'!${sub}:${sub}"),0)),"")'
    )


def open_module_formula(view_row):
    sub = get_column_letter(SUB_COL)
    sh = sheet_lookup(view_row)
    st = get_column_letter(COL["Frontend Status"])
    return (
        f'=IFERROR(HYPERLINK("#\'"&{sh}&"\'!{st}"&'
        f'MATCH(${sub}{view_row},INDIRECT("\'"&{sh}&"\'!${sub}:${sub}"),0),'
        f'"Edit on module sheet"),"")'
    )


def remaining_formula(r):
    bits = []
    for layer in LAYERS:
        sc = get_column_letter(COL[layer])
        st = get_column_letter(COL[f"{layer} Status"])
        bits.append(f'IF(AND({sc}{r}="Yes",{st}{r}<>"Done",{st}{r}<>"Deferred",{st}{r}<>"N/A"),1,0)')
    return "=" + "+".join(bits)


def overall_formula(r):
    scope, done, blocked, wip = [], [], [], []
    for layer in LAYERS:
        sc = get_column_letter(COL[layer])
        st = get_column_letter(COL[f"{layer} Status"])
        scope.append(f'IF({sc}{r}="Yes",1,0)')
        done.append(f'IF(AND({sc}{r}="Yes",{st}{r}="Done"),1,0)')
        blocked.append(f'IF(AND({sc}{r}="Yes",{st}{r}="Blocked"),1,0)')
        wip.append(f'IF(AND({sc}{r}="Yes",OR({st}{r}="In Progress",{st}{r}="Done",{st}{r}="Deferred")),1,0)')
    s = "+".join(scope)
    d = "+".join(done)
    b = "+".join(blocked)
    w = "+".join(wip)
    return (
        f'=IF(({s})=0,"N/A",'
        f'IF(({b})>0,"Blocked",'
        f'IF(({d})=({s}),"Done",'
        f'IF(({w})>0,"In Progress","Not Started"))))'
    )


def apply_status_cf(ws, col_idx, first, last):
    letter = get_column_letter(col_idx)
    rng = f"{letter}{first}:{letter}{last}"
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="Done"'], fill=FILLS["done"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="In Progress"'], fill=FILLS["wip"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="Blocked"'], fill=FILLS["blocked"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="Deferred"'], fill=FILLS["deferred"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="N/A"'], fill=FILLS["na"]))
    ws.conditional_formatting.add(rng, FormulaRule(formula=[f'{letter}{first}="Not Started"'], fill=FILLS["gold"]))


def write_banner(ws, text, fill, is_module=False):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    ws["A1"] = text
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = fill
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 44
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(2, i, h)
        cell.fill = fill
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
        if is_module and (h.endswith(" Status") or h == "Assigned To (lead)" or h.endswith(" Assigned") or h == "Notes"):
            cell.fill = FILLS["edit"]
            cell.font = Font(name="Calibri", bold=True, color="1C2833", size=10)


def identity_vals(d, seq_in_sheet):
    return [
        seq_in_sheet,
        d.get("Original Seq"),
        d.get("Dev Module Code"),
        d.get("Dev Module"),
        d.get("Track"),
        d.get("Phase"),
        d.get("Phase Name"),
        d.get("WBS"),
        d.get("Main Task ID"),
        d.get("Main Task"),
        d.get("Sub Task ID"),
        d.get("Sub Task"),
        d.get("Work Type"),
        d.get("Module / Area"),
        d.get("PDF Section"),
        d.get("PDF Feature (exact wording)"),
        d.get("PDF Delivery Text"),
        d.get("Application Surface"),
        d.get("Layer"),
    ]


def apply_row_style(ws, r, i, d, editable_cols):
    wrap_cols = {4, 7, 10, 12, 16, 17, NOTES_COL}
    for c in range(1, NCOLS + 1):
        cell = ws.cell(r, c)
        cell.border = THIN
        if c == OPEN_COL:
            cell.font = FONT_LINK
            cell.alignment = CENTER
            cell.fill = FILLS["link"]
        elif c in wrap_cols:
            cell.font = FONT_N
            cell.alignment = WRAP
        else:
            cell.font = FONT_N
            cell.alignment = CENTER if c not in wrap_cols else WRAP
        if c == 5:
            cell.fill = TRACK_FILL.get(d.get("Track"), FILLS["white"])
            cell.font = FONT_WHITE
        elif c == 6:
            cell.fill = PHASE_FILL.get(str(d.get("Phase")), FILLS["white"])
            cell.font = FONT_WHITE
        elif c == 13:
            cell.fill = TYPE_FILL.get(d.get("Work Type"), FILLS["white"])
        elif c in editable_cols:
            cell.fill = FILLS["edit"]
        elif HEADERS[c - 1] in LAYERS:
            cell.fill = FILLS["yes"] if cell.value == "Yes" else FILLS["no"]
        elif i % 2 == 0 and c not in (5, 6, 13) and c not in editable_cols and c != OPEN_COL:
            if HEADERS[c - 1] not in LAYERS:
                cell.fill = FILLS["alt"]
    ws.row_dimensions[r].height = 42


def write_module_sheet(wb, name, banner_fill, banner, rows):
    """SOURCE OF TRUTH — yellow Status / Assigned cells are editable values."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = "B9770E"
    write_banner(ws, banner, banner_fill, is_module=True)
    ordered = sort_module_phase(rows)
    editable = {LEAD_COL, NOTES_COL}
    for layer in LAYERS:
        editable.add(COL[f"{layer} Status"])
        editable.add(COL[f"{layer} Assigned"])
    for i, d in enumerate(ordered, 1):
        r = i + 2
        vals = identity_vals(d, i) + ["Unassigned"]
        for layer in LAYERS:
            scope = d.get(layer) or "No"
            vals.append(scope)
            vals.append("Not Started" if scope == "Yes" else "N/A")
            vals.append("Unassigned" if scope == "Yes" else "")
        vals.extend([remaining_formula(r), overall_formula(r), "", "EDIT HERE — this module sheet"])
        for c, v in enumerate(vals, 1):
            ws.cell(r, c, v)
        apply_row_style(ws, r, i, d, editable)
        ws.cell(r, OPEN_COL).font = FONT_BOLD
        ws.cell(r, OPEN_COL).fill = FILLS["edit"]
    last = len(ordered) + 2
    finish_data_sheet(ws, last, is_module=True, use_table=False, emp_names=unique_employee_names())
    return ws


def write_mirror_sheet(wb, name, tab_hex, banner_fill, banner, rows, sort_fn, use_table=False):
    """Reads live from the owning module sheet."""
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_hex
    write_banner(ws, banner, banner_fill, is_module=False)
    ordered = sort_fn(rows)
    pull_cols = [LEAD_COL, NOTES_COL]
    for layer in LAYERS:
        pull_cols.append(COL[f"{layer} Status"])
        pull_cols.append(COL[f"{layer} Assigned"])
    for i, d in enumerate(ordered, 1):
        r = i + 2
        vals = identity_vals(d, i)
        vals.append(None)  # lead — formula below
        for layer in LAYERS:
            vals.append(d.get(layer) or "No")
            vals.append(None)  # status
            vals.append(None)  # assigned
        vals.extend([None, None, None, None])  # remaining, overall, notes, open
        for c, v in enumerate(vals, 1):
            if v is not None:
                ws.cell(r, c, v)
        ws.cell(r, LEAD_COL, pull_from_module(LEAD_COL, r))
        for layer in LAYERS:
            ws.cell(r, COL[f"{layer} Status"], pull_from_module(COL[f"{layer} Status"], r))
            ws.cell(r, COL[f"{layer} Assigned"], pull_from_module(COL[f"{layer} Assigned"], r))
        ws.cell(r, REMAIN_COL, remaining_formula(r))
        ws.cell(r, OVERALL_COL, overall_formula(r))
        ws.cell(r, NOTES_COL, pull_from_module(NOTES_COL, r))
        ws.cell(r, OPEN_COL, open_module_formula(r))
        apply_row_style(ws, r, i, d, editable_cols=set())
    last = len(ordered) + 2
    finish_data_sheet(ws, last, is_module=False, use_table=use_table)
    return ws


def unique_employee_names(wb=None):
    """Names that appear in the Assigned dropdown (same style as Status: in-cell list)."""
    names = []
    seen = set()

    def add(n):
        if n is None:
            return
        n = str(n).strip()
        if not n or n.lower().startswith("replace with name"):
            return
        if "," in n or '"' in n:
            n = n.replace(",", " ").replace('"', "")
        key = n.lower()
        if key in seen:
            return
        seen.add(key)
        names.append(n)

    add("Unassigned")
    if wb is not None and "00_Employees" in wb.sheetnames:
        ws = wb["00_Employees"]
        for r in range(EMP_LIST_FIRST, EMP_LIST_LAST + 1):
            add(ws.cell(r, 2).value)
    else:
        for _eid, name, _fn, _tr in EMPLOYEES:
            add(name)
    return names


def emp_items_formula1(names):
    joined = ",".join(names)
    if len(joined) > 250:
        joined = ",".join(names[: max(1, 250 // 12)])
    return f'"{joined}"'


def add_emp_helper_column(ws, names=None):
    """Live formulas from 00_Employees!B so the name list grows when column B grows."""
    letter = get_column_letter(EMP_HELPER_COL)
    ws.cell(1, EMP_HELPER_COL, "Live copy of 00_Employees column B — do not edit. New names appear here automatically.")
    ws.cell(1, EMP_HELPER_COL).font = FONT_S
    ws.cell(1, EMP_HELPER_COL).fill = FILLS["gold"]
    ws.cell(2, EMP_HELPER_COL, "_EmpList")
    ws.cell(2, EMP_HELPER_COL).font = FONT_BOLD
    ws.cell(2, EMP_HELPER_COL).fill = FILLS["edit"]
    for r in range(EMP_LIST_FIRST, EMP_LIST_LAST + 1):
        cell = ws.cell(r, EMP_HELPER_COL, f"='00_Employees'!B{r}")
        cell.font = FONT_S
        cell.border = THIN
    ws.column_dimensions[letter].width = 22


def make_emp_validation(names=None):
    """Dropdown reads 00_Employees column B. New names in B appear without rebuilding the file."""
    return DataValidation(
        type="list",
        formula1=f"={EMP_RANGE}",
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle="Assigned To",
        prompt="Pick a name from 00_Employees column B. Type a new name there and it appears here.",
    )


def finish_data_sheet(ws, last, is_module, use_table=False, emp_names=None):
    for col_idx in [COL[f"{layer} Status"] for layer in LAYERS] + [OVERALL_COL]:
        apply_status_cf(ws, col_idx, 3, last)
    if is_module:
        add_emp_helper_column(ws)
        dv_st = DataValidation(type="list", formula1=f'"{LAYER_STATUS_VALUES}"', allow_blank=False)
        dv_emp = make_emp_validation()
        ws.add_data_validation(dv_st)
        ws.add_data_validation(dv_emp)
        dv_emp.add(f"{get_column_letter(LEAD_COL)}3:{get_column_letter(LEAD_COL)}{last}")
        for layer in LAYERS:
            sl = get_column_letter(COL[f"{layer} Status"])
            al = get_column_letter(COL[f"{layer} Assigned"])
            dv_st.add(f"{sl}3:{sl}{last}")
            dv_emp.add(f"{al}3:{al}{last}")
    set_widths(ws, DATA_WIDTHS)
    # Freeze header rows only. Freezing many columns (e.g. U3) makes Excel Online
    # show "window is too small to properly display this sheet".
    ws.freeze_panes = "A3"
    if not use_table:
        ws.auto_filter.ref = f"A2:{get_column_letter(NCOLS)}{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_view.zoomScale = 80


def write_employees(wb):
    ws = wb.create_sheet("00_Employees")
    ws.sheet_properties.tabColor = "B9770E"
    ws.merge_cells("A1:E1")
    ws["A1"] = (
        "EMPLOYEES — type real names in column B (yellow). Keep row 3 as Unassigned. "
        "Google Sheets: open 00_Sheets_Dropdown once and connect the dropdowns. After that, a new name in column B appears on every module sheet by itself."
    )
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 36
    for i, h in enumerate(["Emp ID", "Name (shows in Assigned dropdowns)", "Function", "Track", "Active"], 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["edit"]
        cell.font = FONT_BOLD
        cell.border = THIN
    for i, (eid, name, fn, tr) in enumerate(EMPLOYEES, 3):
        ws.cell(i, 1, eid).border = THIN
        ws.cell(i, 2, name if name else None).border = THIN
        ws.cell(i, 2).fill = FILLS["edit"]
        ws.cell(i, 3, fn).border = THIN
        ws.cell(i, 4, tr).border = THIN
        ws.cell(i, 5, "Yes").border = THIN
    last_emp = EMP_LIST_LAST
    for i in range(3 + len(EMPLOYEES), last_emp + 1):
        ws.cell(i, 1, "").border = THIN
        ws.cell(i, 2, "").border = THIN
        ws.cell(i, 2).fill = FILLS["gold"]
        ws.cell(i, 3, "").border = THIN
        ws.cell(i, 4, "").border = THIN
        ws.cell(i, 5, "Yes").border = THIN
    dv_fn = DataValidation(
        type="list",
        formula1='"Frontend,Backend,Database,API,Mobile,Integration,QA,Tech Lead,Any"',
        allow_blank=True,
    )
    dv_tr = DataValidation(type="list", formula1='"Web,UI,Mobile,QA,Any"', allow_blank=True)
    ws.add_data_validation(dv_fn)
    ws.add_data_validation(dv_tr)
    dv_fn.add(f"C3:C{last_emp}")
    dv_tr.add(f"D3:D{last_emp}")
    set_widths(ws, [14, 42, 16, 12, 10])
    # Direct range — module sheets use a same-sheet copy for Excel Online dropdowns.
    wb.defined_names.add(
        DefinedName(
            name="EmpList",
            attr_text=EMP_RANGE,
        )
    )
    return ws


def write_sheets_setup(wb):
    ws = wb.create_sheet("00_Sheets_Dropdown")
    ws.sheet_properties.tabColor = "B9770E"
    ws.merge_cells("A1:B1")
    ws["A1"] = "GOOGLE SHEETS — make Assigned To grow when you type a name in 00_Employees column B"
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 36
    steps = [
        "Excel cannot do this after you upload to Google Sheets. Do this once in Google Sheets. After that, new names in column B appear by themselves.",
        "1. This file must be open in Google Sheets (File → Import → Replace, or upload to Drive and open with Sheets).",
        "2. Menu: Extensions → Apps Script.",
        "3. Delete any code in Code.gs. Copy ALL of the script in cell A20 below. Paste it. Click Save (disk icon).",
        "4. Function dropdown: choose setupAssignedDropdowns. Click Run. Click Allow / Review permissions for your Google account.",
        "5. Reload this spreadsheet. You should see a Homeocentrum menu. You can also use Homeocentrum → Connect Assigned dropdowns to Employees.",
        "6. Go to 00_Employees. Type a real name in a yellow cell in column B. Open M05_Appointments. Click Assigned To (lead) — the new name is in the list. No second setup.",
        "Do not type Assigned names on Master. Edit on the module sheet. Keep row 3 as Unassigned.",
    ]
    for i, line in enumerate(steps, 3):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        ws.cell(i, 1, line).alignment = WRAP
        ws.cell(i, 1).font = FONT_N
        ws.row_dimensions[i].height = 32
        if i == 3:
            ws.cell(i, 1).fill = FILLS["gold"]
            ws.cell(i, 1).font = FONT_BOLD
    ws.merge_cells("A19:B19")
    ws["A19"] = "Copy everything in A20 (the script). Paste into Apps Script."
    ws["A19"].font = FONT_WHITE
    ws["A19"].fill = FILLS["header"]
    ws.merge_cells("A20:B20")
    ws["A20"] = APPS_SCRIPT.strip()
    ws["A20"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A20"].font = Font(name="Consolas", size=9, color="1C2833")
    ws.row_dimensions[20].height = 420
    ws.column_dimensions["A"].width = 100
    ws.column_dimensions["B"].width = 20
    return ws


def write_modmap(wb):
    ws = wb.create_sheet("ModMap")
    ws["A1"] = "Code"
    ws["B1"] = "Sheet"
    ws["A1"].font = FONT_WHITE
    ws["B1"].font = FONT_WHITE
    ws["A1"].fill = FILLS["header"]
    ws["B1"].fill = FILLS["header"]
    for i, (code, _) in enumerate(DEV_MODS, 2):
        ws.cell(i, 1, code)
        ws.cell(i, 2, SHEET_FOR_DEV[code])
    set_widths(ws, [12, 28])
    ws.sheet_state = "hidden"
    return ws


def write_cover(wb, rows):
    ws = wb.active
    ws.title = "00_Cover"
    ws.sheet_properties.tabColor = "1B3A4B"
    ws.merge_cells("B2:H2")
    ws["B2"] = "HOMEOCENTRUM — MODULE REGISTER WITH EMPLOYEE + LAYER STATUS (LIVE SYNC)"
    ws["B2"].font = FONT_TITLE
    ws.merge_cells("B3:H3")
    ws["B3"] = (
        "1000 sub-tasks · edit yellow cells on M00–M19 · Frontend/Backend/Database/API/Mobile/Integration "
        "each has Scope + Status + Assigned · Master and Dashboard follow automatically"
    )
    ws["B3"].font = FONT_S
    ws.merge_cells("B5:H5")
    ws["B5"] = "How to work this file (employees + layer status + sync)"
    ws["B5"].font = FONT_WHITE_MD
    ws["B5"].fill = FILLS["header"]
    rules = [
        "1. Google Sheets: open 00_Sheets_Dropdown, connect Assigned dropdowns once. Then type names on 00_Employees column B — the list grows by itself. Keep “Unassigned” in row 3.",
        "2. EDIT on the module sheet (M05_Appointments, M17_Patient_Mobile, …). Yellow cells: Assigned To (lead), each function’s Status, each function’s Assigned, Notes.",
        "3. Frontend / Backend / Database / API / Mobile / Integration: Scope Yes = this sub-task includes that work. Status N/A = not in this sub-task (Scope No). Status Not Started / In Progress / Done = remaining vs complete.",
        "4. Change Frontend Status (or any layer Status) on the module sheet. 01_MASTER, 03_Web, 04_UI, 05_Mobile, 06_QA, 02_By_Module_Phase and 08_Dashboard update immediately (Excel formulas — no copy-paste).",
        "5. Overall Status is calculated: all in-scope layers Done → Done; any Blocked → Blocked; any started but not all Done → In Progress; else Not Started. Do not type over Overall Status.",
        "6. Layers Remaining = how many in-scope functions are not Done yet. Filter that column to see what is left in the module.",
        "7. Same Sub Task ID is one row of work. Assign a lead plus a person per in-scope function (Frontend Assigned, API Assigned, …) when more than one employee works the row.",
        "8. Do not type over Status on 01_MASTER or track sheets — those cells are live mirrors. Use “Open Module Sheet” to jump to the yellow editable row.",
        "9. Develop sequentially inside a module (Web API → UI → Mobile → QA). Do not skip rows. Dashboard integrity must stay 1000.",
        "10. Work Type (Existing / Existing Improvement / New / …) is the type of work, not completion. Completion is the six function Status columns.",
    ]
    r = 6
    for line in rules:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
        ws.cell(r, 2, line).alignment = WRAP
        ws.cell(r, 2).font = FONT_N
        ws.row_dimensions[r].height = 36
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2, "Legend — function columns").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    for label, fill, txt in [
        ("Yes", FILLS["yes"], "Scope — this sub-task includes this function (Frontend/Backend/Database/API/Mobile/Integration)."),
        ("No", FILLS["no"], "Scope — not part of this sub-task. Status must stay N/A."),
        ("Not Started", FILLS["gold"], "In scope and remaining — not begun."),
        ("In Progress", FILLS["wip"], "In scope and remaining — someone is working it."),
        ("Done", FILLS["done"], "In scope and complete."),
        ("Blocked", FILLS["blocked"], "Cannot finish until a dependency or decision is cleared."),
        ("N/A", FILLS["na"], "Not in scope for this row."),
        ("Yellow cell", FILLS["edit"], "Editable on the MODULE sheet only (assignment + layer Status)."),
    ]:
        ws.cell(r, 2, label).fill = fill
        ws.cell(r, 2).font = FONT_BOLD
        ws.cell(r, 2).border = THIN
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        ws.cell(r, 3, txt).border = THIN
        r += 1
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r, 2, "Modules — open the sheet named here to assign people and mark layer Status").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    for col, txt in enumerate(["Code", "Dev Module", "Edit sheet", "Sub tasks", "Web", "UI", "Mobile", "QA"], 2):
        cell = ws.cell(r, col, txt)
        cell.font = FONT_WHITE
        cell.fill = FILLS["header"]
        cell.border = THIN
    r += 1
    by_mod_track = defaultdict(Counter)
    for d in rows:
        by_mod_track[d["Dev Module Code"]][d["Track"]] += 1
    for code, name in DEV_MODS:
        c = by_mod_track[code]
        ws.cell(r, 2, code).font = FONT_BOLD
        ws.cell(r, 2).border = THIN
        ws.cell(r, 3, name).border = THIN
        ws.cell(r, 4, SHEET_FOR_DEV[code]).border = THIN
        ws.cell(r, 4).fill = FILLS["edit"]
        ws.cell(r, 5, sum(c.values())).alignment = CENTER
        ws.cell(r, 5).border = THIN
        for i, tr in enumerate(("Web", "UI", "Mobile", "QA"), 6):
            cell = ws.cell(r, i, c[tr] or 0)
            cell.alignment = CENTER
            cell.border = THIN
            if c[tr]:
                cell.fill = TRACK_FILL[tr]
                cell.font = FONT_WHITE
        r += 1
    set_widths(ws, [4, 14, 36, 28, 12, 10, 12, 10])
    ws.freeze_panes = "A5"
    return ws


def write_mains(wb, rows):
    ws = wb.create_sheet("07_Main_Tasks_by_Module")
    ws.sheet_properties.tabColor = "1F618D"
    headers = [
        "Seq", "Dev Module Code", "Dev Module", "Phase", "Main Task ID", "Main Task",
        "Work Type(s)", "Tracks", "Web", "UI", "Mobile", "QA", "Sub-tasks",
        "PDF Feature", "AS-IS", "TO-BE", "Overall Status (live)",
    ]
    ws.merge_cells("A1:Q1")
    ws["A1"] = (
        "MAIN TASKS — Overall Status is Done only when every sub-task of this Main Task is Done "
        "(all in-scope Frontend/Backend/Database/API/Mobile/Integration statuses Done on the module sheet)."
    )
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws.row_dimensions[1].height = 36
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["header"]
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    grouped, seen = [], {}
    for d in sort_module_phase(rows):
        mid = d["Main Task ID"]
        if mid not in seen:
            seen[mid] = {
                "code": d["Dev Module Code"], "mod": d["Dev Module"], "phase": d.get("Phase"),
                "mid": mid, "main": d.get("Main Task"), "wts": [], "tracks": set(), "n": 0,
                "feat": d.get("PDF Feature (exact wording)"),
                "as_is": d.get("Existing Capability (AS-IS)"),
                "to_be": d.get("Gap / Change Required (TO-BE)"),
            }
            grouped.append(seen[mid])
        g = seen[mid]
        g["n"] += 1
        g["tracks"].add(d.get("Track"))
        wt = d.get("Work Type")
        if wt and wt not in g["wts"]:
            g["wts"].append(wt)
    for i, g in enumerate(grouped, 1):
        rr = i + 2
        mid_col = tcol("Main Task ID")
        ov_col = tcol("Overall Status")
        status_f = (
            f'=IF(COUNTIF({mid_col},E{rr})=0,"",'
            f'IF(COUNTIFS({mid_col},E{rr},{ov_col},"Done")=COUNTIF({mid_col},E{rr}),"Done",'
            f'IF(COUNTIFS({mid_col},E{rr},{ov_col},"Blocked")>0,"Blocked",'
            f'IF(COUNTIFS({mid_col},E{rr},{ov_col},"In Progress")>0,"In Progress",'
            f'IF(COUNTIFS({mid_col},E{rr},{ov_col},"Not Started")=COUNTIF({mid_col},E{rr}),"Not Started","In Progress")))))'
        )
        tracks = [t for t in ("Web", "UI", "Mobile", "QA") if t in g["tracks"]]
        vals = [
            i, g["code"], g["mod"], g["phase"], g["mid"], g["main"],
            ", ".join(g["wts"]), ", ".join(tracks),
            "Yes" if "Web" in g["tracks"] else "—",
            "Yes" if "UI" in g["tracks"] else "—",
            "Yes" if "Mobile" in g["tracks"] else "—",
            "Yes" if "QA" in g["tracks"] else "—",
            g["n"], g["feat"], g["as_is"], g["to_be"], status_f,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(rr, c, v)
            cell.font = FONT_BOLD if c == 6 else FONT_N
            cell.alignment = WRAP
            cell.border = THIN
            if c in (9, 10, 11, 12) and v == "Yes":
                tr = {9: "Web", 10: "UI", 11: "Mobile", 12: "QA"}[c]
                cell.fill = TRACK_FILL[tr]
                cell.font = FONT_WHITE
                cell.alignment = CENTER
            elif i % 2 == 0 and c not in (9, 10, 11, 12):
                cell.fill = FILLS["alt"]
        ws.row_dimensions[rr].height = 38
    last = len(grouped) + 2
    apply_status_cf(ws, 17, 3, last)
    set_widths(ws, [8, 12, 32, 8, 14, 50, 26, 20, 8, 8, 10, 8, 10, 30, 34, 34, 18])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:Q{last}"
    return ws


def write_dashboard(wb):
    ws = wb.create_sheet("08_Dashboard")
    ws.sheet_properties.tabColor = "6C3483"
    ws.merge_cells("A1:L1")
    ws["A1"] = (
        "LIVE DASHBOARD — counts read T_MASTER, which reads the module sheets. "
        "Mark Frontend Status = Done on M05 and these numbers change. No refresh."
    )
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws.row_dimensions[1].height = 36
    ws["A3"] = "Integrity"
    ws["A3"].font = FONT_WHITE
    ws["A3"].fill = FILLS["header"]
    ws.merge_cells("A3:B3")
    ws["A4"] = "Master rows"
    ws["B4"] = f"=COUNTA({tcol('Sub Task ID')})"
    ws["A5"] = "Must equal"
    ws["B5"] = 1000
    ws["A6"] = "Check"
    ws["B6"] = '=IF(B4=B5,"OK — nothing skipped","ERROR")'
    for r in range(4, 7):
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2).border = THIN

    ws["A8"] = "Overall Status"
    ws["A8"].font = FONT_WHITE
    ws["A8"].fill = FILLS["header"]
    ws.merge_cells("A8:B8")
    ws["A9"] = "Status"
    ws["B9"] = "Count"
    ws["A9"].fill = FILLS["header"]
    ws["B9"].fill = FILLS["header"]
    ws["A9"].font = FONT_WHITE
    ws["B9"].font = FONT_WHITE
    for i, st in enumerate(["Not Started", "In Progress", "Blocked", "Done", "Deferred", "N/A"]):
        ws.cell(10 + i, 1, st).border = THIN
        ws.cell(10 + i, 2, f'=COUNTIF({tcol("Overall Status")},A{10+i})').border = THIN
        ws.cell(10 + i, 2).alignment = CENTER
    pie = PieChart()
    pie.title = "Overall Status (live)"
    labels = Reference(ws, min_col=1, min_row=10, max_row=15)
    data = Reference(ws, min_col=2, min_row=9, max_row=15)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.width = 12
    pie.height = 7
    ws.add_chart(pie, "D3")

    # Layer remaining vs done
    ws["A18"] = "Function remaining vs complete (Scope = Yes only)"
    ws["A18"].font = FONT_WHITE
    ws["A18"].fill = FILLS["header"]
    ws.merge_cells("A18:E18")
    ws["A19"] = "Function"
    ws["B19"] = "In scope (Yes)"
    ws["C19"] = "Done"
    ws["D19"] = "Remaining"
    ws["E19"] = "% Done"
    for c in range(1, 6):
        ws.cell(19, c).fill = FILLS["header"]
        ws.cell(19, c).font = FONT_WHITE
        ws.cell(19, c).border = THIN
    for i, layer in enumerate(LAYERS):
        r = 20 + i
        ws.cell(r, 1, layer).border = THIN
        ws.cell(r, 1).font = FONT_BOLD
        ws.cell(r, 2, f'=COUNTIF({tcol(layer)},"Yes")').border = THIN
        ws.cell(r, 3, f'=COUNTIFS({tcol(layer)},"Yes",{tcol(layer + " Status")},"Done")').border = THIN
        ws.cell(r, 4, f"=B{r}-C{r}").border = THIN
        ws.cell(r, 5, f'=IF(B{r}=0,"",C{r}/B{r})').border = THIN
        ws.cell(r, 5).number_format = "0%"
        ws.cell(r, 4).fill = FILLS["gold"]
        for c in range(2, 5):
            ws.cell(r, c).alignment = CENTER

    ws["A27"] = "Track × Overall Status"
    ws["A27"].font = FONT_WHITE
    ws["A27"].fill = FILLS["header"]
    ws.merge_cells("A27:G27")
    ws["A28"] = "Track"
    for i, st in enumerate(["Not Started", "In Progress", "Blocked", "Done", "Deferred"], 2):
        ws.cell(28, i, st).fill = FILLS["header"]
        ws.cell(28, i).font = FONT_WHITE
        ws.cell(28, i).border = THIN
    ws.cell(28, 7, "Total").fill = FILLS["header"]
    ws.cell(28, 7).font = FONT_WHITE
    ws.cell(28, 1).fill = FILLS["header"]
    ws.cell(28, 1).font = FONT_WHITE
    for ri, tr in enumerate(("Web", "UI", "Mobile", "QA"), 29):
        ws.cell(ri, 1, tr).fill = TRACK_FILL[tr]
        ws.cell(ri, 1).font = FONT_WHITE
        ws.cell(ri, 1).border = THIN
        for ci, st in enumerate(["Not Started", "In Progress", "Blocked", "Done", "Deferred"], 2):
            cl = get_column_letter(ci)
            ws.cell(ri, ci, f'=COUNTIFS({tcol("Track")},$A{ri},{tcol("Overall Status")},{cl}$28)').border = THIN
            ws.cell(ri, ci).alignment = CENTER
        ws.cell(ri, 7, f"=SUM(B{ri}:F{ri})").border = THIN

    ws["A34"] = "Module remaining (sum of Layers Remaining on Master)"
    ws["A34"].font = FONT_WHITE
    ws["A34"].fill = FILLS["header"]
    ws.merge_cells("A34:D34")
    ws["A35"] = "Code"
    ws["B35"] = "Dev Module"
    ws["C35"] = "Sub-tasks"
    ws["D35"] = "Layer remaining"
    for c in range(1, 5):
        ws.cell(35, c).fill = FILLS["header"]
        ws.cell(35, c).font = FONT_WHITE
        ws.cell(35, c).border = THIN
    for i, (code, name) in enumerate(DEV_MODS):
        r = 36 + i
        ws.cell(r, 1, code).border = THIN
        ws.cell(r, 2, name).border = THIN
        ws.cell(r, 3, f'=COUNTIF({tcol("Dev Module Code")},A{r})').border = THIN
        ws.cell(r, 4, f'=SUMIF({tcol("Dev Module Code")},A{r},{tcol("Layers Remaining")})').border = THIN
        ws.cell(r, 4).fill = FILLS["gold"]
        ws.cell(r, 3).alignment = CENTER
        ws.cell(r, 4).alignment = CENTER
    set_widths(ws, [22, 36, 16, 16, 14, 14, 12])
    return ws


def add_master_table(ws, last):
    tab = Table(displayName="T_MASTER", ref=f"A2:{get_column_letter(NCOLS)}{last}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=False)
    ws.add_table(tab)


def main():
    rows = load_rows()
    if len({d["Original Seq"] for d in rows}) != 1000:
        raise SystemExit("Original Seq not unique")

    wb = Workbook()
    write_cover(wb, rows)
    write_employees(wb)
    write_sheets_setup(wb)
    write_modmap(wb)

    by_mod = defaultdict(list)
    for d in rows:
        by_mod[d["Dev Module Code"]].append(d)

    # Module sheets FIRST so names exist for INDIRECT
    for code, name in DEV_MODS:
        subset = by_mod.get(code, [])
        if not subset:
            continue
        n = Counter(d["Track"] for d in subset)
        who = " · ".join(f"{tr} {n[tr]}" for tr in ("Web", "UI", "Mobile", "QA") if n[tr])
        write_module_sheet(
            wb,
            SHEET_FOR_DEV[code],
            FILLS["modhdr"],
            f"{code}  {name}  —  {len(subset)} sub-tasks ({who}).  "
            "YELLOW = edit here (Assigned To, Frontend/Backend/Database/API/Mobile/Integration Status + Assigned).  "
            "Scope Yes/No is fixed. Status N/A = that function is not in this sub-task.  "
            "Master, Web/UI/Mobile/QA sheets and Dashboard follow these cells live.",
            subset,
        )

    write_mirror_sheet(
        wb, "01_MASTER", "1B3A4B", FILLS["header"],
        "MASTER MIRROR — do not type Status here. It is live from M00–M19. "
        "Click Open Module Sheet to assign an employee or mark Frontend/API/… Done. Filter to find work.",
        rows, lambda rs: sorted(rs, key=lambda d: d["Original Seq"]),
        use_table=True,
    )
    # Table on master for dashboard
    ws_m = wb["01_MASTER"]
    add_master_table(ws_m, 1000 + 2)

    write_mirror_sheet(
        wb, "02_By_Module_Phase", "1A5276", FILLS["web"],
        "ALL 1000 ROWS by module then phase. Live from module sheets. Filter Dev Module Code. "
        "To change Status or Assigned, use Open Module Sheet (do not type over formulas).",
        rows, sort_module_phase,
    )
    write_mirror_sheet(
        wb, "03_Web", "1A5276", FILLS["web"],
        "WEB TRACK MIRROR — live from module sheets. Edit layer Status on the module sheet (Open Module Sheet).",
        [d for d in rows if d.get("Track") == "Web"],
        lambda rs: sorted(rs, key=lambda d: d["Original Seq"]),
    )
    write_mirror_sheet(
        wb, "04_UI", "B9770E", FILLS["ui"],
        "UI TRACK MIRROR — live from module sheets. Mark Frontend Status Done on the module sheet; this column updates.",
        [d for d in rows if d.get("Track") == "UI"],
        lambda rs: sorted(rs, key=lambda d: d["Original Seq"]),
    )
    write_mirror_sheet(
        wb, "05_Mobile", "117A65", FILLS["mobile"],
        "MOBILE TRACK MIRROR — live from module sheets. Mark Mobile Status on the module sheet.",
        [d for d in rows if d.get("Track") == "Mobile"],
        lambda rs: sorted(rs, key=lambda d: d["Original Seq"]),
    )
    write_mirror_sheet(
        wb, "06_QA", "6C3483", FILLS["qa"],
        "QA TRACK MIRROR — live from module sheets. QA rows still have Frontend/API Status where those layers are in scope.",
        [d for d in rows if d.get("Track") == "QA"],
        lambda rs: sorted(rs, key=lambda d: d["Original Seq"]),
    )

    write_mains(wb, rows)
    write_dashboard(wb)

    order = [
        "00_Cover",
        "00_Employees",
        "00_Sheets_Dropdown",
        "01_MASTER",
        "02_By_Module_Phase",
        "03_Web",
        "04_UI",
        "05_Mobile",
        "06_QA",
        "07_Main_Tasks_by_Module",
        "08_Dashboard",
    ] + [SHEET_FOR_DEV[c] for c, _ in DEV_MODS if c in by_mod]
    # ModMap stays hidden at end
    for i, name in enumerate(order):
        if name not in wb.sheetnames:
            continue
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)

    wb.save(OUT)
    print(f"Wrote {len(rows)} rows  cols={NCOLS}")
    for code, name in DEV_MODS:
        c = Counter(d["Track"] for d in by_mod.get(code, []))
        print(f"  {code} {sum(c.values()):4}  {name}")
    print("→", OUT)


if __name__ == "__main__":
    main()
