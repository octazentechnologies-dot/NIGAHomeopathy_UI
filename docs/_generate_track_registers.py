#!/usr/bin/env python3
"""
Split Homeocentrum_Feature_Delivery_Task_Register.xlsx into four sequential
development backlogs: Web, UI, Mobile, QA.

Source of truth: the master sequential register (every original row is assigned
to exactly one track — nothing is skipped, nothing is duplicated as work).

Rules (first match wins):
  1. QA     — Layer = QA, or Work Type = Regression, or sub-task is a QA step
  2. Mobile — Layer is Mobile UI / Mobile client / Mobile UX / Mobile
              or surface is Patient/Doctor Mobile App (non-API)
  3. UI     — Layer is Web SPA or Design (React portal screens)
  4. Web    — everything else (DB, API, backend, security, integrations,
              architecture, client decisions, configuration)

Develop each sheet top-to-bottom by Track Seq. Original Seq is kept so teams
can see global order. UI/Mobile/QA of a Main Task wait on that Main Task's
Web (API/DB) rows when Sister Tracks says Web.
"""

from collections import Counter, defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "/Users/OctazenWork/NIGA Project/NigaHomeopathy-UI/docs/Homeocentrum_Feature_Delivery_Task_Register.xlsx"
OUT = "/Users/OctazenWork/NIGA Project/NigaHomeopathy-UI/docs/Homeocentrum_Task_Register_Web_UI_Mobile_QA.xlsx"

THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILLS = {
    "header": PatternFill("solid", fgColor="1B3A4B"),
    "web": PatternFill("solid", fgColor="1A5276"),
    "ui": PatternFill("solid", fgColor="B9770E"),
    "mobile": PatternFill("solid", fgColor="117A65"),
    "qa": PatternFill("solid", fgColor="6C3483"),
    "exist": PatternFill("solid", fgColor="D5F5E3"),
    "improve": PatternFill("solid", fgColor="FCF3CF"),
    "modify": PatternFill("solid", fgColor="FDEBD0"),
    "new": PatternFill("solid", fgColor="D6EAF8"),
    "newint": PatternFill("solid", fgColor="D5D8DC"),
    "cfg": PatternFill("solid", fgColor="E8DAEF"),
    "mig": PatternFill("solid", fgColor="FADBD8"),
    "client": PatternFill("solid", fgColor="F5B7B1"),
    "reg": PatternFill("solid", fgColor="D5F5E3"),
    "alt": PatternFill("solid", fgColor="F8F9F9"),
    "white": PatternFill("solid", fgColor="FFFFFF"),
    "gold": PatternFill("solid", fgColor="FEF9E7"),
    "pre": PatternFill("solid", fgColor="2C3E50"),
}
for i, hexcol in enumerate(
    ["7B241C", "1A5276", "B9770E", "117A65", "6C3483", "1F618D", "0E6655",
     "922B21", "4A235A", "1B4F72", "145A32", "6E2C00", "1A5276", "4A235A",
     "0E6251", "7B241C", "1F618D", "117A65", "6C3483", "34495E"]
):
    FILLS[f"phase{i}"] = PatternFill("solid", fgColor=hexcol)

FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_WHITE_LG = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_WHITE_MD = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="1B3A4B", size=18)
FONT_H = Font(name="Calibri", bold=True, color="1B3A4B", size=13)
FONT_B = Font(name="Calibri", bold=True, color="1B3A4B", size=11)
FONT_N = Font(name="Calibri", size=10, color="1C2833")
FONT_S = Font(name="Calibri", size=9, color="5D6D7E")
FONT_BOLD = Font(name="Calibri", bold=True, size=10, color="1C2833")

TYPE_FILL = {
    "Existing": FILLS["exist"],
    "Existing Improvement": FILLS["improve"],
    "Existing Modification": FILLS["modify"],
    "New": FILLS["new"],
    "New Integration": FILLS["newint"],
    "Configuration": FILLS["cfg"],
    "Data Migration": FILLS["mig"],
    "Client Decision": FILLS["client"],
    "Regression": FILLS["reg"],
}
TRACK_FILL = {
    "Web": FILLS["web"],
    "UI": FILLS["ui"],
    "Mobile": FILLS["mobile"],
    "QA": FILLS["qa"],
}
PHASE_FILL = {str(i): FILLS[f"phase{i}"] for i in range(20)}
PHASE_FILL["PRE"] = FILLS["pre"]

STATUS_VALUES = "Not Started,In Progress,Blocked,Done,Deferred"
WT_VALUES = "Existing,Existing Improvement,Existing Modification,New,New Integration,Configuration,Data Migration,Client Decision,Regression"

TRACK_HEADERS = [
    "Track Seq",
    "Original Seq",
    "Track",
    "Phase",
    "Phase Name",
    "WBS",
    "Main Task ID",
    "Main Task",
    "Sub Task ID",
    "Sub Task",
    "Work Type",
    "Module / Area",
    "PDF Section",
    "PDF Feature (exact wording)",
    "PDF Delivery Text",
    "Application Surface",
    "Layer",
    "Frontend",
    "Backend",
    "Database",
    "API",
    "Mobile",
    "Integration",
    "Priority",
    "Effort Band",
    "Depends On (original)",
    "Depends On (this track)",
    "Sister Tracks (same Main Task)",
    "Wait for other track",
    "Existing Capability (AS-IS)",
    "Gap / Change Required (TO-BE)",
    "Acceptance Criteria",
    "Status",
    "Owner",
    "Notes",
]

MOBILE_LAYERS = {"Mobile UI", "Mobile client", "Mobile UX", "Mobile"}
UI_LAYERS = {"Web SPA", "Design"}
MOBILE_SURFACES = {"Patient Mobile App", "Doctor Mobile App"}


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def classify(d):
    layer = (d.get("Layer") or "").strip()
    surface = (d.get("Application Surface") or "").strip()
    wt = d.get("Work Type") or ""
    sub = (d.get("Sub Task") or "").strip()
    # 1 QA
    if layer == "QA" or wt == "Regression" or sub.startswith("QA:") or sub.lower().startswith("qa "):
        return "QA"
    # 2 Mobile native work only (not Web SPA that also ships later on mobile)
    if layer in MOBILE_LAYERS:
        return "Mobile"
    if surface in MOBILE_SURFACES and layer not in {
        "API", "Database", "Architecture", "API contract", "Backend", "Web SPA", "Design"
    }:
        return "Mobile"
    # 3 UI — all React / web screens
    if layer in UI_LAYERS:
        return "UI"
    fe, be, db, mob = d.get("Frontend"), d.get("Backend"), d.get("Database"), d.get("Mobile")
    if (
        fe == "Yes"
        and be != "Yes"
        and db != "Yes"
        and mob != "Yes"
        and layer not in {
            "API", "Database", "Backend", "Integration", "Architecture",
            "Decision", "Configuration", "Security", "API hygiene",
        }
    ):
        return "UI"
    # 4 Web — platform, DB, API, security, vendors, decisions
    return "Web"


def load_master():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    ws = wb["03_Sequential_Register"]
    headers = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        d = dict(zip(headers, row))
        if not d.get("Sub Task ID"):
            continue
        d["Original Seq"] = d.get("Seq")
        d["Track"] = classify(d)
        rows.append(d)
    wb.close()
    return rows


def sister_map(rows):
    m = defaultdict(set)
    for d in rows:
        m[d["Main Task ID"]].add(d["Track"])
    return {k: sorted(v, key=lambda t: ["Web", "UI", "Mobile", "QA"].index(t)) for k, v in m.items()}


def remap_depends(track_rows):
    """Map original Depends On IDs that live in this track to Track Seq."""
    id_to_tseq = {}
    for d in track_rows:
        id_to_tseq[d["Sub Task ID"]] = d["Track Seq"]
    for d in track_rows:
        raw = (d.get("Depends On") or "").strip()
        if not raw or raw in ("—", "-", "–"):
            d["Depends On (this track)"] = "—"
            d["Wait for other track"] = ""
            continue
        parts = [p.strip() for p in raw.replace(";", ",").split(",") if p.strip() and p.strip() not in ("—", "-")]
        same, other = [], []
        for p in parts:
            # take leading SUB-ID token (e.g. "APT-05.02" or "APT-05.02, DOC-01")
            token = p.split()[0] if p else p
            token = token.rstrip(".")
            if token in id_to_tseq:
                same.append(f"T-{id_to_tseq[token]} ({token})")
            else:
                other.append(p)
        d["Depends On (this track)"] = "; ".join(same) if same else "—"
        d["Wait for other track"] = "; ".join(other) if other else ""


def build_track(rows, sisters, track):
    selected = [d for d in rows if d["Track"] == track]
    # keep original Seq order
    selected.sort(key=lambda d: int(d["Original Seq"]))
    out = []
    for i, d in enumerate(selected, 1):
        nd = dict(d)
        nd["Track Seq"] = i
        sis = [t for t in sisters.get(d["Main Task ID"], []) if t != track]
        nd["Sister Tracks (same Main Task)"] = ", ".join(sis) if sis else "—"
        out.append(nd)
    remap_depends(out)
    return out


def write_header_banner(ws, title, fill, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws["A1"] = title
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = fill
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 36
    for i, h in enumerate(TRACK_HEADERS, 1):
        cell = ws.cell(2, i, h)
        cell.fill = fill
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN


def write_track_sheet(wb, name, color, banner, track_rows, tab_hex):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_hex
    ncols = len(TRACK_HEADERS)
    write_header_banner(ws, banner, color, ncols)
    wrap_cols = {5, 8, 10, 14, 15, 26, 27, 28, 29, 30, 31, 32, 35}
    center_cols = {1, 2, 3, 4, 18, 19, 20, 21, 22, 23, 24, 25, 33}
    for i, d in enumerate(track_rows, 1):
        vals = [
            d["Track Seq"],
            d["Original Seq"],
            d["Track"],
            d.get("Phase"),
            d.get("Phase Name"),
            d.get("WBS"),
            d.get("Main Task ID"),
            d.get("Main Task"),
            d.get("Sub Task ID"),
            d.get("Sub Task"),
            d.get("Work Type"),
            d.get("Module / Area"),
            d.get("PDF Section"),
            d.get("PDF Feature (exact wording)"),
            d.get("PDF Delivery Text"),
            d.get("Application Surface"),
            d.get("Layer"),
            d.get("Frontend"),
            d.get("Backend"),
            d.get("Database"),
            d.get("API"),
            d.get("Mobile"),
            d.get("Integration"),
            d.get("Priority"),
            d.get("Effort Band"),
            d.get("Depends On"),
            d.get("Depends On (this track)"),
            d.get("Sister Tracks (same Main Task)"),
            d.get("Wait for other track"),
            d.get("Existing Capability (AS-IS)"),
            d.get("Gap / Change Required (TO-BE)"),
            d.get("Acceptance Criteria"),
            "Not Started",
            d["Track"],
            d.get("Notes") or "",
        ]
        r = i + 2
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = FONT_N
            cell.alignment = WRAP if c in wrap_cols else (CENTER if c in center_cols else WRAP)
            cell.border = THIN
            if c == 3:
                cell.fill = TRACK_FILL[d["Track"]]
                cell.font = FONT_WHITE
            elif c == 4:
                cell.fill = PHASE_FILL.get(str(d.get("Phase")), FILLS["white"])
                cell.font = FONT_WHITE
            elif c == 11:
                cell.fill = TYPE_FILL.get(d.get("Work Type"), FILLS["white"])
            elif i % 2 == 0 and c not in (3, 4, 11):
                cell.fill = FILLS["alt"]
        ws.row_dimensions[r].height = 46
    last = len(track_rows) + 2
    dv_s = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    dv_w = DataValidation(type="list", formula1=f'"{WT_VALUES}"', allow_blank=False)
    dv_p = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=False)
    dv_e = DataValidation(type="list", formula1='"S,M,L,XL"', allow_blank=False)
    ws.add_data_validation(dv_s)
    ws.add_data_validation(dv_w)
    ws.add_data_validation(dv_p)
    ws.add_data_validation(dv_e)
    dv_s.add(f"AG3:AG{last}")
    dv_w.add(f"K3:K{last}")
    dv_p.add(f"X3:X{last}")
    dv_e.add(f"Y3:Y{last}")
    set_widths(
        ws,
        [10, 12, 10, 8, 40, 10, 14, 52, 16, 58, 22, 20, 14, 34, 44, 26, 16,
         10, 10, 10, 10, 10, 12, 10, 10, 22, 28, 22, 28, 40, 40, 40, 14, 12, 24],
    )
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ncols)}{last}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.sheet_view.zoomScale = 90
    return ws


def write_main_sheet(wb, name, color, tab_hex, track, track_rows):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_hex
    headers = [
        "Seq", "Phase", "Phase Name", "WBS", "Main Task ID", "Main Task",
        "Work Type(s)", "Module", "PDF Section", "PDF Feature", "Surface",
        "Sub-task count", "First Track Seq", "Last Track Seq",
        "Sister Tracks", "Priority (highest)", "AS-IS", "TO-BE", "Status",
    ]
    ws.merge_cells("A1:S1")
    ws["A1"] = (
        f"{track} MAIN TASKS — execute sub-tasks in {track} Sequential sheet filtered by Main Task ID. "
        "Do not start a Main Task if Sister Tracks includes Web and those Web rows are not Done (API/DB first)."
    )
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = color
    ws.row_dimensions[1].height = 32
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = color
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    grouped = []
    seen = {}
    prio_rank = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
    for d in track_rows:
        mid = d["Main Task ID"]
        if mid not in seen:
            seen[mid] = {
                **d,
                "n": 0,
                "wts": [],
                "first": d["Track Seq"],
                "last": d["Track Seq"],
                "prios": [],
            }
            grouped.append(seen[mid])
        g = seen[mid]
        g["n"] += 1
        g["last"] = d["Track Seq"]
        g["wts"].append(d.get("Work Type"))
        g["prios"].append(d.get("Priority") or "P2")
    for i, g in enumerate(grouped, 1):
        wts = []
        for w in g["wts"]:
            if w and w not in wts:
                wts.append(w)
        best = min(g["prios"], key=lambda p: prio_rank.get(p, 9))
        vals = [
            i, g.get("Phase"), g.get("Phase Name"), g.get("WBS"), g.get("Main Task ID"),
            g.get("Main Task"), ", ".join(wts), g.get("Module / Area"), g.get("PDF Section"),
            g.get("PDF Feature (exact wording)"), g.get("Application Surface"), g["n"],
            g["first"], g["last"], g.get("Sister Tracks (same Main Task)"), best,
            g.get("Existing Capability (AS-IS)"), g.get("Gap / Change Required (TO-BE)"),
            "Not Started",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, c, v)
            cell.font = FONT_BOLD if c == 6 else FONT_N
            cell.alignment = WRAP
            cell.border = THIN
            if c == 2:
                cell.fill = PHASE_FILL.get(str(g.get("Phase")), FILLS["white"])
                cell.font = FONT_WHITE
            elif i % 2 == 0 and c != 2:
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i + 2].height = 42
    last = len(grouped) + 2
    dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"S3:S{last}")
    set_widths(ws, [8, 8, 40, 10, 14, 58, 28, 20, 14, 36, 26, 12, 12, 12, 22, 12, 42, 42, 14])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:S{last}"
    return ws


def write_cover(wb, counts, n_orig, n_mains):
    ws = wb.active
    ws.title = "00_Cover"
    ws.sheet_properties.tabColor = "1B3A4B"
    ws.merge_cells("B2:G2")
    ws["B2"] = "HOMEOCENTRUM — WEB / UI / MOBILE / QA TASK REGISTERS"
    ws["B2"].font = FONT_TITLE
    ws.merge_cells("B3:G3")
    ws["B3"] = (
        "Split from Homeocentrum_Feature_Delivery_Task_Register.xlsx  ·  "
        "Source PDF 29 August 2026  ·  Every original sub-task is in exactly one track"
    )
    ws["B3"].font = FONT_S
    ws.merge_cells("B5:G5")
    ws["B5"] = "How the four teams work (read this before coding)"
    ws["B5"].font = FONT_WHITE_MD
    ws["B5"].fill = FILLS["header"]
    rules = [
        "1. This file is the team split of the master sequential register. The master remains the programme source of truth. Do not skip a PDF feature — if it is not on your sheet, another track owns it (see Sister Tracks).",
        "2. Each original sub-task is assigned to EXACTLY one track so work is not double-counted. Sum of Web + UI + Mobile + QA = all 1,000 master sub-tasks.",
        "3. Develop strictly by Track Seq on your sequential sheet. Original Seq is the global programme order (Phase PRE → 0 → … → 19).",
        "4. WEB owns: database, APIs, backend, security (login/RBAC/OTP/audit), payment gateway, SMS/video/push server, Account ledger, client decisions. Start here for every Main Task that has a Web sister.",
        "5. UI owns: React web screens for Admin, Doctor, Reception, Account, Public website, Pharmacy console (Layer = Web SPA). Do not invent a second API — call Web APIs.",
        "6. MOBILE owns: Patient iOS/Android and Doctor iOS/Android screens (Mobile UI / client / UX). Case-taking stays on Doctor web. Consume the same .NET 8 APIs.",
        "7. QA owns: every Regression / QA layer row (acceptance, IDOR, parity, E2E). QA of a Main Task starts when that Main Task’s Web/UI/Mobile implementation rows are Done.",
        "8. Column “Wait for other track” lists original Depends On IDs that are NOT on your sheet (usually a Web API). Do not start until those IDs are Done on the other sheet.",
        "9. Column “Sister Tracks” lists other tracks that have sub-tasks under the same Main Task ID. Typical pattern: Web (DB+API) → UI (screen) → Mobile (app screen) → QA.",
        "10. Work Type is unchanged from the master: Existing = keep/regression only; Existing Improvement = finish partial; Existing Modification = change behaviour; New = build; New Integration = vendor.",
        "11. Status dropdown is the only column to edit daily. Owner defaults to the track name; assign people as needed.",
        "12. Sheet 09_Traceability maps Original Seq → Track. If a master row is missing there, that is a defect — stop and fix the generator before coding.",
    ]
    r = 6
    for line in rules:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(r, 2, line).alignment = WRAP
        ws.cell(r, 2).font = FONT_N
        ws.row_dimensions[r].height = 38
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, "Track counts (exclusive assignment)").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    for h, col in [("Track", 2), ("What you build", 3), ("Sub tasks", 4), ("Main tasks (with ≥1 sub)", 5), ("Sheet to execute", 6)]:
        cell = ws.cell(r, col, h)
        cell.font = FONT_WHITE
        cell.fill = FILLS["header"]
        cell.border = THIN
    r += 1
    blurbs = {
        "Web": "DB, API, backend, security, payments, integrations, PRE client gates",
        "UI": "React SPA screens — Admin, Doctor, Reception, Account, Public, Pharmacy",
        "Mobile": "Patient app + Doctor app (no case-taking on phone)",
        "QA": "Regression packs, acceptance, IDOR, iOS/Android parity, E2E ecosystem proof",
    }
    sheets = {
        "Web": "01_Web_Sequential",
        "UI": "02_UI_Sequential",
        "Mobile": "03_Mobile_Sequential",
        "QA": "04_QA_Sequential",
    }
    for tr in ("Web", "UI", "Mobile", "QA"):
        ws.cell(r, 2, tr).fill = TRACK_FILL[tr]
        ws.cell(r, 2).font = FONT_WHITE
        ws.cell(r, 2).border = THIN
        ws.cell(r, 3, blurbs[tr]).alignment = WRAP
        ws.cell(r, 3).border = THIN
        ws.cell(r, 4, counts[tr][0]).alignment = CENTER
        ws.cell(r, 4).border = THIN
        ws.cell(r, 5, counts[tr][1]).alignment = CENTER
        ws.cell(r, 5).border = THIN
        ws.cell(r, 6, sheets[tr]).border = THIN
        ws.row_dimensions[r].height = 32
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(
        r, 2,
        f"Integrity: original sub-tasks = {n_orig}  ·  assigned = {sum(c[0] for c in counts.values())}  ·  "
        f"unique Main Task IDs in master = {n_mains}  ·  none dropped.",
    ).font = FONT_H
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, "Work Type legend (same as master)").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    legends = [
        ("Existing", "Live today. Keep. Do not rebuild."),
        ("Existing Improvement", "Exists but stub, partial, or missing a role surface."),
        ("Existing Modification", "Exists but behaviour must change."),
        ("New", "Not in the three repos."),
        ("New Integration", "Vendor/gateway (Razorpay, video SDK, SMS, FCM)."),
        ("Configuration", "Rates/flags/templates — not hardcoded."),
        ("Data Migration", "Additive schema / backfill."),
        ("Client Decision", "Blocks engineering until signed (Web PRE rows)."),
        ("Regression", "QA track — prove it still works."),
    ]
    for wt, txt in legends:
        ws.cell(r, 2, wt).fill = TYPE_FILL.get(wt, FILLS["white"])
        ws.cell(r, 2).font = FONT_BOLD
        ws.cell(r, 2).border = THIN
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(r, 3, txt).alignment = WRAP
        ws.cell(r, 3).border = THIN
        ws.row_dimensions[r].height = 24
        r += 1
    set_widths(ws, [4, 28, 62, 14, 22, 24, 20])
    ws.freeze_panes = "B5"
    return ws


def write_trace(wb, rows):
    ws = wb.create_sheet("09_Traceability")
    ws.sheet_properties.tabColor = "2C3E50"
    headers = [
        "Original Seq", "Track Seq", "Track", "Phase", "Main Task ID", "Sub Task ID",
        "Sub Task", "Work Type", "Layer", "Surface", "Sister Tracks",
    ]
    ws.merge_cells("A1:K1")
    ws["A1"] = (
        "LINE-BY-LINE TRACE — every master Seq 1…N appears once. Filter Track to audit a team. "
        "Count of rows must equal the master sequential register."
    )
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws.row_dimensions[1].height = 32
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["header"]
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    # need track seq: build lookup
    tseq = {}
    by_track = defaultdict(int)
    ordered = sorted(rows, key=lambda d: (["Web", "UI", "Mobile", "QA"].index(d["Track"]), int(d["Original Seq"])))
    counters = Counter()
    # actual track seq was assigned in build_track; recompute consistently
    assigned = defaultdict(list)
    for d in rows:
        assigned[d["Track"]].append(d)
    for tr in ("Web", "UI", "Mobile", "QA"):
        assigned[tr].sort(key=lambda d: int(d["Original Seq"]))
        for i, d in enumerate(assigned[tr], 1):
            tseq[(d["Track"], d["Sub Task ID"])] = i
    for i, d in enumerate(sorted(rows, key=lambda x: int(x["Original Seq"])), 1):
        tr = d["Track"]
        vals = [
            d["Original Seq"],
            tseq[(tr, d["Sub Task ID"])],
            tr,
            d.get("Phase"),
            d.get("Main Task ID"),
            d.get("Sub Task ID"),
            d.get("Sub Task"),
            d.get("Work Type"),
            d.get("Layer"),
            d.get("Application Surface"),
            ", ".join(t for t in ["Web", "UI", "Mobile", "QA"] if t in {x["Track"] for x in rows if x["Main Task ID"] == d["Main Task ID"]}),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, c, v)
            cell.font = FONT_N
            cell.alignment = WRAP if c == 7 else CENTER if c in (1, 2, 3, 4) else WRAP
            cell.border = THIN
            if c == 3:
                cell.fill = TRACK_FILL[tr]
                cell.font = FONT_WHITE
            elif c == 8:
                cell.fill = TYPE_FILL.get(d.get("Work Type"), FILLS["white"])
            elif i % 2 == 0 and c not in (3, 8):
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i + 2].height = 32
    last = len(rows) + 2
    set_widths(ws, [12, 12, 10, 8, 14, 16, 62, 22, 16, 28, 22])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{last}"
    return ws


def write_dashboard(wb, tracks):
    ws = wb.create_sheet("10_Dashboard")
    ws.sheet_properties.tabColor = "6C3483"
    ws.merge_cells("A1:F1")
    ws["A1"] = "PLANNING COUNTS — exclusive track assignment from the master register"
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws["A3"] = "Sub-tasks by track"
    ws["A3"].font = FONT_WHITE
    ws["A3"].fill = FILLS["header"]
    ws.merge_cells("A3:B3")
    ws["A4"] = "Track"
    ws["B4"] = "Sub tasks"
    ws["A4"].fill = FILLS["header"]
    ws["B4"].fill = FILLS["header"]
    ws["A4"].font = FONT_WHITE
    ws["B4"].font = FONT_WHITE
    r = 5
    for tr in ("Web", "UI", "Mobile", "QA"):
        ws.cell(r, 1, tr).fill = TRACK_FILL[tr]
        ws.cell(r, 1).font = FONT_WHITE
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2, len(tracks[tr])).alignment = CENTER
        ws.cell(r, 2).border = THIN
        r += 1
    pie = PieChart()
    pie.title = "Sub-tasks by track"
    labels = Reference(ws, min_col=1, min_row=5, max_row=8)
    data = Reference(ws, min_col=2, min_row=4, max_row=8)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = True
    pie.width = 12
    pie.height = 8
    ws.add_chart(pie, "D3")

    # work type per track
    ws["A11"] = "Work Type × Track"
    ws["A11"].font = FONT_WHITE
    ws["A11"].fill = FILLS["header"]
    ws.merge_cells("A11:F11")
    wtypes = [
        "Existing", "Existing Improvement", "Existing Modification", "New",
        "New Integration", "Configuration", "Data Migration", "Client Decision", "Regression",
    ]
    ws["A12"] = "Work Type"
    ws["A12"].font = FONT_WHITE
    ws["A12"].fill = FILLS["header"]
    for i, tr in enumerate(("Web", "UI", "Mobile", "QA"), 2):
        cell = ws.cell(12, i, tr)
        cell.font = FONT_WHITE
        cell.fill = TRACK_FILL[tr]
        cell.border = THIN
    ws.cell(12, 6, "Total").font = FONT_WHITE
    ws.cell(12, 6).fill = FILLS["header"]
    r = 13
    for wt in wtypes:
        ws.cell(r, 1, wt).fill = TYPE_FILL.get(wt, FILLS["white"])
        ws.cell(r, 1).border = THIN
        total = 0
        for i, tr in enumerate(("Web", "UI", "Mobile", "QA"), 2):
            n = sum(1 for d in tracks[tr] if d.get("Work Type") == wt)
            total += n
            ws.cell(r, i, n).alignment = CENTER
            ws.cell(r, i).border = THIN
        ws.cell(r, 6, total).alignment = CENTER
        ws.cell(r, 6).border = THIN
        r += 1

    # phase × track
    r += 1
    ws.cell(r, 1, "Phase × Track (sub-tasks)").font = FONT_WHITE
    ws.cell(r, 1).fill = FILLS["header"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 1, "Phase").font = FONT_WHITE
    ws.cell(r, 1).fill = FILLS["header"]
    for i, tr in enumerate(("Web", "UI", "Mobile", "QA"), 2):
        cell = ws.cell(r, i, tr)
        cell.font = FONT_WHITE
        cell.fill = TRACK_FILL[tr]
        cell.border = THIN
    ws.cell(r, 6, "Total").font = FONT_WHITE
    ws.cell(r, 6).fill = FILLS["header"]
    r += 1
    phases = []
    for tr in ("Web", "UI", "Mobile", "QA"):
        for d in tracks[tr]:
            ph = str(d.get("Phase"))
            if ph not in phases:
                phases.append(ph)
    for ph in phases:
        ws.cell(r, 1, ph).fill = PHASE_FILL.get(ph, FILLS["white"])
        ws.cell(r, 1).font = FONT_WHITE
        ws.cell(r, 1).border = THIN
        total = 0
        for i, tr in enumerate(("Web", "UI", "Mobile", "QA"), 2):
            n = sum(1 for d in tracks[tr] if str(d.get("Phase")) == ph)
            total += n
            ws.cell(r, i, n).alignment = CENTER
            ws.cell(r, i).border = THIN
        ws.cell(r, 6, total).alignment = CENTER
        ws.cell(r, 6).border = THIN
        r += 1
    set_widths(ws, [28, 14, 14, 14, 14, 14])
    return ws


def main():
    rows = load_master()
    n_orig = len(rows)
    n_mains = len({d["Main Task ID"] for d in rows})
    sisters = sister_map(rows)
    tracks = {tr: build_track(rows, sisters, tr) for tr in ("Web", "UI", "Mobile", "QA")}
    assigned = sum(len(v) for v in tracks.values())
    if assigned != n_orig:
        raise SystemExit(f"Integrity fail: assigned {assigned} != original {n_orig}")

    wb = Workbook()
    counts = {tr: (len(tracks[tr]), len({d["Main Task ID"] for d in tracks[tr]})) for tr in tracks}
    write_cover(wb, counts, n_orig, n_mains)
    banners = {
        "Web": (
            "WEB SEQUENTIAL BACKLOG — DB, API, backend, security, payments, integrations, PRE decisions. "
            "Execute top to bottom by Track Seq. Do not skip. UI/Mobile/QA of the same Main Task wait on your Done status."
        ),
        "UI": (
            "UI SEQUENTIAL BACKLOG — React SPA screens (Admin, Doctor, Reception, Account, Public, Pharmacy). "
            "Execute by Track Seq. If Wait for other track lists a Web API ID, that API must be Done first."
        ),
        "Mobile": (
            "MOBILE SEQUENTIAL BACKLOG — Patient iOS/Android then Doctor iOS/Android. No case-taking on phone. "
            "Execute by Track Seq. Consume .NET 8 APIs already built on the Web track."
        ),
        "QA": (
            "QA SEQUENTIAL BACKLOG — every Regression / acceptance / IDOR / parity / E2E row from the master. "
            "Execute by Track Seq. Start a Main Task’s QA when its Web/UI/Mobile implementation rows are Done."
        ),
    }
    colors = {"Web": FILLS["web"], "UI": FILLS["ui"], "Mobile": FILLS["mobile"], "QA": FILLS["qa"]}
    tabs = {"Web": "1A5276", "UI": "B9770E", "Mobile": "117A65", "QA": "6C3483"}
    names = {
        "Web": "01_Web_Sequential",
        "UI": "02_UI_Sequential",
        "Mobile": "03_Mobile_Sequential",
        "QA": "04_QA_Sequential",
    }
    main_names = {
        "Web": "05_Web_Main_Tasks",
        "UI": "06_UI_Main_Tasks",
        "Mobile": "07_Mobile_Main_Tasks",
        "QA": "08_QA_Main_Tasks",
    }
    for tr in ("Web", "UI", "Mobile", "QA"):
        write_track_sheet(wb, names[tr], colors[tr], banners[tr], tracks[tr], tabs[tr])
    for tr in ("Web", "UI", "Mobile", "QA"):
        write_main_sheet(wb, main_names[tr], colors[tr], tabs[tr], tr, tracks[tr])
    write_trace(wb, rows)
    write_dashboard(wb, tracks)
    order = [
        "00_Cover",
        "01_Web_Sequential",
        "02_UI_Sequential",
        "03_Mobile_Sequential",
        "04_QA_Sequential",
        "05_Web_Main_Tasks",
        "06_UI_Main_Tasks",
        "07_Mobile_Main_Tasks",
        "08_QA_Main_Tasks",
        "09_Traceability",
        "10_Dashboard",
    ]
    for i, name in enumerate(order):
        current = wb.sheetnames.index(name)
        if current != i:
            wb.move_sheet(name, offset=i - current)
    wb.save(OUT)
    print(f"Original {n_orig} sub-tasks / {n_mains} main tasks")
    for tr in ("Web", "UI", "Mobile", "QA"):
        print(f"  {tr:6} {counts[tr][0]:4} sub  {counts[tr][1]:4} main")
    print("→", OUT)


if __name__ == "__main__":
    main()
