#!/usr/bin/env python3
"""Generate NIGA Homeocentrum sequential development task register (Excel)."""

from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from copy import copy

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILLS = {
    "header": PatternFill("solid", fgColor="1B3A4B"),
    "cover_title": PatternFill("solid", fgColor="1B3A4B"),
    "section": PatternFill("solid", fgColor="1B3A4B"),
    "phase0": PatternFill("solid", fgColor="7B241C"),
    "phase1": PatternFill("solid", fgColor="1A5276"),
    "phase2": PatternFill("solid", fgColor="B9770E"),
    "phase3": PatternFill("solid", fgColor="117A65"),
    "phase4": PatternFill("solid", fgColor="6C3483"),
    "phase5": PatternFill("solid", fgColor="1F618D"),
    "phase6": PatternFill("solid", fgColor="0E6655"),
    "phase7": PatternFill("solid", fgColor="922B21"),
    "phase8": PatternFill("solid", fgColor="4A235A"),
    "pre": PatternFill("solid", fgColor="2C3E50"),
    "cross": PatternFill("solid", fgColor="34495E"),
    "reg": PatternFill("solid", fgColor="1E8449"),
    "exist": PatternFill("solid", fgColor="D5F5E3"),
    "improve": PatternFill("solid", fgColor="FCF3CF"),
    "modify": PatternFill("solid", fgColor="FDEBD0"),
    "new": PatternFill("solid", fgColor="D6EAF8"),
    "newint": PatternFill("solid", fgColor="D5D8DC"),
    "cfg": PatternFill("solid", fgColor="E8DAEF"),
    "mig": PatternFill("solid", fgColor="FADBD8"),
    "client": PatternFill("solid", fgColor="F5B7B1"),
    "p0": PatternFill("solid", fgColor="F5B7B1"),
    "p1": PatternFill("solid", fgColor="FAD7A0"),
    "p2": PatternFill("solid", fgColor="D6EAF8"),
    "p3": PatternFill("solid", fgColor="E8F8F5"),
    "alt": PatternFill("solid", fgColor="F8F9F9"),
    "white": PatternFill("solid", fgColor="FFFFFF"),
    "light": PatternFill("solid", fgColor="F4F6F7"),
    "gold": PatternFill("solid", fgColor="FEF9E7"),
    "redsoft": PatternFill("solid", fgColor="FDEDEC"),
}

FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_WHITE_LG = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_WHITE_MD = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="1B3A4B", size=18)
FONT_H = Font(name="Calibri", bold=True, color="1B3A4B", size=13)
FONT_B = Font(name="Calibri", bold=True, color="1B3A4B", size=11)
FONT_N = Font(name="Calibri", size=10, color="1C2833")
FONT_S = Font(name="Calibri", size=9, color="5D6D7E")
FONT_BOLD = Font(name="Calibri", bold=True, size=10, color="1C2833")

TYPE_FILL = {
    "Existing": FILLS["exist"],
    "Existing Improvement": FILLS["improve"],
    "Existing Modification": FILLS["modify"],
    "New": FILLS["new"],
    "New Integration": FILLS["newint"],
    "Configuration": FILLS["cfg"],
    "Data Migration": FILLS["mig"],
    "Client Decision": FILLS["client"],
    "Regression": FILLS["exist"],
}

PHASE_FILL = {
    "PRE": FILLS["pre"],
    "0": FILLS["phase0"],
    "1": FILLS["phase1"],
    "2": FILLS["phase2"],
    "3": FILLS["phase3"],
    "4": FILLS["phase4"],
    "5": FILLS["phase5"],
    "6": FILLS["phase6"],
    "7": FILLS["phase7"],
    "8": FILLS["phase8"],
    "QA": FILLS["cross"],
    "SEC": FILLS["phase0"],
    "NFR": FILLS["cross"],
    "INT": FILLS["cross"],
    "REG": FILLS["reg"],
}

HEADERS = [
    "Seq",
    "Phase",
    "Phase Name",
    "WBS",
    "Main Task ID",
    "Main Task",
    "Sub Task ID",
    "Sub Task",
    "Work Type",
    "Gap Code",
    "Module / Area",
    "Requirement IDs",
    "Application Surface",
    "Frontend",
    "Backend",
    "Database",
    "API",
    "Integration",
    "Priority",
    "Effort Band",
    "Depends On",
    "Source Section",
    "Existing Capability (AS-IS)",
    "Gap / Change Required (TO-BE)",
    "Acceptance Criteria",
    "Status",
    "Owner",
    "Notes",
]

# Work Type values used throughout:
# Existing | Existing Improvement | Existing Modification | New | New Integration
# Configuration | Data Migration | Client Decision | Regression

tasks = []  # list of dicts


def T(
    phase,
    phase_name,
    wbs_main,
    main_id,
    main,
    sub_id,
    sub,
    work_type,
    gap,
    module,
    req,
    surface,
    fe,
    be,
    db,
    api,
    integ,
    prio,
    effort,
    depends,
    source,
    existing,
    gap_text,
    accept,
    notes="",
):
    tasks.append(
        {
            "phase": phase,
            "phase_name": phase_name,
            "wbs_main": wbs_main,
            "main_id": main_id,
            "main": main,
            "sub_id": sub_id,
            "sub": sub,
            "work_type": work_type,
            "gap": gap,
            "module": module,
            "req": req,
            "surface": surface,
            "fe": fe,
            "be": be,
            "db": db,
            "api": api,
            "integ": integ,
            "prio": prio,
            "effort": effort,
            "depends": depends,
            "source": source,
            "existing": existing,
            "gap_text": gap_text,
            "accept": accept,
            "notes": notes,
        }
    )


Y, N, P = "Yes", "No", "Partial"

# =============================================================================
# PRE-PROGRAMME — Client gates, vendors, legal (must complete before / in parallel
# with engineering). Sequential because Phase 2 cannot start without Critical OQs.
# =============================================================================

# --- PRE.1 Critical open questions ---
T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.01",
  "OQ-A1: Decide settlement model — Razorpay Route / linked accounts vs Homeocentrum-held funds + periodic NEFT payouts (v1 recommendation: platform-held + NEFT)",
  "Client Decision", "CLR", "Finance / Payments", "OQ-A1, SC-02, FIN-02, FIN-03",
  "Account / Gateway", N, N, N, N, Y, "P0", "S", "—",
  "§25, §31.1, §21.1",
  "No settlement model exists; only S1 subscription via client-side Razorpay",
  "Written decision on marketplace split vs platform-held + NEFT; this determines payout architecture",
  "Signed decision recorded; payout architecture unblocked",
  "Ecosystem plan recommends platform-held + NEFT for v1, Route for v2")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.02",
  "OQ-A2: Decide whether cash collected at reception is retained by the clinic or remitted to Homeocentrum",
  "Client Decision", "CLR", "Finance / Reception", "OQ-A2, REC-03, FIN-10, SC-03",
  "Account / Reception", N, N, N, N, N, "P0", "S", "PRE-01.01",
  "§25, §11.3, §21.1",
  "No reception collection or S3 ledger exists",
  "S3 ledger semantics and doctor settlement arithmetic depend on retain vs remit",
  "Written policy; S3 ledger design can proceed", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.03",
  "OQ-A3: Confirm commission percentages, GST treatment (on platform fee, consult, or both), invoice numbering series, and settlement hold period (T+N)",
  "Client Decision", "CLR", "Finance", "OQ-A3, FIN-01, FIN-08, FIN-02",
  "Account", N, N, N, N, N, "P0", "S", "PRE-01.01",
  "§25, §21.1, §12.2",
  "No commission/GST configuration; no invoice entity",
  "Cannot compute or report money without these values; must be configuration not hardcoded",
  "Finance function provides rates, GST treatment, invoice series, hold days", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.04",
  "OQ-A4: Decide unpaid-booking slot policy — hold slot (duration?) vs release immediately if patient does not pay",
  "Client Decision", "CLR", "Booking", "OQ-A4, GST-01, GST-02",
  "Public booking", N, N, N, N, N, "P0", "S", "—",
  "§25, §12.5",
  "Staff-only appointments; no public hold semantics",
  "Determines booking concurrency and no-show economics",
  "Hold-and-release rule documented and implemented in PublicBooking/Create", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.05",
  "OQ-A5: Confirm live-doctor credentialing backfill — auto-Approve existing doctors? Can unverified doctors still serve own clinic patients while invisible to public directory?",
  "Client Decision", "CLR", "Credentialing", "OQ-A5, ADM-01, GST-07, A-7",
  "Admin / Doctor", N, N, Y, N, N, "P0", "S", "—",
  "§25, §12.1, §16.5, §31.8",
  "RegisterDoctor activates immediately; no verification",
  "Wrong backfill locks out paying customers or exposes unverified doctors publicly",
  "Written backfill rule; migration script follows it", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.06",
  "OQ-A6: Define refund policy for cancellations — by whom, notice period, percentage",
  "Client Decision", "CLR", "Finance / Appointments", "OQ-A6, DOC-02, FIN-07",
  "Account / Doctor / Patient", N, N, N, N, N, "P0", "S", "PRE-01.01",
  "§25, §11.2, §21.1",
  "No refund path anywhere",
  "Cancel flow ships incomplete without a money path",
  "Refund matrix signed; FIN-07 can encode it", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.07",
  "OQ-A7: Select video vendor (Twilio / Daily / Agora / self-hosted WebRTC) and whether recording is required at launch",
  "Client Decision", "CLR", "Telemedicine", "OQ-A7, DOC-11, F-10",
  "Telemedicine / Mobile", N, N, N, N, Y, "P1", "S", "—",
  "§25, §2.5, §21.2, §31.4",
  "E-CONSULT status + WhatsApp video stub only",
  "Largest engineering variable in Phase 5; both mobile consult flows depend on it",
  "Vendor contracted; interface can be abstracted even before SDK wiring",
  "Build availability/queue/consent vendor-independently")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.08",
  "OQ-A8: Confirm whether a teleconsultation must be paid before the patient can join",
  "Client Decision", "CLR", "Telemedicine / Payments", "OQ-A8, DOC-10, DOC-14",
  "Telemedicine", N, N, N, N, N, "P1", "S", "PRE-01.01",
  "§25, §12.10",
  "No payment gate on E-CONSULT",
  "Changes queue and join gating",
  "Join-gating rule documented", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.1", "PRE-01",
  "Answer Critical open questions that block Phase 2 (money spine)",
  "PRE-01.09",
  "OQ-A9: Select SMS provider and confirm who owns DLT sender-ID and template registration (India)",
  "Client Decision", "CLR", "Outreach / OTP", "OQ-A9, DOC-05, GST-01",
  "SMS", N, N, N, N, Y, "P0", "S", "—",
  "§25, §2.5, §21.2, §31.4",
  "WhatsApp only; no SMS/DLT",
  "Regulatory prerequisite for OTP, booking confirmations, registration SMS",
  "Provider selected; DLT registration started (long lead time)", "")

# --- PRE.2 High clarifications ---
T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.01",
  "OQ-B1: Confirm code-then-name rule — names hidden until licensed pharmacy OTP-accepts; does it also apply to eRx PDF/print?",
  "Client Decision", "CLR", "eRx / HomeoMeds", "OQ-B1, PHR-10, DOC-08, F-08",
  "Patient / Pharmacy / Doctor", N, N, N, N, N, "P1", "S", "—",
  "§25, §11.5, §31.3",
  "Free-text prescription; names always visible to doctor; patient has no eRx view",
  "Most consequential interpretation; shapes eRx, patient app and HomeoMeds together",
  "Written confirmation of disclosure rule for app AND print", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.02",
  "OQ-B2: Confirm whether frequency and duration are mandatory alongside potency, or optional",
  "Client Decision", "CLR", "Prescription", "OQ-B2, DOC-07",
  "Patient Board / Pharmacy", N, N, N, N, N, "P1", "S", "—",
  "§25, §12.7",
  "Dose is empty free-text",
  "Affects validation and pharmacy dispensing clarity",
  "Validation rules signed", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.03",
  "OQ-B3: Provide exact Center of Gravity algorithm with a worked example from a practising homeopath (weighted centroid / miasmatic-kingdom / hierarchical section weighting / other)",
  "Client Decision", "CLR", "Patient Board", "OQ-B3, DOC-06, F-09",
  "Doctor web", N, N, N, N, N, "P2", "S", "—",
  "§25, §12.9, §31.2",
  "Repertorize + elimination exist; no COG",
  "Feature is unbuildable and unestimable until algorithm is signed",
  "Client-signed algorithm + worked example",
  "Sequence COG last within Phase 4; ship potency and notes split first")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.04",
  "OQ-B4: Define doctor directory ranking rules and what the ranking-explanation screen discloses to patients",
  "Client Decision", "CLR", "Directory / Trust", "OQ-B4, PAT-M2, ADM-01",
  "Patient web/app", N, N, N, N, N, "P2", "S", "PRE-01.05",
  "§25, §12.1, §21.1",
  "No public directory",
  "Trust feature and potential regulatory sensitivity",
  "Ranking rules documented for honest explanation UI", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.05",
  "OQ-B5: Confirm which roles may see remedy names before pharmacy acceptance (treating doctor only, or also reception and admin)",
  "Client Decision", "CLR", "eRx ACL", "OQ-B5, DOC-08, PHR-10",
  "API / all roles", N, N, N, N, N, "P1", "S", "PRE-02.01",
  "§25",
  "No code-then-name disclosure",
  "Access control on eRx payload",
  "Role matrix for eRx name disclosure signed", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.2", "PRE-02",
  "Answer High-priority feature clarifications",
  "PRE-02.06",
  "OQ-B6: Decide whether unused roles Management / Supervisor / Inspector need real screens and permissions, or should be retired",
  "Client Decision", "CLR", "RBAC", "OQ-B6, PLT-05, PLT-09",
  "Platform", Y, Y, Y, N, N, "P1", "S", "—",
  "§25, §4.4",
  "Present in role enum only; not specially routed",
  "Affects the RBAC matrix",
  "Retain-and-build or retire decision recorded", "")

# --- PRE.3 Medium / Low ---
for sid, q, notes in [
    ("PRE-03.01", "OQ-C1: Define availability, RPO/RTO targets, and retention periods for audio, recordings, documents and financial records", "NFR / DR"),
    ("PRE-03.02", "OQ-C2: Define telemedicine recording policy — enabled at launch? who may access? retention?", "Telemedicine legal"),
    ("PRE-03.03", "OQ-C3: Confirm what Hello Homeo Doc module 7 was (list runs 1–6, 8, 9, 10)", "Patient mobile scope"),
    ("PRE-03.04", "OQ-C4: Confirm whether patient WEB must be multilingual, or language selection is mobile-only", "i18n"),
    ("PRE-03.05", "OQ-C5: Define CliniSight progress (module 6) precisely and which data feeds it", "Phase 8"),
    ("PRE-03.06", "OQ-C6: Confirm launch geography constraint (city/state) for booking, telemedicine and HomeoMeds", "Launch scope"),
    ("PRE-03.07", "OQ-C7: Decide full patient web portal (/me/*) at app parity vs web limited to booking and payment (Assumption A-6)", "Patient web scope"),
    ("PRE-03.08", "OQ-C8: Who operates the pharmacy console at launch — pharmacy staff or Homeocentrum ops on their behalf?", "HomeoMeds ops"),
    ("PRE-03.09", "OQ-D1: Deliver Deep Analytics placeholder in this programme or formally defer", "Analytics"),
    ("PRE-03.10", "OQ-D2: Remove unused Velzon demo surface from repository, or only block it", "Platform hygiene"),
    ("PRE-03.11", "OQ-D3: Is a patient loyalty / wallet / package concept anticipated (cheaper to accommodate in ledger now)", "Ledger design"),
]:
    T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.3", "PRE-03",
      "Answer Medium and Low open questions (scope, NFR, deferred items)",
      sid, q, "Client Decision", "CLR", "Programme", sid.replace("PRE-03.0", "OQ-").replace("PRE-03.", "OQ-"),
      "Cross-cutting", N, N, N, N, N, "P2", "S", "—",
      "§25", "Not specified in provided documentation",
      "Written answer so scope does not silently expand",
      "Decision logged against the OQ", notes)

# --- PRE.4 Vendors & commercial setup ---
T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.01",
  "Open Homeocentrum-owned payment gateway account with KYC and marketplace/settlement configuration",
  "Client Decision", "DEP", "Payments", "SC-02, F-02, §2.5",
  "Gateway", N, N, N, N, Y, "P0", "M", "PRE-01.01",
  "§2.5, §21.2, §31.5",
  "Razorpay used only for doctor SaaS; keys in appsettings; no webhook",
  "Merchant-of-record account with webhook capability in every environment",
  "Sandbox + production accounts; webhook URL reachable; keys in secret store", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.02",
  "Contract SMS provider and begin DLT sender-ID + template registration for OTP, booking, registration, doctor-unavailable, reschedule, cancel, payment receipt",
  "New Integration", "NEW-INT", "Outreach", "DOC-05, OQ-A9",
  "SMS", N, Y, Y, Y, Y, "P0", "M", "PRE-01.09",
  "§17, §12.11, §31.4",
  "WhatsApp Cloud only",
  "SMS is a parallel channel; DLT is a regulatory prerequisite in India",
  "Templates registered; sandbox whitelist of test numbers", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.03",
  "Select and contract video vendor; obtain test project for token issuance",
  "New Integration", "NEW-INT", "Telemedicine", "DOC-11, OQ-A7",
  "Video", N, Y, N, Y, Y, "P2", "M", "PRE-01.07",
  "§17, §31.4",
  "WhatsApp stub",
  "Telemedicine unbuildable until vendor chosen; abstract behind interface",
  "Sandbox tokens work; vendor security review and DPA started", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.04",
  "Start legal review of privacy, terms, telemedicine practice guidelines, DPDP consent, e-pharmacy rules, recording policy — do not wait for Phase 3",
  "Client Decision", "DEP", "Legal", "GST-05, GST-06, OQ-C2, R-8",
  "Public web", Y, N, N, N, N, "P1", "M", "—",
  "§21.1, §31.6, §21.4",
  "Privacy/terms pages exist at 50% without DPDP/payments/tele/pharmacy",
  "Legal sign-off blocks patient go-live regardless of engineering",
  "Counsel review started in Phase 0; signed pages required before patient launch", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.05",
  "Decide mobile delivery approach (native vs React Native / cross-platform) and secure Apple + Google store accounts, signing, crash monitoring",
  "Client Decision", "CLR", "Mobile", "PAT-M*, DOCM-*, R-12, A-3",
  "Patient app / Doctor app", Y, Y, N, N, N, "P2", "S", "—",
  "§2.5, §31.9, §7.7",
  "No mobile code in any current repository",
  "Determines team shape for Phases 6–7; interim = responsive patient web",
  "Approach chosen; store accounts created", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.06",
  "Select object-storage provider for documents, audio, eRx PDFs and licences (signed URLs) — AD-9",
  "New Integration", "NEW-INT", "Platform", "AD-9, PLT-08, ADM-01",
  "Storage", N, Y, N, N, Y, "P0", "M", "—",
  "§10.4 AD-9, §17",
  "Static file hosting with directory browsing on /attachments and /Blogs",
  "Unsafe for PHI; signed URLs required before credentialing and patient uploads",
  "Provider selected; signed-URL issuance working in non-prod", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.07",
  "Approve Phase 0 as an immediate standalone workstream (security remediation is independent of commercial questions)",
  "Client Decision", "CLR", "Programme", "§31.7, R-2",
  "All", N, N, N, N, N, "P0", "S", "—",
  "§31.7, §33",
  "Closed clinic-tool security posture",
  "Must not wait for OQs; public perimeter cannot open until Phase 0 is done",
  "Sponsor approval to start Phase 0 immediately", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.08",
  "Validate pharmacy partner pipeline (licensed premises willing to onboard) before committing to Phase 7",
  "Client Decision", "DEP", "HomeoMeds", "R-14, PHR-01",
  "Pharmacy", N, N, N, N, N, "P3", "S", "—",
  "§2.5, §22 R-14",
  "No pharmacy entity or partners",
  "HomeoMeds requires licensed partners; supply risk",
  "At least a launch-partner pipeline confirmed before Phase 7 build", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.09",
  "Freeze API contracts in Swagger once Critical OQs 1–3 (settlement, cash, commission/GST) plus code-then-name and COG status are answered — convert PROPOSED endpoints to agreed contracts",
  "New", "NEW", "API design", "§15, §31.10",
  ".NET 8", N, Y, N, Y, N, "P0", "M", "PRE-01, PRE-02.01",
  "§31.10, §20.3",
  "All new endpoints are marked PROPOSED — not existing facts",
  "Unblocks estimation and parallel FE/BE work",
  "Swagger published for Phase 0–2 surfaces first", "")

T("PRE", "Pre-Programme — Client Gates & Decisions", "PRE.4", "PRE-04",
  "Vendor, commercial account and legal setup (long-lead items)",
  "PRE-04.10",
  "Produce phase-level estimate only AFTER steps PRE-01 through PRE-04.03 (vendors + policy) — do not invent timelines before that",
  "Client Decision", "CLR", "Programme", "R-17, §31.12, §26",
  "PMO", N, N, N, N, N, "P1", "S", "PRE-01, PRE-04.01, PRE-04.02, PRE-04.03",
  "§26, §31.12, §22 R-17",
  "Source documents provide no timelines",
  "An estimate before vendor and policy decisions would not survive contact with them",
  "Effort-band estimate issued after gates; no fake calendar", "")

# =============================================================================
# PHASE 0 — Platform hygiene (unblocks everything)
# =============================================================================

# P0.1 Password hashing
T("0", "Phase 0 — Platform hygiene", "0.1", "PLT-01",
  "Password hashing on UserMaster and credential migration (no lockout)",
  "PLT-01.01",
  "Implement hash+salt on write for UserMaster on NigaHomeopathy-API (.NET 8)",
  "Existing Modification", "E-MOD", "Auth", "PLT-01, S-1",
  ".NET 8 API", N, Y, Y, Y, N, "P0", "M", "—",
  "§7.1, §12.15, §18 S-1",
  "Plaintext compare on both APIs",
  "Hash + salt; never store or compare clear passwords",
  "New passwords stored hashed; compare uses hash", "")

T("0", "Phase 0 — Platform hygiene", "0.1", "PLT-01",
  "Password hashing on UserMaster and credential migration (no lockout)",
  "PLT-01.02",
  "Implement the same hash+salt verify path on classic NIGA_Latest_Code_API login (POST /Account/Login) until login is fully ported",
  "Existing Modification", "E-MOD", "Auth", "PLT-01, S-1, S-11",
  "Classic API", N, Y, Y, Y, N, "P0", "M", "PLT-01.01",
  "§14.1, §15.2, §18 S-11",
  "Classic API is the live login path (EOL .NET Core 2.2)",
  "Both APIs must not diverge during transition (R-6)",
  "Login on classic accepts hashed credentials with legacy fallback", "")

T("0", "Phase 0 — Platform hygiene", "0.1", "PLT-01",
  "Password hashing on UserMaster and credential migration (no lockout)",
  "PLT-01.03",
  "Legacy-compare fallback then re-hash on next successful login (MIG) so live users are not locked out",
  "Data Migration", "MIG", "Auth", "PLT-01, R-9, §16.5",
  "Both APIs / DB", N, Y, Y, Y, N, "P0", "M", "PLT-01.01",
  "§16.5, §12.15",
  "All existing UserMaster passwords are plaintext",
  "Hash on next login; force reset for dormant accounts after a cut-off",
  "No lockout of active users; dormant accounts forced through reset after cut-off",
  "Rehearse on a production copy")

T("0", "Phase 0 — Platform hygiene", "0.1", "PLT-01",
  "Password hashing on UserMaster and credential migration (no lockout)",
  "PLT-01.04",
  "Security tests: no plaintext credential exists or is logged; login still works for migrated and unmigrated users during fallback window",
  "Existing Modification", "E-MOD", "Auth / QA", "PLT-01, §20.1",
  "QA", N, Y, N, Y, N, "P0", "S", "PLT-01.03",
  "§20.1, §29",
  "Passwords comparable in clear",
  "Acceptance: no plaintext password is stored, compared or emailed anywhere",
  "Security + migration suite green", "")

# P0.2 Password reset
T("0", "Phase 0 — Platform hygiene", "0.2", "PLT-02",
  "Tokenised password reset — stop emailing plaintext passwords; replace fake/Firebase thunk",
  "PLT-02.01",
  "Create PasswordResetToken table (token hash, TTL, single-use, user FK)",
  "Existing Improvement", "P-ENH", "Auth", "PLT-02, S-2",
  "DB", N, Y, Y, N, N, "P0", "S", "PLT-01.01",
  "§12.15, §16.2",
  "ForgetPassword API emails the plaintext password; UI thunk uses fake/Firebase",
  "Reset token table with TTL and single use",
  "Token cannot be reused; expires", "")

T("0", "Phase 0 — Platform hygiene", "0.2", "PLT-02",
  "Tokenised password reset — stop emailing plaintext passwords; replace fake/Firebase thunk",
  "PLT-02.02",
  "Implement POST /api/users/ForgotPassword, /ResetPassword, /ChangePassword on .NET 8 — email a time-limited link, never the password",
  "Existing Improvement", "P-ENH", "Auth", "PLT-02, S-2",
  ".NET 8 API", Y, Y, Y, Y, Y, "P0", "M", "PLT-02.01",
  "§12.15, §15.2, §17 SMTP",
  "Classic ForgetPassword emails plaintext; SMTP exists",
  "Stop sending passwords; use tokenised links",
  "Link expires, single use, no password emailed", "")

T("0", "Phase 0 — Platform hygiene", "0.2", "PLT-02",
  "Tokenised password reset — stop emailing plaintext passwords; replace fake/Firebase thunk",
  "PLT-02.03",
  "Replace ForgetPassword.js + slices/auth/forgetpwd/thunk.js fake/Firebase path with real backend_helper calls",
  "Existing Improvement", "P-ENH", "Auth", "PLT-02",
  "Web SPA", Y, N, N, Y, N, "P0", "S", "PLT-02.02",
  "§5.3, §13.1",
  "/forgot-password UI exists; thunk is fake",
  "This is a security defect, not a 50% feature",
  "Fake/Firebase thunk gone; real APIs called", "")

T("0", "Phase 0 — Platform hygiene", "0.2", "PLT-02",
  "Tokenised password reset — stop emailing plaintext passwords; replace fake/Firebase thunk",
  "PLT-02.04",
  "Add reset-link status page (expired / already used / success) aligned with activation-link patterns",
  "Existing Improvement", "P-ENH", "Auth", "PLT-02, GST-08",
  "Web SPA", Y, Y, N, Y, N, "P1", "S", "PLT-02.02",
  "§7.4 GST-08",
  "Activation exists without expiry/resend UX completeness",
  "User-facing clarity for failed reset attempts",
  "Expired/used tokens show a resend path, not a crash", "")

# P0.3 Logout
T("0", "Phase 0 — Platform hygiene", "0.3", "PLT-03",
  "Server-side logout and session invalidation (covers REC-09 Clear session)",
  "PLT-03.01",
  "Activate UserLoginStatus and implement POST /api/Account/Logout with optional token denylist",
  "Existing Improvement", "P-ENH", "Auth", "PLT-03, REC-09, S-8",
  ".NET 8 API", N, Y, Y, Y, N, "P1", "S", "PLT-01.01",
  "§7.1, §12.15, §18 S-8",
  "Client clears sessionStorage only; UserLoginStatus unused",
  "Invalidate session server-side",
  "A token presented after logout is rejected", "")

T("0", "Phase 0 — Platform hygiene", "0.3", "PLT-03",
  "Server-side logout and session invalidation (covers REC-09 Clear session)",
  "PLT-03.02",
  "Wire header/logout in Admin, Doctor and Reception layouts to the Logout API (REC-09 50% → 100%)",
  "Existing Improvement", "P-ENH", "Auth", "PLT-03, REC-09",
  "Web SPA", Y, N, N, Y, N, "P1", "S", "PLT-03.01",
  "§7.6 REC-09",
  "Clear session is client-only",
  "Reception and doctor logout must hit the server",
  "All staff layouts call Logout; local storage cleared after 200", "")

# P0.4 Profiles
T("0", "Phase 0 — Platform hygiene", "0.4", "PLT-04",
  "Real profile management for Admin, Doctor and Reception (replace Velzon first_name screen)",
  "PLT-04.01",
  "Build GET/PUT /api/Profile/Me and POST /api/Profile/Photo (signed-URL storage) — no role escalation via this endpoint",
  "Existing Improvement", "P-ENH", "Auth", "PLT-04",
  ".NET 8 API", Y, Y, Y, Y, Y, "P1", "M", "PLT-08 / AD-9",
  "§12.15, §13.1",
  "Velzon first_name screen only",
  "Role-appropriate profile DTO; photo upload",
  "Each role edits own real fields; role cannot be changed here", "")

T("0", "Phase 0 — Platform hygiene", "0.4", "PLT-04",
  "Real profile management for Admin, Doctor and Reception (replace Velzon first_name screen)",
  "PLT-04.02",
  "Replace pages/Authentication/user-profile.js with role-appropriate forms (Admin: identity + photo + password)",
  "Existing Improvement", "P-ENH", "Auth", "PLT-04",
  "Web SPA — Admin", Y, N, N, Y, N, "P1", "S", "PLT-04.01",
  "§13.1",
  "Shared Velzon profile",
  "Admin edits real fields",
  "Admin profile saves and reloads", "")

T("0", "Phase 0 — Platform hygiene", "0.4", "PLT-04",
  "Real profile management for Admin, Doctor and Reception (replace Velzon first_name screen)",
  "PLT-04.03",
  "Doctor profile — clinic, fees, KYC/bank-PAN capture fields, photo, qualifications, hours, verification status (DOC-15 20% → 100%)",
  "Existing Improvement", "P-ENH", "Auth / Doctor", "PLT-04, DOC-15, FIN-09",
  "Web SPA — Doctor", Y, Y, Y, Y, N, "P1", "M", "PLT-04.01, ADM-01",
  "§7.5 DOC-15, §12.15",
  "Shared Velzon profile; registration has university/cert/qualification",
  "Complete doctor profile module used by directory and payouts",
  "Doctor sees verification status and can complete KYC without role change", "")

T("0", "Phase 0 — Platform hygiene", "0.4", "PLT-04",
  "Real profile management for Admin, Doctor and Reception (replace Velzon first_name screen)",
  "PLT-04.04",
  "Reception profile screen + API (REC-04 currently 0%)",
  "New", "NEW", "Auth / Reception", "PLT-04, REC-04",
  "Web SPA — Reception", Y, Y, Y, Y, N, "P2", "S", "PLT-04.01",
  "§7.6 REC-04, §12.15",
  "No reception profile at all",
  "Reception profile fields and screen",
  "Reception can edit own name, mobile, photo, password", "")

# P0.5 RBAC
T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.01",
  "Restore GetMenuByRole on .NET 8; seed RoleDetail/MenuMaster rows for existing + new menus",
  "Existing Modification", "E-MOD", "Platform", "PLT-05, PLT-09",
  ".NET 8 API / DB", Y, Y, Y, Y, N, "P0", "M", "PLT-09.01",
  "§4.4, §7.1, §12.15",
  "Menu permissions exist in DB but GetMenuByRole is commented out",
  "Menu built from GetMenuByRole",
  "Sidebar matches role; Account/Pharmacy get own menus", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.02",
  "Implement per-route frontend ACL in Routes/allRoutes.js and Routes/AuthProtected.js — UI hiding is never the only control",
  "Existing Modification", "E-MOD", "Platform", "PLT-05, S-3",
  "Web SPA", Y, N, N, N, N, "P0", "M", "PLT-05.01",
  "§4.4, §13.1, §18 S-3",
  "No per-route frontend ACL; any authenticated user who knows a URL can open any protected page",
  "Doctor typing an admin URL is refused by UI",
  "Cross-role URL access refused by UI", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.03",
  "Re-enable commented [Authorize] on repertory/admin controllers on BOTH APIs; apply [Authorize(Roles=)] to every new controller",
  "Existing Modification", "E-MOD", "Platform", "PLT-05, S-3",
  "Both APIs", N, Y, N, Y, N, "P0", "M", "PLT-05.01",
  "§4.4, §14.4, §18 S-3",
  "[Authorize] commented out on many controllers",
  "Public perimeter cannot be opened until closed",
  "API refuses cross-role calls with 403", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.04",
  "Enable JWT issuer and audience validation on .NET 8; add role claims for all roles; Recommended: separate token audiences (staff / patient / pharmacy)",
  "Existing Modification", "E-MOD", "Auth", "PLT-05, S-4, §14.4",
  ".NET 8 API", N, Y, N, Y, N, "P0", "M", "PLT-09.01",
  "§14.4, §18 S-4",
  "JWT issuer and audience are not validated",
  "Enable validation; separate audiences per client type",
  "Tokens with wrong audience rejected", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.05",
  "Implement ownership checks (not just role): doctor→own appointments; reception→JWT DoctorId; pharmacy→assigned orders; patient→own records",
  "Existing Modification", "E-MOD", "Platform", "PLT-05, §14.4",
  ".NET 8 API", N, Y, N, Y, N, "P0", "L", "PLT-05.03",
  "§14.4, §20.1 Permission matrix",
  "Plan gating is flag-based in login payload/UI, not API gateway",
  "IDOR protection required before PHI/payments",
  "Permission matrix suite: role × route × API", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.06",
  "Deprecate/remove classic POST api/Login/authenticate (second auth path, different signing key)",
  "Existing Modification", "E-MOD", "Auth", "S-9, §15.1",
  "Classic API", Y, Y, N, Y, N, "P1", "S", "PLT-05.04",
  "§15.1, §18 S-9",
  "Two auth paths with different signing keys",
  "Removal reduces attack surface",
  "Legacy path returns 410/404; no client still calls it", "")

T("0", "Phase 0 — Platform hygiene", "0.5", "PLT-05",
  "Restore RBAC: menu-by-role, per-route frontend ACL, [Authorize(Roles=)] on APIs, ownership checks",
  "PLT-05.07",
  "Restrict CORS from AllowAnyOrigin to known origins on both APIs",
  "Configuration", "CFG", "Platform", "S-10",
  "Both APIs", N, Y, N, N, N, "P1", "S", "—",
  "§18 S-10",
  "CORS AllowAnyOrigin on both APIs",
  "Restrict to known SPA and future mobile/web origins",
  "Unknown origins rejected", "")

# P0.6 Admin home
T("0", "Phase 0 — Platform hygiene", "0.6", "PLT-06",
  "Admin lands on /admin/dashboard with real KPI widgets (not Velzon ecommerce demo) — also covers ADM-05",
  "PLT-06.01",
  "Fix helpers/dashboard_helper.js getHomeDashboardPath so Admin goes to /admin/dashboard not /dashboard",
  "New", "NEW", "Admin dashboard", "PLT-06, ADM-05",
  "Web SPA", Y, N, N, N, N, "P1", "S", "—",
  "§5.3, §13.1",
  "Admin currently lands on Velzon ecommerce demo",
  "Redirect fix",
  "Admin login lands on real KPI home", "")

T("0", "Phase 0 — Platform hygiene", "0.6", "PLT-06",
  "Admin lands on /admin/dashboard with real KPI widgets (not Velzon ecommerce demo) — also covers ADM-05",
  "PLT-06.02",
  "Replace pages/Admin/Dashboard ecommerce widgets with real KPI cards; GET /api/AdminDashboard/Overview aggregate API",
  "New", "NEW", "Admin dashboard", "PLT-06, ADM-05, DOC-04",
  "Web SPA + .NET 8", Y, Y, N, Y, N, "P1", "M", "PLT-06.01",
  "§12.13, §13.1, §13.2",
  "Velzon ecommerce / NFT / crypto / CRM demo surface",
  "KPI widgets from real aggregates (extend as money/tele come online)",
  "Admin sees product KPIs; demo widgets gone from this route",
  "Full money KPIs complete after Phase 2")

# P0.7 Demo / fake auth
T("0", "Phase 0 — Platform hygiene", "0.7", "PLT-07",
  "Remove fake auth backend from production path; quarantine Velzon demo routes",
  "PLT-07.01",
  "Remove fake backend from App.js; set REACT_APP_DEFAULTAUTH off fake in production .env",
  "Configuration", "CFG", "Platform", "PLT-07, S-7, §4.10 #4",
  "Web SPA", Y, N, N, N, N, "P1", "S", "—",
  "§4.10, §13.1, §18 S-7",
  "Fake backend still activated; .env REACT_APP_DEFAULTAUTH=fake",
  "Confusing auth surface; demo code in production bundle",
  "Production path uses real APIs only", "")

T("0", "Phase 0 — Platform hygiene", "0.7", "PLT-07",
  "Remove fake auth backend from production path; quarantine Velzon demo routes",
  "PLT-07.02",
  "Block or remove Velzon demo routes (ecommerce, NFT, crypto, CRM, tickets, kanban, mailbox, apps-crypto-kyc) per OQ-D2 — do NOT build product on them",
  "Configuration", "CFG", "Platform", "PLT-07, OQ-D2, R-18, §26 traps",
  "Web SPA", Y, N, N, N, N, "P1", "S", "PLT-05.02, PRE-03.10",
  "§4.10 #5, §13.4, §26 traps",
  "Large unused Velzon demo surface still routed",
  "Must not be mistaken for Support, KYC or Admin KPI product",
  "Demo routes excluded from production or ACL-blocked", "")

T("0", "Phase 0 — Platform hygiene", "0.7", "PLT-07",
  "Remove fake auth backend from production path; quarantine Velzon demo routes",
  "PLT-07.03",
  "Fix start/build scripts that use Windows `set` syntax so heap bump applies on macOS/Linux CI agents",
  "Configuration", "CFG", "Platform", "§4.10 #6",
  "Web SPA", Y, N, N, N, N, "P2", "S", "—",
  "§4.10 #6",
  "Windows set syntax in npm scripts",
  "Cross-platform build reliability",
  "CI build on macOS/Linux applies NODE_OPTIONS heap bump", "")

# P0.8 Secrets & files
T("0", "Phase 0 — Platform hygiene", "0.8", "PLT-08",
  "Move secrets out of appsettings.json; disable directory browsing; signed URLs (AD-9)",
  "PLT-08.01",
  "Move gateway, OpenAI, WhatsApp, SMTP and JWT secrets to a secret store; rotate all keys previously committed in appsettings.json",
  "Configuration", "CFG", "Platform", "PLT-08, S-5",
  "Both APIs", N, Y, N, N, Y, "P0", "M", "—",
  "§4.10 #9, §18 S-5, §21.2",
  "Secrets present in appsettings.json",
  "Rotate; secret store before public exposure",
  "No secrets in source; production reads from store", "")

T("0", "Phase 0 — Platform hygiene", "0.8", "PLT-08",
  "Move secrets out of appsettings.json; disable directory browsing; signed URLs (AD-9)",
  "PLT-08.02",
  "Disable directory browsing on /attachments and /Blogs; serve patient documents, eRx PDFs and licences via time-limited signed URLs only",
  "Configuration", "CFG", "Platform", "PLT-08, S-6, AD-9",
  "Both APIs / storage", N, Y, N, Y, Y, "P0", "M", "PRE-04.06",
  "§4.10 #10, §18 S-6, §17",
  "Directory browsing enabled; static file hosting",
  "Information disclosure once patient documents are stored",
  "Guessable public URL cannot retrieve documents; access logged", "")

# P0.9 Roles
T("0", "Phase 0 — Platform hygiene", "0.9", "PLT-09",
  "Add roles Account, Patient, Pharmacy to RoleMaster and frontend constants; Account and Pharmacy layout shells",
  "PLT-09.01",
  "Seed RoleMaster rows Account, Patient, Pharmacy (and apply OQ-B6 decision for Management/Supervisor/Inspector)",
  "New", "DB", "Platform", "PLT-09, SC-04",
  "DB / API / SPA", Y, Y, Y, N, N, "P0", "S", "PRE-02.06",
  "§7.1, §4.4",
  "6-value role enum, 3 used (Admin, Doctor, Reception); Patient/Account/Pharmacy do not exist",
  "Seed roles + menu rows + FE constants in Components/constants/roles.js",
  "Login can issue tokens with new role claims", "")

T("0", "Phase 0 — Platform hygiene", "0.9", "PLT-09",
  "Add roles Account, Patient, Pharmacy to RoleMaster and frontend constants; Account and Pharmacy layout shells",
  "PLT-09.02",
  "Create AccountLayout and PharmacyLayout shells; extend LayoutMenuData.js for new menus (Credentialing, Payments, Support, Pharmacies, Account sidebar)",
  "New", "NEW", "Platform", "PLT-09, SC-04, §13.3",
  "Web SPA", Y, N, N, N, N, "P0", "M", "PLT-09.01, PLT-05.01",
  "§13.2, §13.3, §13.4",
  "Admin/Doctor/Reception layouts only",
  "New role shells; doctor layout continues to hide admin sidebar",
  "Account user cannot open repertory masters (even if URL guessed — depends on PLT-05)", "")

# P0.10 Dual-API discipline / freeze classic
T("0", "Phase 0 — Platform hygiene", "0.10", "PLT-10",
  "Architecture guardrails: freeze classic API; no third backend; all new domain on .NET 8 (AD-1, AD-2, AD-7)",
  "PLT-10.01",
  "Institute freeze: no new domain endpoints on NIGA_Latest_Code_API (ASP.NET Core 2.2 EOL); all new modules on NigaHomeopathy-API",
  "Existing Modification", "E-MOD", "Architecture", "AD-1, AD-2, R-3, R-6, §26 trap",
  "Classic API", N, Y, N, Y, N, "P0", "S", "—",
  "§2.4, §10.4, §26",
  "One UI against two backends; classic still on login, Rx save, notes, Razorpay",
  "Avoid third drift surface; classic frozen",
  "PR check / convention: new controllers only in .NET 8", "")

T("0", "Phase 0 — Platform hygiene", "0.10", "PLT-10",
  "Architecture guardrails: freeze classic API; no third backend; all new domain on .NET 8 (AD-1, AD-2, AD-7)",
  "PLT-10.02",
  "Plan and execute port of POST /Account/Login to .NET 8 (dated plan) — S-11",
  "Existing Modification", "E-MOD", "Auth", "S-11, AD-2, R-3",
  ".NET 8 API", Y, Y, Y, Y, N, "P0", "L", "PLT-01, PLT-05.04",
  "§14.4, §18 S-11, §15.2",
  "Login is on classic EOL stack",
  "EOL API on critical path under a public perimeter",
  "Staff login works entirely on .NET 8; classic login retired on a dated cutover", "")

T("0", "Phase 0 — Platform hygiene", "0.10", "PLT-10",
  "Architecture guardrails: freeze classic API; no third backend; all new domain on .NET 8 (AD-1, AD-2, AD-7)",
  "PLT-10.03",
  "Port Razorpay order generation off classic OrderController to .NET 8; remove hardcoded keys (Phase 0 seed of money spine — webhook + S1 ledger write continues in Phase 2)",
  "Existing Improvement", "P-ENH", "Payments", "DOC-19, AD-2, C-3, F-02",
  ".NET 8 API", Y, Y, Y, Y, Y, "P0", "M", "PLT-08.01, PRE-04.01",
  "§26 Phase 0, §15.2, C-3",
  "classic POST /Order/GenerateOrderId + client callback only",
  "Port to .NET 8; thin reuse of classic only as time-boxed interim with agreed removal date",
  "Order create on .NET 8; keys not in source",
  "Full webhook/ledger is Phase 2; this task is the port")

T("0", "Phase 0 — Platform hygiene", "0.10", "PLT-10",
  "Architecture guardrails: freeze classic API; no third backend; all new domain on .NET 8 (AD-1, AD-2, AD-7)",
  "PLT-10.04",
  "Add integration tests that exercise shared entities written by both APIs so dual-API drift is caught (R-6)",
  "Existing Modification", "E-MOD", "QA", "R-6, §16.6",
  "QA", N, Y, Y, Y, N, "P1", "M", "PLT-10.01",
  "§16.6, §22 R-6",
  "Two APIs, one DB, overlapping entity names",
  "Contract tests over shared entities",
  "CI fails if overlapping writes diverge", "")

T("0", "Phase 0 — Platform hygiene", "0.10", "PLT-10",
  "Architecture guardrails: freeze classic API; no third backend; all new domain on .NET 8 (AD-1, AD-2, AD-7)",
  "PLT-10.05",
  "Adopt one migration path for new tables (versioned scripts, not ad-hoc SQL only) — schema today is largely Database/Scripts/",
  "Existing Modification", "E-MOD", "Database", "§4.7, §16.6",
  "DB", N, N, Y, N, N, "P1", "S", "—",
  "§4.7, §16.6",
  "Schema evolution largely by SQL scripts, not exclusively EF migrations",
  "One migration path for new tables; version and review scripts",
  "New commercial/patient tables follow the agreed migration process", "")

# P0.11 Shared keys & OTP control plane foundation
T("0", "Phase 0 — Platform hygiene", "0.11", "PLT-11",
  "Ecosystem shared-key design and OTP control-plane foundation (SC-01, AD-6)",
  "PLT-11.01",
  "Document and enforce shared keys: DoctorId/UserId, PatientId, PatientAppId, CaseId, ErxId, LedgerTxnId, OrderId — reject any module that invents its own patient/appointment identifier at code review",
  "New", "NEW", "Architecture", "SC-01, §10.2",
  "All APIs", N, Y, Y, Y, N, "P0", "S", "—",
  "§10.2, §26 SC-01",
  "Clinic roles share some keys; no patient/ledger/eRx keys",
  "Interlinking is identifiers, not diagrams",
  "Design review checklist includes shared-key rule", "")

T("0", "Phase 0 — Platform hygiene", "0.11", "PLT-11",
  "Ecosystem shared-key design and OTP control-plane foundation (SC-01, AD-6)",
  "PLT-11.02",
  "Create OtpChallenge + OtpAuditLog tables and generic OTP service (rate limit, attempt lockout, short expiry, single use) reused by patient auth, payout, exception resolve, pharmacy accept, sensitive eRx",
  "New", "NEW", "Control plane", "AD-6, PHR-11, SC-03, §18.2",
  ".NET 8 / DB", N, Y, Y, Y, Y, "P0", "M", "PLT-09.01, PRE-04.02",
  "§12.2, §16.2, §18.2",
  "No OTP control plane (audio consent log is a pattern only)",
  "OTP-based monitoring so Homeocentrum can trace every high-risk action",
  "OTP challenges audited; brute force throttled",
  "Patient OTP SMS needs DLT; can stub in Phase 0 tests")

T("0", "Phase 0 — Platform hygiene", "0.11", "PLT-11",
  "Ecosystem shared-key design and OTP control-plane foundation (SC-01, AD-6)",
  "PLT-11.03",
  "Background job: OTP expiry cleanup (purge expired challenges; retain audit log)",
  "New", "NEW", "Control plane", "§14.3",
  ".NET 8 worker", N, Y, Y, N, N, "P1", "S", "PLT-11.02",
  "§14.3",
  "Existing hosted services: audio, embeddings, WhatsApp bulk, Excel import, retention",
  "Extend hosted-service pattern",
  "Expired OTPs purged; audit retained", "")

# P0.12 Logging, errors, public hardening foundation
T("0", "Phase 0 — Platform hygiene", "0.12", "PLT-12",
  "Cross-cutting API hygiene: correlation IDs, error mapping, rate-limit namespace for public routes (AD-8)",
  "PLT-12.01",
  "Extend BaseAPIController error mapping with correlation ids; never return gateway errors verbatim; structured logs with actor, role, entity reference",
  "Existing Improvement", "P-ENH", "Platform", "§11.9, §14.5, §19.2",
  ".NET 8 API", N, Y, N, Y, N, "P1", "M", "—",
  "§11.9, §14.5",
  "Existing BaseAPIController error mapping",
  "400/401/403/404/409/500 classes as specified in §11.9; money failures never silent",
  "Unhandled 500 returns correlation id; financial/OTP logs retained longer", "")

T("0", "Phase 0 — Platform hygiene", "0.12", "PLT-12",
  "Ecosystem shared-key design and OTP control-plane foundation (SC-01, AD-6)",
  "PLT-12.02",
  "Add rate limiting + recommended WAF on public namespaces /api/PublicBooking/* and /api/PatientAuth/* (AD-8) — implement before Phase 3 go-live",
  "New", "NEW", "Platform", "AD-8, R-15, GST-01",
  ".NET 8 API", N, Y, N, Y, N, "P1", "M", "PLT-12.01",
  "§10.4 AD-8, §18.2, §22 R-15",
  "No public patient endpoints",
  "Blast-radius containment; OTP/booking abuse protection",
  "Public endpoints throttled; cannot enumerate patient data", "")

T("0", "Phase 0 — Platform hygiene", "0.12", "PLT-12",
  "Cross-cutting API hygiene: correlation IDs, error mapping, rate-limit namespace for public routes (AD-8)",
  "PLT-12.03",
  "Idempotency keys required on webhooks, booking create and payment order create — shared middleware",
  "New", "NEW", "Platform", "§14.5, F-02, GST-01",
  ".NET 8 API", N, Y, Y, Y, N, "P0", "M", "—",
  "§14.5, §11.1",
  "No idempotency layer",
  "Duplicate webhook / retry must be no-op",
  "Replayed requests do not duplicate entities", "")

# P0.13 Regression suite + Patient Board spike
T("0", "Phase 0 — Platform hygiene", "0.13", "PLT-13",
  "Establish automated regression suite for all 100% modules (R-11) and schedule Patient Board decomposition spike",
  "PLT-13.01",
  "Create automated regression suite covering 25 admin master modules, Patient Board tabs, repertorization, audio pipeline, WhatsApp, 3D, import/export, board backup/restore — run every phase",
  "Regression", "E-NC", "QA", "ADM-10, DOC-33, R-11, §20.1",
  "QA", Y, Y, N, Y, N, "P0", "L", "—",
  "§20.1, §22 R-11, §29",
  "These modules are at 100% and must remain so",
  "Regression is the cost of not rebuilding them",
  "Suite green in CI; run at end of every phase", "")

T("0", "Phase 0 — Platform hygiene", "0.13", "PLT-13",
  "Establish automated regression suite for all 100% modules (R-11) and schedule Patient Board decomposition spike",
  "PLT-13.02",
  "Patient Board decomposition spike: extract Repertorize, Prescription, Notes, History sub-panels from PatientBoard.js (~13.5k LOC) BEFORE Phase 4 features land (R-5, §31.11)",
  "Existing Improvement", "P-ENH", "Patient Board", "R-5, §13.3, §2.4",
  "Web SPA", Y, N, N, N, N, "P1", "L", "PLT-13.01",
  "§13.3, §22 R-5, §31.11",
  "Four new clinical capabilities all land in one mega-component",
  "Extraction before, not after, COG/potency/notes/export",
  "Sub-components exist; no behaviour change; regression green",
  "Largest avoidable delivery risk")

T("0", "Phase 0 — Platform hygiene", "0.13", "PLT-13",
  "Establish automated regression suite for all 100% modules (R-11) and schedule Patient Board decomposition spike",
  "PLT-13.03",
  "Independent security review of Phase 0 findings before any public surface opens",
  "Existing Modification", "E-MOD", "Security", "R-2, §20.3",
  "Security", Y, Y, N, Y, N, "P0", "M", "PLT-01 through PLT-08",
  "§20.3, §18",
  "Eleven existing findings, three Critical",
  "Security review is an exit criterion for any phase that opens a public surface",
  "Review completed; Critical findings closed", "")

# Remaining phases, existing 100% catalogue, QA / NFR / integrations
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _task_rows_p1_p8 import add_all

add_all(T)

# =============================================================================
# Workbook
# =============================================================================
OUT = Path(__file__).resolve().parent / "NIGA_Homeocentrum_Development_Task_Register.xlsx"

wb = Workbook()

# Number tasks sequentially
for i, t in enumerate(tasks, 1):
    t["seq"] = i
    t["wbs"] = f"{t['wbs_main']}.{t['sub_id'].split('.')[-1]}" if "." in t["sub_id"] else t["sub_id"]


def apply_header(ws, headers, fill=FILLS["header"]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = FONT_WHITE
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = THIN
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    ws.sheet_view.showGridLines = False


def style_data_cell(cell, wrap=True, center=False):
    cell.font = FONT_N
    cell.alignment = CENTER if center else WRAP
    cell.border = THIN


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fill_for_type(wt):
    return TYPE_FILL.get(wt, FILLS["white"])


def fill_for_prio(p):
    return {"P0": FILLS["p0"], "P1": FILLS["p1"], "P2": FILLS["p2"], "P3": FILLS["p3"]}.get(p, FILLS["white"])


# ---------- Cover ----------
ws = wb.active
ws.title = "00_Cover_Legend"
ws.sheet_properties.tabColor = "1B3A4B"
ws.merge_cells("A1:G1")
ws["A1"] = "NIGA HOMEOCENTRUM — Development Task Register"
ws["A1"].font = FONT_WHITE_LG
ws["A1"].fill = FILLS["cover_title"]
ws["A1"].alignment = Alignment(vertical="center", horizontal="left")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:G2")
ws["A2"] = "Sequential WBS from NIGA_Homeocentrum_Solution_Analysis_Document (1).md  ·  Version 1.0  ·  29 August 2026  ·  Source analysis dated 28 August 2026"
ws["A2"].font = FONT_S
ws["A2"].fill = FILLS["light"]

cover_info = [
    ("Source document", "NigaHomeopathy-UI/docs/NIGA_Homeocentrum_Solution_Analysis_Document (1).md"),
    ("Programme", "Homeocentrum Single-Ecosystem Expansion (Clinic SaaS → Patient-Facing Care Marketplace)"),
    ("How to use this workbook", "Start on 02_Master_Task_Register. Filter by Phase, Work Type, Priority. Do not skip Seq order — dependencies are encoded left-to-right in Depends On."),
    ("Governing sequence", "PRE (client gates) → Phase 0 hygiene → 1 appointments → 2 money spine → 3 patient access → 4 clinical → 5 telemedicine → 6 mobile → 7 HomeoMeds → 8 continuity. Phase 2 is the commercial spine. Phase 0 is unconditional."),
    ("Accuracy rule", "No invented existing APIs. PROPOSED endpoints are design targets. Priorities are Recommended (client document states none). Effort bands S/M/L/XL from §27."),
]
r = 4
ws.merge_cells("A4:B4")
ws["A4"] = "DOCUMENT CONTROL"
ws["A4"].font = FONT_WHITE_MD
ws["A4"].fill = FILLS["section"]
ws.merge_cells("C4:G4")
ws["C4"].fill = FILLS["section"]
for label, val in cover_info:
    r += 1
    ws.cell(r, 1, label).font = FONT_B
    ws.cell(r, 1).fill = FILLS["light"]
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, val).font = FONT_N
    ws.cell(r, 2).alignment = WRAP
    ws.row_dimensions[r].height = 36

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "WORK TYPE (primary classification — use this column for planning)").font = FONT_WHITE_MD
ws.cell(r, 1).fill = FILLS["section"]
r += 1
type_legend = [
    ("Existing", "Confirmed present and complete in code (client 100%). Work is regression + ACL/hygiene only — do not rebuild.", FILLS["exist"]),
    ("Existing Improvement", "Partially existing (client N% < 100, or code partial). Enhancement / completion of a working asset.", FILLS["improve"]),
    ("Existing Modification", "Present but must change behaviour (e.g. hashing, RegisterDoctor must not auto-activate, [Authorize] re-enabled).", FILLS["modify"]),
    ("New", "Not built. Full specification, DB, API, UI, tests.", FILLS["new"]),
    ("New Integration", "New third-party (SMS, video, push, storage, payout rail) or inbound webhook.", FILLS["newint"]),
    ("Configuration", "Config/ops: secrets store, CORS, demo route quarantine, env flags — not a product feature.", FILLS["cfg"]),
    ("Data Migration", "Live-data change (password re-hash, note defaults, doctor verification backfill, S1 ledger backfill).", FILLS["mig"]),
    ("Client Decision", "Open question or vendor/legal/commercial gate. Engineering is blocked or shaped by the answer.", FILLS["client"]),
    ("Regression", "Protect a 100% module after other work. Same as Existing for planning; listed so nothing is skipped.", FILLS["exist"]),
]
ws.cell(r, 1, "Work Type").font = FONT_WHITE
ws.cell(r, 1).fill = FILLS["header"]
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
ws.cell(r, 2, "Meaning").font = FONT_WHITE
ws.cell(r, 2).fill = FILLS["header"]
for name, meaning, fill in type_legend:
    r += 1
    c = ws.cell(r, 1, name)
    c.font = FONT_BOLD
    c.fill = fill
    c.border = THIN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    c2 = ws.cell(r, 2, meaning)
    c2.font = FONT_N
    c2.alignment = WRAP
    c2.border = THIN
    c2.fill = fill
    ws.row_dimensions[r].height = 32

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "GAP CODE (from analysis §9.1)").font = FONT_WHITE_MD
ws.cell(r, 1).fill = FILLS["section"]
r += 1
ws.cell(r, 1, "Code").font = FONT_WHITE
ws.cell(r, 1).fill = FILLS["header"]
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
ws.cell(r, 2, "Meaning").font = FONT_WHITE
ws.cell(r, 2).fill = FILLS["header"]
for code, meaning in [
    ("E-NC", "Existing — No Change Required (regression only)"),
    ("E-MOD", "Existing — Modification Required"),
    ("P-ENH", "Partially Existing — Enhancement Required"),
    ("NEW", "New Development"),
    ("NEW-INT", "New Integration"),
    ("DB", "Database Change Required"),
    ("CFG", "Configuration Change"),
    ("MIG", "Data Migration Required"),
    ("DEP", "Requirement Dependency"),
    ("CLR", "Clarification Needed (open question)"),
]:
    r += 1
    ws.cell(r, 1, code).font = FONT_BOLD
    ws.cell(r, 1).border = THIN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, meaning).font = FONT_N
    ws.cell(r, 2).border = THIN

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "PRIORITY (Recommended — client document states none)  ·  EFFORT BANDS").font = FONT_WHITE_MD
ws.cell(r, 1).fill = FILLS["section"]
r += 1
for p, meaning, fill in [
    ("P0 Critical", "Security defects, or blockers without which money and public access cannot exist", FILLS["p0"]),
    ("P1 High", "Core client-requested capability with direct operational or revenue impact", FILLS["p1"]),
    ("P2 Medium", "Depends on P0/P1 foundations, or completion of partial modules", FILLS["p2"]),
    ("P3 Low", "Valuable but sequenced last (mobile doctor, HomeoMeds, continuity)", FILLS["p3"]),
    ("S — days", "Cancel formalisation, payment badge, potency picker, reception staff UI, export button, complaints wiring, admin redirect", FILLS["light"]),
    ("M — 1–2 sprints", "Reception collection, notes/eRx split, SMS events, credentialing, tickets MVP, profiles, RBAC, schedule screen, waiting queue", FILLS["light"]),
    ("L — multi-sprint", "Payments spine, Account console, public booking, clinic performance, COG, password migration on live users", FILLS["gold"]),
    ("XL", "Full telemedicine; patient mobile; doctor mobile; HomeoMeds marketplace", FILLS["redsoft"]),
]:
    r += 1
    ws.cell(r, 1, p).font = FONT_BOLD
    ws.cell(r, 1).fill = fill
    ws.cell(r, 1).border = THIN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, meaning).font = FONT_N
    ws.cell(r, 2).fill = fill
    ws.cell(r, 2).alignment = WRAP
    ws.cell(r, 2).border = THIN
    ws.row_dimensions[r].height = 28

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "SHEETS IN THIS WORKBOOK").font = FONT_WHITE_MD
ws.cell(r, 1).fill = FILLS["section"]
r += 1
sheets_help = [
    ("00_Cover_Legend", "This page — classification, sequence, how to read columns"),
    ("01_Phase_Summary", "Counts by phase and work type; dependency spine"),
    ("02_Master_Task_Register", "THE register — every sequential sub-task. Filter and assign Status/Owner."),
    ("03_Existing_vs_New", "Pivot-style counts: Existing / Improvement / New / etc."),
    ("04_Requirement_Traceability", "Requirement ID → tasks (PLT, ADM, FIN, GST, DOC, REC, PAT, DOCM, PHR, SC, OQ)"),
    ("05_Open_Questions_Gates", "Client decisions that gate engineering (from §25 / §31)"),
    ("06_Screen_API_DB_Index", "New screens, proposed APIs, new tables — implementation index"),
]
ws.cell(r, 1, "Sheet").font = FONT_WHITE
ws.cell(r, 1).fill = FILLS["header"]
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
ws.cell(r, 2, "Contents").font = FONT_WHITE
ws.cell(r, 2).fill = FILLS["header"]
for name, meaning in sheets_help:
    r += 1
    ws.cell(r, 1, name).font = FONT_BOLD
    ws.cell(r, 1).border = THIN
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, meaning).font = FONT_N
    ws.cell(r, 2).border = THIN

r += 2
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "STATUS VALUES (column Status on the master register)").font = FONT_WHITE_MD
ws.cell(r, 1).fill = FILLS["section"]
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.cell(r, 1, "Not Started  ·  Blocked (waiting OQ/vendor)  ·  In Progress  ·  In Review  ·  Done  ·  Deferred").font = FONT_N

set_widths(ws, [28, 28, 18, 18, 18, 18, 40])
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = "1:2"
ws.oddHeader.left.text = "NIGA Homeocentrum Task Register"
ws.oddFooter.right.text = "Page &P of &N"

# ---------- Master register (build first so summaries can count) ----------
wsM = wb.create_sheet("02_Master_Task_Register", 1)
wsM.sheet_properties.tabColor = "B9770E"
apply_header(wsM, HEADERS)
status_dv = DataValidation(
    type="list",
    formula1='"Not Started,Blocked,In Progress,In Review,Done,Deferred"',
    allow_blank=True,
)
status_dv.error = "Pick a status from the list"
status_dv.errorTitle = "Invalid status"
wsM.add_data_validation(status_dv)

for i, t in enumerate(tasks, 2):
    row = [
        t["seq"],
        t["phase"],
        t["phase_name"],
        t["wbs_main"] + " / " + t["sub_id"],
        t["main_id"],
        t["main"],
        t["sub_id"],
        t["sub"],
        t["work_type"],
        t["gap"],
        t["module"],
        t["req"],
        t["surface"],
        t["fe"],
        t["be"],
        t["db"],
        t["api"],
        t["integ"],
        t["prio"],
        t["effort"],
        t["depends"],
        t["source"],
        t["existing"],
        t["gap_text"],
        t["accept"],
        "Not Started",
        "",
        t["notes"],
    ]
    for col, val in enumerate(row, 1):
        cell = wsM.cell(i, col, val)
        style_data_cell(cell, center=col in (1, 2, 14, 15, 16, 17, 18, 19, 20, 26))
        if col == 9:
            cell.fill = fill_for_type(t["work_type"])
            cell.font = FONT_BOLD
        elif col == 19:
            cell.fill = fill_for_prio(t["prio"])
            cell.font = FONT_BOLD
        elif col == 2:
            pf = PHASE_FILL.get(str(t["phase"]), FILLS["white"])
            cell.fill = pf
            cell.font = FONT_WHITE if t["phase"] not in () else FONT_WHITE
            cell.font = FONT_WHITE
        elif i % 2 == 0 and col not in (2, 9, 19):
            cell.fill = FILLS["alt"]
    # phase font already white on col 2
    status_dv.add(wsM.cell(i, 26))
    # row height
    wsM.row_dimensions[i].height = 48

wsM.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(tasks)+1}"
set_widths(
    wsM,
    [6, 8, 28, 22, 14, 42, 14, 55, 22, 12, 22, 28, 22, 10, 10, 10, 10, 12, 10, 12, 28, 18, 40, 40, 40, 14, 14, 28],
)
wsM.auto_filter.ref = f"A1:AB{len(tasks)+1}"
wsM.freeze_panes = "A2"
wsM.page_setup.orientation = "landscape"
wsM.page_setup.fitToPage = True
wsM.page_setup.fitToWidth = 1
wsM.page_setup.fitToHeight = 0
wsM.page_setup.paperSize = wsM.PAPERSIZE_A3
wsM.print_title_rows = "1:1"
wsM.oddHeader.left.text = "NIGA Homeocentrum — Master Task Register (sequential — do not skip)"
wsM.oddFooter.right.text = "Page &P of &N"
wsM.auto_filter.ref = f"A1:AB{1+len(tasks)}"

# ---------- Phase summary ----------
wsP = wb.create_sheet("01_Phase_Summary", 1)
wsP.sheet_properties.tabColor = "1A5276"
wsP.merge_cells("A1:L1")
wsP["A1"] = "Phase summary — sequential delivery spine (dependency-driven, not calendar-driven)"
wsP["A1"].font = FONT_WHITE_LG
ws["A1"].fill = FILLS["header"]
wsP["A1"].fill = FILLS["header"]
wsP.row_dimensions[1].height = 26

phase_order = [
    ("PRE", "Pre-Programme — Client Gates & Decisions", "Must complete Critical OQs, vendors, legal start, Phase 0 approval. Blocks honest estimates and Phase 2."),
    ("0", "Phase 0 — Platform hygiene", "Unconditional. Hashing, reset, logout, profiles, RBAC, admin home, secrets, roles, freeze classic API, OTP foundation, regression suite, Patient Board spike."),
    ("1", "Phase 1 — Appointment product completeness", "Reschedule, cancel, schedule screen, reception staff UI, visit type, queue, complaints wiring, case export."),
    ("2", "Phase 2 — Money spine and Account role", "KEYSTONE. Ledger, webhook, S2/S3, recon, exceptions, settlements, OTP payouts, refunds, GST, KYC. Booking/tele/HomeoMeds are commercially meaningless without this."),
    ("3", "Phase 3 — Patient access", "Public booking, credentialing, legal pages, SMS, WhatsApp completion, support tickets."),
    ("4", "Phase 4 — Clinical differentiation", "Potency, notes/eRx snapshot, COG (gated), audio accuracy, 3D, analytics, timeline. After Patient Board extract."),
    ("5", "Phase 5 — Telemedicine", "Availability, queue, vendor video, rejoin, consent. Abstract vendor; build queue independently."),
    ("6", "Phase 6 — Mobile applications", "Patient app (~60 screens) then Doctor app. Same APIs. No case-taking on mobile. Interim: responsive patient web."),
    ("7", "Phase 7 — HomeoMeds", "Pharmacy, licence gate, immutable eRx handoff, quote-before-pay, medicine ledger, exceptions, OTP."),
    ("8", "Phase 8 — Continuity and trust", "Diary, CliniSight, reviews, consent centre, ranking explanation."),
    ("REG", "Existing 100% catalogue", "Every admin master, doctor 100% item, reception 100% item — regression so nothing is skipped."),
    ("QA", "Cross-cutting QA, NFR, integrations, events, traps", "Test types, recommended NFRs, nine integrations, event fan-out, explicit traps."),
]

ph_headers = ["Phase", "Phase name", "Why this sequence", "Total tasks", "Existing", "Existing Improvement", "Existing Modification", "New", "New Integration", "Config / Migration / Client", "P0", "P1"]
for col, h in enumerate(ph_headers, 1):
    cell = wsP.cell(3, col, h)
    cell.font = FONT_WHITE
    cell.fill = FILLS["header"]
    cell.alignment = CENTER
    cell.border = THIN

from collections import Counter, defaultdict

def count_phase(ph):
    rows = [t for t in tasks if str(t["phase"]) == str(ph)]
    wt = Counter(t["work_type"] for t in rows)
    pr = Counter(t["prio"] for t in rows)
    cfg = sum(1 for t in rows if t["work_type"] in ("Configuration", "Data Migration", "Client Decision", "Regression"))
    return len(rows), wt, pr, cfg

r = 4
for ph, name, why in phase_order:
    n, wt, pr, cfg = count_phase(ph)
    vals = [
        ph, name, why, n,
        wt.get("Existing", 0) + wt.get("Regression", 0),
        wt.get("Existing Improvement", 0),
        wt.get("Existing Modification", 0),
        wt.get("New", 0),
        wt.get("New Integration", 0),
        wt.get("Configuration", 0) + wt.get("Data Migration", 0) + wt.get("Client Decision", 0),
        pr.get("P0", 0),
        pr.get("P1", 0),
    ]
    for col, val in enumerate(vals, 1):
        cell = wsP.cell(r, col, val)
        style_data_cell(cell)
        if col == 1:
            cell.fill = PHASE_FILL.get(str(ph), FILLS["light"])
            cell.font = FONT_WHITE
        elif col == 4:
            cell.font = FONT_BOLD
            cell.alignment = CENTER
        elif col >= 5:
            cell.alignment = CENTER
    wsP.row_dimensions[r].height = 48
    r += 1

wsP.cell(r + 1, 1, "Total tasks in register").font = FONT_BOLD
wsP.cell(r + 1, 2, len(tasks)).font = FONT_BOLD
wsP.merge_cells(start_row=r + 3, start_column=1, end_row=r + 3, end_column=12)
wsP.cell(r + 3, 1, "Dependency note: P0 hygiene → P1 appointments → P2 money → P3 patient access → P5 tele → P6 mobile → P7 HomeoMeds → P8. P2 also feeds P4 (analytics/revenue). Phase 4 eRx snapshot is required before Phase 7. Legal starts in PRE, not Phase 3.").font = FONT_N
wsP.cell(r + 3, 1).alignment = WRAP
wsP.row_dimensions[r + 3].height = 40

set_widths(wsP, [10, 42, 55, 12, 12, 18, 20, 10, 16, 24, 8, 8])
wsP.freeze_panes = "A4"

# ---------- Classification summary ----------
wsC = wb.create_sheet("03_Existing_vs_New", 3)
wsC.sheet_properties.tabColor = "117A65"
wsC.merge_cells("A1:D1")
wsC["A1"] = "Existing vs New vs Existing Improvement — counts (do not skip Existing rows — they are regression/hygiene)"
wsC["A1"].font = FONT_WHITE_LG
wsC["A1"].fill = FILLS["header"]
wsC.row_dimensions[1].height = 26

wt_all = Counter(t["work_type"] for t in tasks)
wsC["A3"] = "Work Type"
wsC["B3"] = "Task count"
wsC["C3"] = "% of register"
wsC["D3"] = "How to treat in the plan"
for col in range(1, 5):
    wsC.cell(3, col).font = FONT_WHITE
    wsC.cell(3, col).fill = FILLS["header"]
    wsC.cell(3, col).border = THIN

treat = {
    "Existing": "Regression pack each phase. No rebuild.",
    "Existing Improvement": "Complete the delta. Reuse existing APIs/screens.",
    "Existing Modification": "Change behaviour of a live path (security, gates, ports).",
    "New": "Full build: DB + API + UI + tests + ACL.",
    "New Integration": "Vendor contract + sandbox + failure/retry + secrets.",
    "Configuration": "Ops/config PR, not a sprint 'feature' but still P0 when security.",
    "Data Migration": "Rehearse on production copy; rollback plan.",
    "Client Decision": "Owner = client. Track as a gate, not an engineering estimate.",
    "Regression": "Same as Existing — listed item-by-item so 100% modules are not forgotten.",
}
r = 4
for name in ["Existing", "Existing Improvement", "Existing Modification", "New", "New Integration", "Configuration", "Data Migration", "Client Decision", "Regression"]:
    n = wt_all.get(name, 0)
    wsC.cell(r, 1, name).fill = fill_for_type(name)
    wsC.cell(r, 1).font = FONT_BOLD
    wsC.cell(r, 1).border = THIN
    wsC.cell(r, 2, n).alignment = CENTER
    wsC.cell(r, 2).border = THIN
    wsC.cell(r, 3, round(100 * n / len(tasks), 1) if tasks else 0).alignment = CENTER
    wsC.cell(r, 3).border = THIN
    wsC.cell(r, 4, treat.get(name, "")).alignment = WRAP
    wsC.cell(r, 4).border = THIN
    r += 1

wsC.cell(r, 1, "TOTAL").font = FONT_BOLD
wsC.cell(r, 2, len(tasks)).font = FONT_BOLD

# Priority x type matrix
r += 3
wsC.cell(r, 1, "Priority × Work Type").font = FONT_WHITE_MD
wsC.cell(r, 1).fill = FILLS["section"]
wsC.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
prios = ["P0", "P1", "P2", "P3"]
types_row = ["Existing", "Existing Improvement", "Existing Modification", "New", "New Integration", "Client Decision"]
wsC.cell(r, 1, "Priority").font = FONT_WHITE
wsC.cell(r, 1).fill = FILLS["header"]
for i, tp in enumerate(types_row, 2):
    c = wsC.cell(r, i, tp)
    c.font = FONT_WHITE
    c.fill = FILLS["header"]
    c.alignment = CENTER
    c.border = THIN
r += 1
start_chart = r
for p in prios:
    wsC.cell(r, 1, p).fill = fill_for_prio(p)
    wsC.cell(r, 1).font = FONT_BOLD
    wsC.cell(r, 1).border = THIN
    for i, tp in enumerate(types_row, 2):
        n = sum(1 for t in tasks if t["prio"] == p and t["work_type"] == tp)
        wsC.cell(r, i, n).alignment = CENTER
        wsC.cell(r, i).border = THIN
    r += 1

set_widths(wsC, [28, 14, 14, 70, 22, 22])

# ---------- Traceability ----------
wsT = wb.create_sheet("04_Requirement_Traceability", 4)
wsT.sheet_properties.tabColor = "6C3483"
wsT.merge_cells("A1:F1")
wsT["A1"] = "Requirement traceability — every analysis ID maps to sequential tasks (nothing dropped from §7 / §9 / §25 / §28)"
wsT["A1"].font = FONT_WHITE_LG
wsC["A1"].fill = FILLS["header"]
wsT["A1"].fill = FILLS["header"]

tr_headers = ["Requirement ID family", "IDs covered in this register", "Task count", "Primary phases", "Classification mix", "Notes"]
for col, h in enumerate(tr_headers, 1):
    cell = wsT.cell(3, col, h)
    cell.font = FONT_WHITE
    cell.fill = FILLS["header"]
    cell.border = THIN
    cell.alignment = CENTER

families = [
    ("PRE / OQ-A Critical", "OQ-A1–A9, §31 steps", "PRE", "Client gates for Phase 2"),
    ("PRE / OQ-B High", "OQ-B1–B6", "PRE", "Code-then-name, COG, ranking, unused roles"),
    ("PRE / OQ-C/D", "OQ-C1–C8, OQ-D1–D3", "PRE", "NFR, recording, missing module 7, i18n, CliniSight, geo, patient web, pharmacy ops, Deep Analytics, Velzon, wallet"),
    ("PLT-01–13 Platform", "PLT-01 … PLT-13, S-1–S-11", "0", "Hygiene, RBAC, dual-API freeze"),
    ("ADM Admin", "ADM-01–10", "0,2,3,7,REG", "Credentialing, recon, tickets, users, enquiry, pharmacy, 25 masters"),
    ("FIN / SC Account", "FIN-01–10, SC-01–04", "2, QA", "Money spine and constraints"),
    ("GST Guest/Public", "GST-01–09", "2,3", "Booking, pay, legal, registration"),
    ("DOC Doctor web", "DOC-01–33", "1,2,4,5,REG", "New + improvements + 100% regression"),
    ("REC Reception", "REC-01–10", "0,1,2,REG", "Shared APIs + reception-first UX"),
    ("PAT Patient mobile", "PAT-M1–M11 (no M7 — OQ-C3)", "6,7,8", "~60 screens; M8 with HomeoMeds"),
    ("DOCM Doctor mobile", "DOCM-01–12", "6", "No case-taking"),
    ("PHR HomeoMeds", "PHR-01–11, ADM-08/09", "7", "Marketplace"),
    ("F-01–F-14 Features", "Mapped via ADM/FIN/GST/DOC/PHR IDs", "1–7", "§12 specifications exploded into sub-tasks"),
    ("QA / NFR / INT / EVT", "§17–20, §10.3 events, §26 traps", "QA", "Cross-cutting"),
]

def ids_in(*prefixes):
    return [t for t in tasks if any(p in (t["req"] + t["main_id"] + t["sub_id"]) for p in prefixes)]

r = 4
for fam, ids, phases, notes in families:
    rows = []
    # rough filter
    key = fam.split()[0].split("/")[0].strip()
    if "OQ-A" in fam:
        rows = [t for t in tasks if "OQ-A" in t["req"] or t["main_id"] == "PRE-01"]
    elif "OQ-B" in fam:
        rows = [t for t in tasks if "OQ-B" in t["req"] or t["main_id"] == "PRE-02"]
    elif "OQ-C" in fam:
        rows = [t for t in tasks if t["main_id"] == "PRE-03"]
    elif "PLT" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("PLT") or t["phase"] == "0"]
    elif "ADM" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("ADM") or "ADM-" in t["req"]]
    elif "FIN" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("FIN") or t["main_id"].startswith("SC") or "FIN-" in t["main_id"]]
    elif "GST" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("GST")]
    elif "DOC Doctor" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("DOC")]
    elif "REC" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("REC")]
    elif "PAT" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("PAT")]
    elif "DOCM" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("DOCM") or t["main_id"] == "DOCM"]
    elif "PHR" in fam:
        rows = [t for t in tasks if t["main_id"].startswith("PHR") or t["main_id"] == "ADM-08"]
    elif "F-01" in fam:
        rows = [t for t in tasks if t["phase"] in ("1", "2", "3", "4", "5", "7")]
    else:
        rows = [t for t in tasks if t["phase"] in ("QA",)]
    mix = Counter(t["work_type"] for t in rows)
    mix_s = ", ".join(f"{k}:{v}" for k, v in mix.most_common(4))
    vals = [fam, ids, len(rows), phases, mix_s, notes]
    for col, val in enumerate(vals, 1):
        cell = wsT.cell(r, col, val)
        style_data_cell(cell)
    wsT.row_dimensions[r].height = 36
    r += 1

wsT.cell(r + 1, 1, "Note").font = FONT_BOLD
wsT.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=6)
wsT.cell(r + 1, 2, "Hello Homeo Doc module 7 is absent from the client list (OQ-C3). Management/Supervisor/Inspector are enum-only (OQ-B6). Deep Analytics is a placeholder (OQ-D1). These are rows in PRE, not silent drops.")
wsT.cell(r + 1, 2).alignment = WRAP
wsT.row_dimensions[r + 1].height = 36
set_widths(wsT, [28, 45, 12, 22, 45, 50])

# ---------- Open questions ----------
wsQ = wb.create_sheet("05_Open_Questions_Gates", 5)
wsQ.sheet_properties.tabColor = "922B21"
wsQ.merge_cells("A1:G1")
wsQ["A1"] = "Open questions and next-step gates — filter Master register Work Type = Client Decision for the full task rows"
wsQ["A1"].font = FONT_WHITE_LG
wsQ["A1"].fill = FILLS["header"]
qh = ["Seq in register", "ID", "Question / gate", "Blocks", "Priority", "Owner (suggested)", "Status"]
for col, h in enumerate(qh, 1):
    cell = wsQ.cell(3, col, h)
    cell.font = FONT_WHITE
    cell.fill = FILLS["header"]
    cell.border = THIN
r = 4
for t in tasks:
    if t["work_type"] != "Client Decision":
        continue
    vals = [t["seq"], t["sub_id"], t["sub"], t["depends"] if t["depends"] != "—" else t["phase_name"], t["prio"], "Client (business / finance / legal / clinical as applicable)", "Not Started"]
    for col, val in enumerate(vals, 1):
        cell = wsQ.cell(r, col, val)
        style_data_cell(cell)
        if col == 5:
            cell.fill = fill_for_prio(t["prio"])
    wsQ.row_dimensions[r].height = 40
    r += 1
wsQ.auto_filter.ref = f"A3:G{r-1}"
wsQ.freeze_panes = "A4"
set_widths(wsQ, [14, 14, 80, 28, 10, 42, 14])
status_dv2 = DataValidation(type="list", formula1='"Not Started,Blocked,In Progress,Answered,Deferred"', allow_blank=True)
wsQ.add_data_validation(status_dv2)
status_dv2.add(f"G4:G{r-1}")

# ---------- Screen / API / DB index ----------
wsI = wb.create_sheet("06_Screen_API_DB_Index", 6)
wsI.sheet_properties.tabColor = "1F618D"
wsI.merge_cells("A1:D1")
wsI["A1"] = "Implementation index — new screens, proposed APIs, new tables (from analysis §§12–16). Tasks in the master register implement these."
wsI["A1"].font = FONT_WHITE_LG
wsI["A1"].fill = FILLS["header"]

wsI["A3"] = "NEW / CHANGED SCREENS (from §13)"
wsI["A3"].font = FONT_WHITE
wsI["A3"].fill = FILLS["section"]
wsI.merge_cells("A3:D3")
screens = [
    ("Admin", "Credentialing queue and detail · Consult reconciliation · Payment exceptions · Support tickets list/detail · Enquiries inbox · Pharmacy activation · HomeoMeds exceptions · Admin KPI home"),
    ("Account", "Ledger · Consult reconciliation · Medicine ledger · Exceptions · Settlements · Payouts · Refunds · Tax · Payees · Clinic collections"),
    ("Doctor", "Schedule · Reception staff · Follow-up analysis · Clinic performance · SMS templates · Teleconsult room · Doctor profile (real)"),
    ("Reception", "Reception home/queue · Read-only schedule · Case paper · Reception profile · Collect payment modal"),
    ("Public", "/book · /book/:id · /book/:id/slots · /book/confirm · /book/pay/:bookingId · success/failure"),
    ("Pharmacy", "Onboarding · Licence · Order inbox · Order detail (OTP accept, quote, dispatch) · History"),
    ("Patient app", "~60 Hello Homeo Doc screens (modules 1–6, 8, 9, 10 + push) — listed as PAT-M* sub-tasks"),
    ("Doctor app", "Login, onboarding, queue, availability, join, refill, earnings, offline/crash — DOCM-01–12"),
    ("Modified SPA", "Admin dashboard replace · Users list · ForgetPassword thunk · user-profile · Register · Widgets.js · BestSellingProducts · PatientBoard.js · DailySchedule modal promote · WhatsApp modal · LayoutMenuData · allRoutes/AuthProtected · roles.js · App.js · landing/legal"),
]
wsI["A4"] = "Area"
wsI["B4"] = "Screens"
wsI["A4"].font = FONT_WHITE
wsI["B4"].font = FONT_WHITE
wsI["A4"].fill = FILLS["header"]
wsI["B4"].fill = FILLS["header"]
wsI.merge_cells("B4:D4")
r = 5
for area, sc in screens:
    wsI.cell(r, 1, area).font = FONT_BOLD
    wsI.cell(r, 1).border = THIN
    wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    wsI.cell(r, 2, sc).alignment = WRAP
    wsI.cell(r, 2).border = THIN
    wsI.row_dimensions[r].height = 36
    r += 1

r += 1
wsI.cell(r, 1, "PROPOSED APIs (target contracts — design sign-off required)").font = FONT_WHITE
wsI.cell(r, 1).fill = FILLS["section"]
wsI.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
wsI.cell(r, 1, "Group").font = FONT_WHITE
wsI.cell(r, 1).fill = FILLS["header"]
wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
wsI.cell(r, 2, "Endpoints (abbreviated)").font = FONT_WHITE
wsI.cell(r, 2).fill = FILLS["header"]
apis = [
    ("Auth / profile", "Logout · Forgot/Reset/ChangePassword · Profile/Me · Profile/Photo"),
    ("Credentialing", "Queue · {id} · Approve · Reject · RequestInfo · Documents · MyStatus · RegisterDoctor (modify)"),
    ("Payments", "Orders · Webhook/Razorpay · Verify · Admin/Reconciliation · Exceptions · Exceptions/{id}/Resolve · ConsultFee/Config · Earnings/Summary · Order/GenerateOrderId (port)"),
    ("Consult payment", "CollectAtReception · ByAppointment/{patientAppId}"),
    ("Ledger / settle", "Ledger · Ledger/Medicine · Settlements · ApprovePayout (OTP) · Refunds"),
    ("Public booking", "Doctors · Doctors/{id}/Profile · Doctors/{id}/Slots · Create · {bookingToken} · Waitlist · InstantConsult"),
    ("Patient auth", "RequestOtp · VerifyOtp"),
    ("Appointments", "RescheduleAppointment · CancelAppointment · ChangeLog/{id} · PATCH VisitType"),
    ("Clinical", "PotencyMaster CRUD · Repertorization/CenterOfGravity · Erx/ByAppointment · Erx/Export · Erx/RefillInbox · Refill Approve/Reject · SavePrescriptionDetail (extend) · notes extend"),
    ("Telemedicine", "Availability · Availability/{doctorId} · Queue · Sessions · Sessions/{id} · End · Rejoin · Consent"),
    ("SMS / push", "Sms/Templates · Send · History · Events/Test · Devices/Register · Notifications/Send · Notifications"),
    ("Support", "SupportTicket · {id} · Messages · Status · Assign"),
    ("Analytics", "FollowUpDue · FollowUpSummary · ClinicPerformance · AdminDashboard/Overview"),
    ("HomeoMeds", "Pharmacy/Onboard · Licence · Admin Activate · Sellers · Orders · Accept (OTP) · Reject · Quote · Pay · Status · Exceptions · RefillRequest"),
    ("Deprecated", "POST api/Login/authenticate (classic second auth path) — remove"),
]
r += 1
for g, ep in apis:
    wsI.cell(r, 1, g).font = FONT_BOLD
    wsI.cell(r, 1).border = THIN
    wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    wsI.cell(r, 2, ep).alignment = WRAP
    wsI.cell(r, 2).border = THIN
    wsI.row_dimensions[r].height = 32
    r += 1

r += 1
wsI.cell(r, 1, "NEW / MODIFIED TABLES (from §16)").font = FONT_WHITE
wsI.cell(r, 1).fill = FILLS["section"]
wsI.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
wsI.cell(r, 1, "Domain").font = FONT_WHITE
wsI.cell(r, 1).fill = FILLS["header"]
wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
wsI.cell(r, 2, "Objects").font = FONT_WHITE
wsI.cell(r, 2).fill = FILLS["header"]
tables = [
    ("Modified existing", "UserMaster · Doctor · Patient · PatientAppointment · AppointmentHistoryNote · PrescriptionRemedyDetail · RoleMaster/RoleDetail/MenuMaster · EnquiryDetail · UserLoginStatus (activate) · PackageEntryDetail (shape UNCHANGED, S1 link only)"),
    ("Trust", "DoctorVerification · DoctorCredentialDocument"),
    ("Money", "PaymentOrder · LedgerEntry · ConsultPayment · ConsultFeeConfig · PaymentException · SettlementRun · Payout · Refund · Invoice (Proposed)"),
    ("Control plane", "OtpChallenge · OtpAuditLog · PasswordResetToken"),
    ("Appointments", "AppointmentChangeLog · BookingWaitlist"),
    ("Patient ecosystem", "FamilyMember · CaregiverAuth · ConsentRecord"),
    ("Prescription", "PotencyMaster · ErxSnapshot · ErxSnapshotItem · RefillRequest"),
    ("Telemedicine", "TeleAvailability · TeleSession · TeleConsentLog · TeleSessionEvent"),
    ("Messaging", "SmsTemplate · SmsMessageLog · DeviceToken · AppNotification"),
    ("Support", "SupportTicket · SupportTicketMessage · SupportTicketAttachment"),
    ("HomeoMeds", "PharmacyPartner · PharmacyLicence · SellerRoutingRule · MedicineOrder · MedicineOrderItem · MedicineQuote · MedicineOrderEvent"),
    ("Continuity", "Review · SymptomDiary · FollowUpTask"),
]
r += 1
for g, ep in tables:
    wsI.cell(r, 1, g).font = FONT_BOLD
    wsI.cell(r, 1).border = THIN
    wsI.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    wsI.cell(r, 2, ep).alignment = WRAP
    wsI.cell(r, 2).border = THIN
    wsI.row_dimensions[r].height = 36
    r += 1

set_widths(wsI, [28, 40, 20, 40])
wsI.freeze_panes = "A4"

# ---------- Print / freeze extras ----------
for sheet in wb.worksheets:
    wsx = sheet
    wsx.page_setup.orientation = "landscape"
    wsx.page_setup.fitToPage = True
    wsx.page_setup.fitToWidth = 1
    wsx.page_setup.fitToHeight = 0
    wsx.page_setup.paperSize = wsM.PAPERSIZE_A3
    wsx.page_setup.leftMargin = 0.4
    wsx.page_setup.rightMargin = 0.4
    wsx.sheet_view.zoomScale = 90 if sheet.title.startswith("02") else 100

wb.save(OUT)
print(f"Wrote {len(tasks)} sequential tasks → {OUT}")

