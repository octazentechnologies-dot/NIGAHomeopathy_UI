#!/usr/bin/env python3
"""
Homeocentrum Feature Delivery Document → Sequential Development Task Register.

Source of truth: docs/Homeocentrum_Feature_Delivery_Document.pdf
Prepared for Homeocentrum / NIGA · 29 August 2026

Every PDF feature line is a Main Task. Each Main Task is broken into sequential
Sub Tasks (DB → API → Web/Mobile → Integration → QA). Develop strictly by Seq.
Do not start a later Seq if its Depends On items are not Done.

Work Type:
  Existing              — live in code today; keep, do not rebuild; regression + ACL
  Existing Improvement  — screen/API exists but incomplete, stub, or missing a role
  Existing Modification — exists but behaviour must change (e.g. plaintext password)
  New                   — not in any of the three repos
  New Integration       — vendor/gateway (Razorpay webhook, video SDK, FCM, SMS)
  Configuration         — rates, flags, templates — not hardcoded
  Data Migration        — additive schema / backfill
  Client Decision       — blocks engineering until signed
  Regression            — prove an Existing surface still works after later phases
"""

from collections import Counter, defaultdict
from copy import copy
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "/Users/OctazenWork/NIGA Project/NigaHomeopathy-UI/docs/Homeocentrum_Feature_Delivery_Task_Register.xlsx"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

FILLS = {
    "header": PatternFill("solid", fgColor="1B3A4B"),
    "section": PatternFill("solid", fgColor="1B3A4B"),
    "pre": PatternFill("solid", fgColor="2C3E50"),
    "phase0": PatternFill("solid", fgColor="7B241C"),
    "phase1": PatternFill("solid", fgColor="1A5276"),
    "phase2": PatternFill("solid", fgColor="B9770E"),
    "phase3": PatternFill("solid", fgColor="117A65"),
    "phase4": PatternFill("solid", fgColor="6C3483"),
    "phase5": PatternFill("solid", fgColor="1F618D"),
    "phase6": PatternFill("solid", fgColor="0E6655"),
    "phase7": PatternFill("solid", fgColor="922B21"),
    "phase8": PatternFill("solid", fgColor="4A235A"),
    "phase9": PatternFill("solid", fgColor="1B4F72"),
    "phase10": PatternFill("solid", fgColor="145A32"),
    "phase11": PatternFill("solid", fgColor="6E2C00"),
    "phase12": PatternFill("solid", fgColor="1A5276"),
    "phase13": PatternFill("solid", fgColor="4A235A"),
    "phase14": PatternFill("solid", fgColor="0E6251"),
    "phase15": PatternFill("solid", fgColor="7B241C"),
    "phase16": PatternFill("solid", fgColor="1F618D"),
    "phase17": PatternFill("solid", fgColor="117A65"),
    "phase18": PatternFill("solid", fgColor="6C3483"),
    "phase19": PatternFill("solid", fgColor="34495E"),
    "exist": PatternFill("solid", fgColor="D5F5E3"),
    "improve": PatternFill("solid", fgColor="FCF3CF"),
    "modify": PatternFill("solid", fgColor="FDEBD0"),
    "new": PatternFill("solid", fgColor="D6EAF8"),
    "newint": PatternFill("solid", fgColor="D5D8DC"),
    "cfg": PatternFill("solid", fgColor="E8DAEF"),
    "mig": PatternFill("solid", fgColor="FADBD8"),
    "client": PatternFill("solid", fgColor="F5B7B1"),
    "reg": PatternFill("solid", fgColor="D5F5E3"),
    "alt": PatternFill("solid", fgColor="F8F9F9"),
    "white": PatternFill("solid", fgColor="FFFFFF"),
    "gold": PatternFill("solid", fgColor="FEF9E7"),
    "redsoft": PatternFill("solid", fgColor="FDEDEC"),
    "light": PatternFill("solid", fgColor="F4F6F7"),
}

FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_WHITE_LG = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_WHITE_MD = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="1B3A4B", size=18)
FONT_H = Font(name="Calibri", bold=True, color="1B3A4B", size=13)
FONT_B = Font(name="Calibri", bold=True, color="1B3A4B", size=11)
FONT_N = Font(name="Calibri", size=10, color="1C2833")
FONT_S = Font(name="Calibri", size=9, color="5D6D7E")
FONT_BOLD = Font(name="Calibri", bold=True, size=10, color="1C2833")

TYPE_FILL = {
    "Existing": FILLS["exist"],
    "Existing Improvement": FILLS["improve"],
    "Existing Modification": FILLS["modify"],
    "New": FILLS["new"],
    "New Integration": FILLS["newint"],
    "Configuration": FILLS["cfg"],
    "Data Migration": FILLS["mig"],
    "Client Decision": FILLS["client"],
    "Regression": FILLS["reg"],
}

PHASE_FILL = {str(i): FILLS.get(f"phase{i}", FILLS["section"]) for i in range(20)}
PHASE_FILL["PRE"] = FILLS["pre"]
PHASE_FILL["QA"] = FILLS["phase19"]

HEADERS = [
    "Seq",
    "Phase",
    "Phase Name",
    "WBS",
    "Main Task ID",
    "Main Task",
    "Sub Task ID",
    "Sub Task",
    "Work Type",
    "Module / Area",
    "PDF Section",
    "PDF Feature (exact wording)",
    "PDF Delivery Text",
    "Application Surface",
    "Layer",
    "Frontend",
    "Backend",
    "Database",
    "API",
    "Mobile",
    "Integration",
    "Priority",
    "Effort Band",
    "Depends On",
    "Existing Capability (AS-IS)",
    "Gap / Change Required (TO-BE)",
    "Acceptance Criteria",
    "Status",
    "Owner",
    "Notes",
]

Y, N, P = "Yes", "No", "Partial"
tasks = []
coverage = []  # one row per PDF feature line


def T(
    phase,
    phase_name,
    wbs,
    main_id,
    main,
    sub_id,
    sub,
    work_type,
    module,
    pdf_section,
    pdf_feature,
    pdf_delivery,
    surface,
    layer,
    fe,
    be,
    db,
    api,
    mobile,
    integ,
    prio,
    effort,
    depends,
    existing,
    gap,
    accept,
    notes="",
):
    tasks.append(
        {
            "phase": str(phase),
            "phase_name": phase_name,
            "wbs": wbs,
            "main_id": main_id,
            "main": main,
            "sub_id": sub_id,
            "sub": sub,
            "work_type": work_type,
            "module": module,
            "pdf_section": pdf_section,
            "pdf_feature": pdf_feature,
            "pdf_delivery": pdf_delivery,
            "surface": surface,
            "layer": layer,
            "fe": fe,
            "be": be,
            "db": db,
            "api": api,
            "mobile": mobile,
            "integ": integ,
            "prio": prio,
            "effort": effort,
            "depends": depends,
            "existing": existing,
            "gap": gap,
            "accept": accept,
            "notes": notes,
        }
    )


def cov(section, feature, delivered, work_type, main_id, as_is, phase):
    coverage.append(
        {
            "section": section,
            "feature": feature,
            "delivered": delivered,
            "work_type": work_type,
            "main_id": main_id,
            "as_is": as_is,
            "phase": str(phase),
        }
    )


def add(meta, subs):
    """Expand one PDF feature into sequential sub-tasks.

    meta keys: phase, pn, wbs, mid, main, module, pdf, feat, deliv, surface,
               as_is, to_be, notes, work_type (default for coverage)
    subs: list of dicts with at least sub, and optional overrides.
    """
    n = len(subs)
    for i, s in enumerate(subs, 1):
        sid = s.get("id") or f"{meta['mid']}.{i:02d}"
        wt = s.get("wt", meta.get("wt", "New"))
        T(
            meta["phase"],
            meta["pn"],
            meta["wbs"],
            meta["mid"],
            meta["main"],
            sid,
            s["sub"],
            wt,
            meta["module"],
            meta["pdf"],
            meta["feat"],
            meta["deliv"],
            s.get("surface", meta["surface"]),
            s.get("layer", "Full"),
            s.get("fe", N),
            s.get("be", N),
            s.get("db", N),
            s.get("api", N),
            s.get("mob", N),
            s.get("int", N),
            s.get("prio", meta.get("prio", "P1")),
            s.get("eff", meta.get("eff", "M")),
            s.get("dep", meta.get("dep", "—") if i == 1 else f"{meta['mid']}.{i-1:02d}"),
            s.get("as_is", meta["as_is"]),
            s.get("gap", meta["to_be"]),
            s.get("ac", meta.get("ac", "Meets PDF delivery text; no silent failure")),
            s.get("notes", meta.get("notes", "")),
        )
    cov(
        meta["pdf"],
        meta["feat"],
        meta["deliv"],
        meta.get("wt", "New"),
        meta["mid"],
        meta["as_is"],
        meta["phase"],
    )


# ---------------------------------------------------------------------------
# PHASE HELPERS — common sub-task patterns
# ---------------------------------------------------------------------------
def db_api_ui_qa(db_text, api_text, ui_text, qa_text, ui_layer="Web SPA", extra=None):
    rows = [
        {"sub": db_text, "layer": "Database", "db": Y, "be": Y, "eff": "S"},
        {"sub": api_text, "layer": "API", "be": Y, "db": Y, "api": Y, "eff": "M"},
        {"sub": ui_text, "layer": ui_layer, "fe": Y, "api": Y, "eff": "M",
         "mob": Y if "Mobile" in ui_layer else N},
        {"sub": qa_text, "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Regression"},
    ]
    if extra:
        rows[2:2] = extra if isinstance(extra, list) else [extra]
    return rows


def exist_keep(verify, acl, notes_api):
    return [
        {"sub": verify, "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Existing"},
        {"sub": acl, "layer": "Security", "fe": Y, "be": Y, "api": Y, "eff": "S",
         "wt": "Existing Improvement"},
        {"sub": notes_api, "layer": "API hygiene", "be": Y, "api": Y, "eff": "S",
         "wt": "Existing"},
    ]


def improve_rows(gap_db, gap_api, gap_ui, qa):
    return [
        {"sub": gap_db, "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": gap_api, "layer": "API", "be": Y, "api": Y, "db": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": gap_ui, "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": qa, "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Regression"},
    ]


def new_full(db, api, ui, integ, qa, ui_layer="Web SPA"):
    rows = [
        {"sub": db, "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": api, "layer": "API", "be": Y, "db": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": ui, "layer": ui_layer, "fe": Y, "api": Y, "eff": "M", "wt": "New",
         "mob": Y if "Mobile" in ui_layer else N},
    ]
    if integ:
        rows.append(
            {"sub": integ, "layer": "Integration", "be": Y, "api": Y, "int": Y, "eff": "M",
             "wt": "New Integration"}
        )
    rows.append(
        {"sub": qa, "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Regression"}
    )
    return rows


def mobile_consume(screen, wire, states, qa):
    return [
        {"sub": screen, "layer": "Mobile UI", "fe": Y, "mob": Y, "eff": "M", "wt": "New"},
        {"sub": wire, "layer": "Mobile client", "fe": Y, "mob": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": states, "layer": "Mobile UX", "fe": Y, "mob": Y, "eff": "S", "wt": "New"},
        {"sub": qa, "layer": "QA", "fe": Y, "mob": Y, "api": Y, "eff": "S", "wt": "Regression"},
    ]


# =============================================================================
# PRE — Client decisions that block money / legal / vendors
# =============================================================================
PN_PRE = "PRE — Client gates (must sign before Phase 6 money spine)"

add(
    dict(phase="PRE", pn=PN_PRE, wbs="PRE.1", mid="PRE-01",
         main="Decide settlement model — Homeocentrum holds all money (food-delivery / quick-commerce pattern)",
         module="Finance", pdf="§9 + §8", feat="Every transaction passes through Homeocentrum",
         deliv="Built on the principle: every transaction passes through Homeocentrum, in the same way a food-delivery or quick-commerce platform holds all payments centrally.",
         surface="Account / Gateway", as_is="Only doctor SaaS subscription via client-side Razorpay; no consult/medicine ledger",
         to_be="Written decision: platform-held funds + periodic NEFT (v1) vs Razorpay Route (v2)",
         wt="Client Decision", prio="P0", eff="S", ac="Signed decision recorded; payout architecture unblocked"),
    [
        {"sub": "Workshop with finance: confirm Homeocentrum is single merchant of record for consult + medicine + subscription",
         "layer": "Decision", "eff": "S", "wt": "Client Decision"},
        {"sub": "Choose v1 payout mechanic: Homeocentrum-held + NEFT/IMPS vs Razorpay Route / linked accounts",
         "layer": "Decision", "int": Y, "eff": "S", "wt": "Client Decision"},
        {"sub": "Record decision in config (SettlementModel) — engineering must not hardcode a split model",
         "layer": "Configuration", "be": Y, "eff": "S", "wt": "Configuration"},
    ],
)

add(
    dict(phase="PRE", pn=PN_PRE, wbs="PRE.2", mid="PRE-02",
         main="Decide reception cash policy — clinic retains vs remits to Homeocentrum",
         module="Finance / Reception", pdf="§9", feat="Reception payment collection",
         deliv="Cash, UPI, card or payment link with printed receipt",
         surface="Account / Reception", as_is="No reception collection; no S3 ledger",
         to_be="Policy for cash/UPI at clinic: retain at clinic (still ledgered) or remit",
         wt="Client Decision", prio="P0", ac="Written policy; clinic collection view can be designed"),
    [
        {"sub": "OQ: cash collected at reception retained by clinic or remitted to Homeocentrum?",
         "layer": "Decision", "wt": "Client Decision", "eff": "S"},
        {"sub": "Document GST treatment on consult fee, platform fee, and medicine (on which amount, which GSTIN)",
         "layer": "Decision", "wt": "Client Decision", "eff": "S"},
        {"sub": "Confirm commission %, settlement hold T+N days, invoice numbering series — configuration not code",
         "layer": "Configuration", "wt": "Configuration", "be": Y, "eff": "S"},
    ],
)

add(
    dict(phase="PRE", pn=PN_PRE, wbs="PRE.3", mid="PRE-03",
         main="Vendor and legal gates — video SDK, SMS, WhatsApp, push, DPDP copy",
         module="Platform", pdf="§10 + §14 + §17", feat="Telemedicine + notifications + consent",
         deliv="Video consultation; SMS/WhatsApp/Push/Email; consent records",
         surface="Shared", as_is="WhatsApp Meta Cloud partial; no video SDK; no SMS provider; no FCM; privacy/terms incomplete for DPDP/tele/pharmacy",
         to_be="Named vendors + DPDP-ready legal copy before those phases start",
         wt="Client Decision", prio="P0", ac="Vendors contracted or shortlisted; legal pages drafted"),
    [
        {"sub": "Select in-browser + mobile video vendor (Agora / Twilio / Daily / WebRTC self-host) — required for §10",
         "layer": "Decision", "int": Y, "wt": "Client Decision", "eff": "S"},
        {"sub": "Select SMS provider (MSG91 / Twilio / etc.) and confirm DLT templates for appointment/OTP/cancel",
         "layer": "Decision", "int": Y, "wt": "Client Decision", "eff": "S"},
        {"sub": "Legal: privacy (data protection, recording, pharmacy consent) and terms (payments, refunds, telemedicine, medicine orders) drafted for §4 + §17",
         "layer": "Legal", "wt": "Client Decision", "eff": "M"},
        {"sub": "OQ: unpaid-booking slot hold duration vs release immediately if patient does not pay",
         "layer": "Decision", "wt": "Client Decision", "eff": "S"},
        {"sub": "OQ: live-doctor credentialing backfill — auto-Approve existing doctors? Unverified doctors visible only to own clinic?",
         "layer": "Decision", "wt": "Client Decision", "eff": "S"},
    ],
)


# =============================================================================
# PHASE 0 — Foundation: one ecosystem, identity, security (§1 + §17)
# =============================================================================
PN0 = "Phase 0 — Foundation: one ecosystem, identity, security (PDF §1 + §17)"

add(
    dict(phase=0, pn=PN0, wbs="0.1", mid="FND-01",
         main="One connected ecosystem — shared keys so a patient, appointment, prescription and payment created anywhere are visible everywhere",
         module="Platform", pdf="§1", feat="One connected ecosystem serving six types of users",
         deliv="A patient, an appointment, a prescription and a payment created anywhere in the system are visible everywhere else — there are no separate systems to reconcile.",
         surface="Shared", as_is="Doctor/Admin/Reception web only; Patient, Account, Pharmacy roles do not exist; no shared payment or eRx keys",
         to_be="Canonical keys: UserId/DoctorId, PatientId, PatientAppId, CaseId, ErxId, LedgerTxnId, MedicineOrderId — used by every surface",
         wt="New", prio="P0", ac="Any write on one surface is readable on all authorised surfaces within SLA; no dual ledgers"),
    [
        {"sub": "Publish shared-key design (DoctorId, PatientId, PatientAppId, CaseId, ErxId, LedgerTxnId, MedicineOrderId) and event list (created/rescheduled/cancelled/paid/signed/accepted)",
         "layer": "Architecture", "be": Y, "db": Y, "eff": "M", "wt": "New"},
        {"sub": "Add role enum values Patient, Account, PharmacyPartner alongside existing Admin/Doctor/Reception (keep unused Management/Supervisor/Inspector unused)",
         "layer": "Database + Auth", "be": Y, "db": Y, "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "All new domain modules on NigaHomeopathy-API (.NET 8); do not create a third API; classic API only where login/Rx-write/Razorpay already live until ported",
         "layer": "Architecture", "be": Y, "api": Y, "eff": "S", "wt": "Existing Modification",
         "as_is": "Dual API api.homeocentrum.com + api1.homeocentrum.com",
         "gap": "New work on .NET 8; port Razorpay off classic OrderController as part of Phase 6"},
        {"sub": "QA: create appointment on Reception → visible on Doctor web; pay on Patient web → same PaymentStatus on Doctor and Account (after Phase 6)",
         "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Regression", "prio": "P2"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.2", mid="FND-02",
         main="Delivery footprint — 2 mobile apps · 5 web portals · payment · telemedicine · digital prescription · HomeoMeds (programme skeleton)",
         module="Platform", pdf="§1", feat="Delivery covers",
         deliv="2 mobile applications · 5 web portals · payment system · telemedicine · digital prescription · online medicine delivery.",
         surface="Shared", as_is="1 web SPA (Admin/Doctor/Reception + marketing). No patient app, doctor app, Account portal, Pharmacy console",
         to_be="Repo/route skeleton: Patient Website, Doctor Web, Reception, Admin, Account (+ Pharmacy console under HomeoMeds); Patient + Doctor mobile later",
         wt="New", prio="P0", ac="Five web portal shells routable by role; mobile apps scheduled in Phases 16–17 not started here"),
    [
        {"sub": "Web: add Account layout + home redirect; Pharmacy layout stub; do not build screens yet",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "New"},
        {"sub": "Extend roles.js and AuthProtected with per-route ACL skeleton (deny-by-default for new routes)",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Mobile: create Patient app and Doctor app repository/bootstrap (React Native or agreed stack) — empty shells only",
         "layer": "Mobile", "mob": Y, "eff": "L", "wt": "New"},
        {"sub": "Document the five portals: Patient Website, Doctor Web Portal, Reception Portal, Admin Portal, Account Department (Pharmacy console is HomeoMeds, not a 6th portal in the PDF count)",
         "layer": "Architecture", "eff": "S", "wt": "New"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.3", mid="SEC-01",
         main="Secure login — protected credentials for every user type",
         module="Security", pdf="§17", feat="Secure login",
         deliv="Protected credentials for every user type",
         surface="All web + later mobile", as_is="POST /Account/Login on classic API; JWT in sessionStorage; doctor/user passwords compared plaintext; reception hashed on classic; fakeBackend() still mounted",
         to_be="Hashed passwords; Role claim on JWT; disable fake backend for production paths; Patient OTP login is Phase 8 (not password)",
         wt="Existing Modification", prio="P0", ac="No plaintext password compare; every role including future Patient/Account/Pharmacy authenticates through a protected path"),
    [
        {"sub": "DB/Migration: hash existing UserMaster passwords (bcrypt/PBKDF2); never store new plaintext",
         "layer": "Database", "db": Y, "be": Y, "eff": "M", "wt": "Data Migration"},
        {"sub": "API: Login verifies hash; embed Role + DoctorId claims; keep classic login URL until .NET 8 login is cut over",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "Existing Modification"},
        {"sub": "UI: remove fakeBackend() from App.js production path; login still Formik/Yup",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Modification"},
        {"sub": "QA: Admin, Doctor, Reception login; invalid password; deactivated user; SQL injection/brute-force rate limit",
         "layer": "QA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.4", mid="SEC-02",
         main="Secure password reset — time-limited reset links (stop plaintext email + fake thunk)",
         module="Security", pdf="§17", feat="Secure password reset",
         deliv="Time-limited reset links",
         surface="Admin / Doctor / Reception web", as_is="/forgot-password UI exists; thunk uses fake/Firebase; POST /users/ForgetPassword emails plaintext password",
         to_be="Tokenised reset (hash, TTL); email contains link only; ChangePassword for logged-in user",
         wt="Existing Modification", prio="P0", ac="Reset email never contains the password; token expires; used token cannot be reused"),
    [
        {"sub": "DB: PasswordResetToken (UserId, TokenHash, ExpiresAt, UsedAt)",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API: ForgotPassword sends reset link; ResetPassword consumes token; stop emailing plaintext; ChangePassword for authenticated user",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "Existing Modification"},
        {"sub": "UI: replace slices/auth/forgetpwd fake/Firebase with real APIs; /reset-password/:token success/expiry states",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "QA: expired token, reused token, unknown email (no user enumeration), logged-in change password",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.5", mid="SEC-03",
         main="Session control — proper sign-out across web and mobile",
         module="Security", pdf="§17", feat="Session control",
         deliv="Proper sign-out across web and mobile",
         surface="All", as_is="Client clears sessionStorage; no server logout / token revoke; UserLoginStatus unused",
         to_be="POST Logout persists login status; optional token denylist; mobile will call the same API in Phase 16/17",
         wt="Existing Improvement", prio="P0", ac="After logout the JWT cannot call protected APIs; board session cleared (already on client)"),
    [
        {"sub": "API: POST /Account/Logout; persist UserLoginStatus; optional denylist until token expiry",
         "layer": "API", "be": Y, "db": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "Web: Logout.js calls API then clears authUser + patient-board session (keep existing board clear)",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "Contract: mobile apps MUST use the same Logout endpoint (implemented in Phases 16–17)",
         "layer": "API contract", "api": Y, "mob": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: logout on one tab; expired session message; 401 interceptor",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.6", mid="SEC-04",
         main="Role-based access — each role sees only what it is permitted to see",
         module="Security", pdf="§17", feat="Role-based access",
         deliv="Each role sees only what it is permitted to see",
         surface="Admin / All web", as_is="RoleMaster + RoleDetail + MenuMaster exist; GetMenuByRole commented out on .NET 8; no per-route frontend ACL — any authenticated URL works",
         to_be="Re-enable GetMenuByRole; AuthProtected deny-by-default; [Authorize(Roles=)] on money/clinical APIs; Account cannot edit repertory; Admin cannot approve payouts",
         wt="Existing Improvement", prio="P0", ac="Unknown URL for a role returns 403/redirect; admin menus match DB permissions"),
    [
        {"sub": "Backend: restore GetMenuByRole; add role attributes on new money and PII APIs",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Frontend: hide routes the role cannot view; keep Velzon demo routes out of production menus",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Separation of duties seed: Account role controls money; Admin controls platform and clinical data — neither can perform the other’s role (PDF §8)",
         "layer": "Configuration", "be": Y, "db": Y, "eff": "S", "wt": "New",
         "notes": "PDF: 'the Account department controls money. Admin controls the platform and clinical data.'"},
        {"sub": "QA: Reception cannot open repertory masters; Account cannot edit diagnosis; Doctor cannot see platform-wide ledger",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.7", mid="SEC-05",
         main="Patient data protection — health records restricted to the treating doctor and the patient",
         module="Security", pdf="§17", feat="Patient data protection",
         deliv="Health records restricted to the treating doctor and the patient",
         surface="API", as_is="Many repertory APIs have [Authorize] commented; directory browsing enabled for /attachments and /Blogs; no patient-scoped ACL on case data",
         to_be="Every case/Rx/note/lab query filtered by treating DoctorId (or PatientId for patient token); disable directory browsing",
         wt="Existing Improvement", prio="P0", ac="Doctor A cannot read Doctor B’s patient board payload; unauthenticated attachment URLs fail"),
    [
        {"sub": "API: enforce DoctorId ownership on Patient, Appointment, Case, Rx, notes, labs, board backup",
         "layer": "API", "be": Y, "api": Y, "eff": "L", "wt": "Existing Improvement"},
        {"sub": "Disable directory browsing on /attachments and /Blogs; signed-URL or authorised download for documents",
         "layer": "API", "be": Y, "eff": "S", "wt": "Existing Modification"},
        {"sub": "QA: IDOR tests (swap PatientAppId / CaseId in requests)",
         "layer": "QA", "be": Y, "api": Y, "eff": "M", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.8", mid="SEC-06",
         main="Consent records infrastructure — every consent captured, stored and reviewable",
         module="Security / Trust", pdf="§17", feat="Consent records",
         deliv="Every consent captured, stored and reviewable",
         surface="Shared", as_is="AudioCaseConsentLog exists for audio case taking only",
         to_be="ConsentRecord (PatientId, Type, Version, GrantedAt, WithdrawnAt, SourceSurface) covering privacy, booking, recording, pharmacy share, marketing",
         wt="New", prio="P0", ac="Any later feature that needs consent writes this table; withdrawal is queryable"),
    [
        {"sub": "DB: ConsentRecord + ConsentType master (Privacy, Booking, TeleRecording, PharmacyShare, Marketing, Caregiver)",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API: Grant / Withdraw / List-mine / Admin-audit (no clinical content in the API response beyond type+time)",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "Reuse AudioCaseConsentLog pattern; do not fork a second consent model for telemedicine (Phase 11 will write TeleRecording into ConsentRecord)",
         "layer": "API", "be": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: grant, withdraw, re-grant; audit log immutable",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.9", mid="SEC-07",
         main="OTP verification infrastructure — applied to sensitive actions including payouts and pharmacy acceptance",
         module="Security", pdf="§17 + §8 + §12", feat="OTP verification",
         deliv="Applied to sensitive actions including payouts and pharmacy acceptance",
         surface="Shared", as_is="No generic OTP control plane; patient OTP login does not exist",
         to_be="OtpChallenge + OtpAuditLog; Account/Admin can query who/when/entity/success without seeing card data",
         wt="New", prio="P0", ac="Payout and pharmacy-accept cannot complete without a successful OTP; failures are auditable"),
    [
        {"sub": "DB: OtpChallenge, OtpAuditLog (Action, EntityType, EntityId, ToMasked, Success, At)",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API: RequestOtp / VerifyOtp (generic); rate-limit; lockout; SMS provider adapter stub until PRE-03 vendor is live",
         "layer": "API", "be": Y, "api": Y, "int": Y, "eff": "M", "wt": "New"},
        {"sub": "GET /api/Otp/Audit for Account and Admin (masked destination)",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: replay OTP, expiry, max attempts, audit completeness",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.10", mid="SEC-08",
         main="Complete audit trail — who did what and when across payments, prescriptions and approvals",
         module="Security", pdf="§17 + §9", feat="Complete audit trail",
         deliv="Who did what, and when, across payments, prescriptions and approvals / Every transaction traceable end to end by Homeocentrum",
         surface="Shared", as_is="No platform-wide audit; no AppointmentChangeLog; no payment webhook store",
         to_be="Append-only AuditEvent for money, eRx sign, credentialing, payout, pharmacy accept; payments get a dedicated trail in Phase 6",
         wt="New", prio="P0", ac="Account can reconstruct any sensitive action without calling a doctor"),
    [
        {"sub": "DB: AuditEvent (ActorUserId, Role, Action, Entity, OldJson, NewJson, At, CorrelationId)",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API middleware: write audit on mutating money/Rx/approval endpoints (start with a helper used by later phases)",
         "layer": "API", "be": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: sample sign-prescription / approve-payout (when those exist) produce audit rows",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression", "prio": "P2"},
    ],
)

add(
    dict(phase=0, pn=PN0, wbs="0.11", mid="SEC-09",
         main="Secure documents — prescriptions, reports and uploads protected from unauthorised access",
         module="Security", pdf="§17", feat="Secure documents",
         deliv="Prescriptions, reports and uploads protected from unauthorised access",
         surface="Shared", as_is="Attachments served with directory browsing; no signed URLs; no patient upload module",
         to_be="Authorised download; virus-scan hook; patient uploads in Phase 15 use this store",
         wt="New", prio="P1", ac="Guessable URL does not leak a prescription PDF"),
    [
        {"sub": "DB: SecureDocument (OwnerType, OwnerId, BlobPath, Mime, Hash, CreatedBy)",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API: upload (multipart) + authorised GET; no public folder listing",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: doctor of record can download; other doctor 403; anonymous 401",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 1 — Admin clinical knowledge masters (PDF §7.1) — Existing 100%
# =============================================================================
PN1 = "Phase 1 — Admin clinical knowledge masters (PDF §7.1) — keep live, do not rebuild"

ADMIN_MASTERS = [
    ("1.1", "ADM-R01", "Repertory", "Repertory sections", "Manage repertory sections",
     "Admin Repertory section CRUD live (pages/Admin/Repertory)"),
    ("1.2", "ADM-R02", "Repertory", "Subsection & rubric tree", "Manage the full rubric hierarchy",
     "Hierarchical subsection/rubric tree + Excel import/export live"),
    ("1.3", "ADM-R03", "Repertory", "Rubric–remedy mapping", "Map remedies to rubrics with author and grade",
     "Rubric-remedy mapping by author/grade live"),
    ("1.4", "ADM-R04", "Repertory", "Remedy-linked rubrics", "View all rubrics linked to a remedy",
     "Remedial-rubrics viewer live"),
    ("1.5", "ADM-R05", "Repertory", "Language master", "Manage languages",
     "Language master live"),
    ("1.6", "ADM-R06", "Repertory", "Body part master", "Manage body parts",
     "Body part master live"),
    ("1.7", "ADM-R07", "Repertory", "Intensity master", "Manage intensity levels",
     "Intensity master live"),
    ("1.8", "ADM-R08", "Repertory", "Remedy master", "Manage the remedy library",
     "Remedy master live"),
    ("1.9", "ADM-R09", "Repertory", "Remedy grade master", "Manage remedy grading",
     "Remedy grade master live"),
    ("1.10", "ADM-M01", "Materia Medica", "Author master", "Manage materia medica authors",
     "Author master live"),
    ("1.11", "ADM-M02", "Materia Medica", "Materia medica master", "Manage the materia medica library",
     "MM master live"),
    ("1.12", "ADM-M03", "Materia Medica", "Materia medica heads", "Manage chapter headings",
     "MM heads live"),
    ("1.13", "ADM-M04", "Materia Medica", "Materia medica remedies", "View remedy-wise materia medica",
     "Remedy-wise MM viewer live"),
    ("1.14", "ADM-D01", "Diagnosis", "Diagnosis system", "Manage diagnosis systems",
     "Diagnosis system CRUD live"),
    ("1.15", "ADM-D02", "Diagnosis", "Diagnosis therapeutics", "Manage therapeutics",
     "Therapeutics details live"),
    ("1.16", "ADM-D03", "Diagnosis", "Diagnosis conditions", "Manage conditions and linked rubrics",
     "Conditions + keyword→rubric live"),
    ("1.17", "ADM-A01", "Adverse Effect", "Drug system", "Manage conventional drug systems",
     "Drug system CRUD live"),
    ("1.18", "ADM-A02", "Adverse Effect", "Drug group", "Manage drug groups",
     "Drug group CRUD live"),
    ("1.19", "ADM-A03", "Adverse Effect", "Allopathic drug & side effects",
     "Manage drugs with side effects and adverse reactions",
     "Allopathic drug + serious/other/adverse live"),
    ("1.20", "ADM-Q01", "Clinical Questions", "Question section, group & sub-group",
     "Manage the clinical questionnaire structure",
     "Existance / question group / subgroup CRUD live"),
    ("1.21", "ADM-Q02", "Clinical Questions", "Clinical question mapping",
     "Map questions to body parts and rubrics",
     "Clinical questions mapped to body parts and rubrics live"),
    ("1.22", "ADM-3D1", "3D Body", "3D mesh key master", "Manage 3D model references",
     "Mesh key master on .NET 8 live"),
    ("1.23", "ADM-3D2", "3D Body", "3D section master", "Manage 3D body sections",
     "3D section master live"),
    ("1.24", "ADM-3D3", "3D Body", "3D hotspots", "Manage clickable points on the 3D model",
     "Hotspot master live; doctor viewer consumes them"),
]

for wbs, mid, module, feat, deliv, as_is in ADMIN_MASTERS:
    add(
        dict(phase=1, pn=PN1, wbs=wbs, mid=mid,
             main=f"Keep and harden Admin {feat} (do not rebuild)",
             module=module, pdf="§7.1", feat=feat, deliv=deliv,
             surface="Admin Portal", as_is=as_is,
             to_be="Regression pack + ACL + dual-API hygiene; Excel import/export remains where it already works",
             wt="Existing", prio="P2", eff="S",
             ac=f"{feat} still works after Phase 0 ACL; unauthorised role cannot mutate"),
        exist_keep(
            f"Regression: list / add / edit / (import-export if present) for {feat}",
            f"ACL: only Admin (and permitted roles) can mutate {feat}; Doctor is read-only consumer on Patient Board",
            f"Dual-API: confirm UI still hits the correct host (classic vs .NET 8) for {feat}; do not silently switch",
        ),
    )

add(
    dict(phase=1, pn=PN1, wbs="1.25", mid="ADM-B01",
         main="Business Management — Doctor qualifications (existing, used later by credentialing)",
         module="Business Management", pdf="§7.2", feat="Doctor qualifications",
         deliv="Manage the qualification list",
         surface="Admin Portal", as_is="Qualifications CRUD on newer API live; used by RegisterDoctor",
         to_be="Keep; credentialing Phase 8 will reference QualificationId",
         wt="Existing", prio="P2", ac="Qualification list still drives doctor registration"),
    exist_keep(
        "Regression: qualifications list/add/edit",
        "ACL on qualification APIs",
        "Confirm UI uses newer API (not classic) for qualifications",
    ),
)

add(
    dict(phase=1, pn=PN1, wbs="1.26", mid="ADM-B02",
         main="Business Management — Lab & imaging test catalog (existing, used by eRx lab orders)",
         module="Business Management", pdf="§7.2", feat="Lab & imaging test catalog",
         deliv="Manage the test catalogue",
         surface="Admin Portal", as_is="Labs & imaging CRUD live; Patient Board orders from this catalogue",
         to_be="Keep; eRx Phase 10 continues to use the same catalogue",
         wt="Existing", prio="P2", ac="Doctor can still order a lab from the catalogue"),
    exist_keep(
        "Regression: lab catalog CRUD and doctor order pick-list",
        "ACL on catalog mutate",
        "Confirm classic PatientLab APIs still read this catalogue",
    ),
)

add(
    dict(phase=1, pn=PN1, wbs="1.27", mid="ADM-B03",
         main="Business Management — Subscription packages (existing S1 SaaS; later tied to Account ledger)",
         module="Business Management", pdf="§7.2", feat="Subscription packages",
         deliv="Manage doctor subscription plans",
         surface="Admin Portal", as_is="Package CRUD live; Razorpay buy/renew partial (no webhook)",
         to_be="Do not overload PackageEntryDetail for consult/medicine money; Phase 6 ports payment verification",
         wt="Existing", prio="P2", ac="Admin can still create/edit packages; doctor still sees plan flags"),
    exist_keep(
        "Regression: package list/add/edit; doctor Widgets still reads packages",
        "ACL on package mutate",
        "Note in code: PackageEntryDetail is S1 only — never reuse for S2 consult or S5 medicine",
    ),
)

add(
    dict(phase=1, pn=PN1, wbs="1.28", mid="ADM-B04",
         main="Business Management — Roles & menu permissions (existing DB, runtime ACL leftover)",
         module="Business Management", pdf="§7.2", feat="Roles & menu permissions",
         deliv="Control exactly what each role can access",
         surface="Admin Portal", as_is="Role CRUD screens live; runtime GetMenuByRole commented; no frontend route ACL",
         to_be="This is the product surface for SEC-04; complete wiring here",
         wt="Existing Improvement", prio="P1", ac="Changing a role’s menu immediately hides those routes after refresh"),
    improve_rows(
        "No schema change unless Account/Pharmacy/Patient roles need new MenuMaster rows — seed those menus",
        "Restore GetMenuByRole on .NET 8; seed menus for Account and (later) Pharmacy",
        "Admin role pages remain; frontend consumes menu API instead of hard-coded LayoutMenuData for domain items",
        "QA: edit role menu → user of that role no longer sees the item",
    ),
)


# =============================================================================
# PHASE 2 — Doctor clinical workspace (PDF §5.3)
# =============================================================================
PN2 = "Phase 2 — Doctor clinical workspace (PDF §5.3) — complete the existing heart of the product"

add(
    dict(phase=2, pn=PN2, wbs="2.1", mid="CLN-01",
         main="Clinical workspace — single screen for the entire consultation",
         module="Patient Board", pdf="§5.3", feat="Clinical workspace",
         deliv="Single screen for the entire consultation",
         surface="Doctor Web Portal", as_is="PatientBoard.js (~13.5k) is the workspace; ~90% — missing payment/tele/eRx header status",
         to_be="Keep one screen; add header chips for payment, tele session, signed eRx (chips light up as later phases land)",
         wt="Existing Improvement", prio="P1", ac="Doctor still opens one board per patient; new chips do not break tabs"),
    [
        {"sub": "UI: Patient Board header shows appointment, visit type, consult mode placeholders (wired in Phases 4 and 6)",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "Do not split the board into multiple apps; mobile doctor app will NOT get case-taking (PDF §3)",
         "layer": "Architecture", "eff": "S", "wt": "Existing"},
        {"sub": "Regression: open case from dashboard still remounts via PatientBoardRoute",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=2, pn=PN2, wbs="2.2", mid="CLN-02",
         main="Manual case taking — structured case taking (keep)",
         module="Patient Board", pdf="§5.3", feat="Manual case taking",
         deliv="Structured case taking",
         surface="Doctor Web Portal", as_is="Classic case-taking tabs live",
         to_be="Regression only; do not rebuild",
         wt="Existing", prio="P2", ac="Manual case taking still saves/restores"),
    exist_keep(
        "Regression: create case, add rubrics via manual path, repertorize",
        "ACL: only treating doctor (not Reception) can run full case taking",
        "Confirm clipboard/repertorization APIs still on the correct host",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.3", mid="CLN-03",
         main="Body parts & intensity selection (keep)",
         module="Patient Board", pdf="§5.3", feat="Body parts & intensity selection",
         deliv="Select affected areas and severity",
         surface="Doctor Web Portal", as_is="Body Parts tab + intensity chips live",
         to_be="Regression; intensity master from Phase 1 still drives chips",
         wt="Existing", prio="P2", ac="Selecting a body part + intensity still adds rubrics"),
    exist_keep(
        "Regression: body part → rubrics; intensity chip",
        "ACL: doctor-only",
        "API host unchanged",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.4", mid="CLN-04",
         main="Clinical questions — guided questions by section, group and sub-group (keep)",
         module="Patient Board", pdf="§5.3", feat="Clinical questions",
         deliv="Guided questions by section, group and sub-group",
         surface="Doctor Web Portal", as_is="Questions tab loads Existance taxonomy live",
         to_be="Regression",
         wt="Existing", prio="P2", ac="Section → group → subgroup → keywords → rubrics still works"),
    exist_keep(
        "Regression: full question drill-down",
        "ACL doctor-only",
        "API host unchanged",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.5", mid="CLN-05",
         main="Diagnosis, therapeutics & pattern search (keep)",
         module="Patient Board", pdf="§5.3", feat="Diagnosis, therapeutics & pattern search",
         deliv="Search diagnoses and linked therapeutics",
         surface="Doctor Web Portal", as_is="Clinical Pattern tab live (13 keyword tabs + therapeutics)",
         to_be="Regression",
         wt="Existing", prio="P2", ac="Diagnosis search still returns linked rubrics and therapeutics"),
    exist_keep(
        "Regression: diagnosis search + therapeutics",
        "ACL doctor-only",
        "API host unchanged",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.6", mid="CLN-06",
         main="Repertory rubric search & selection (keep)",
         module="Patient Board", pdf="§5.3", feat="Repertory rubric search & selection",
         deliv="Search and select from the full repertory",
         surface="Doctor Web Portal", as_is="Repertory tree + search live",
         to_be="Regression; small UX polish allowed but not a rewrite",
         wt="Existing", prio="P2", ac="Search still adds to clipboard"),
    exist_keep(
        "Regression: tree + search + add to clipboard",
        "ACL doctor-only",
        "Confirm subsection search host (newer vs classic) documented",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.7", mid="CLN-07",
         main="Manage selected rubrics — review, edit and remove (remaining ~10%)",
         module="Patient Board", pdf="§5.3", feat="Manage selected rubrics",
         deliv="Review, edit and remove selected rubrics",
         surface="Doctor Web Portal", as_is="Clipboard exists; confirm-delete / intensity-edit incomplete",
         to_be="Confirm delete; edit intensity on selected rubric; no accidental wipe",
         wt="Existing Improvement", prio="P1", ac="Doctor can remove one rubric without losing the rest"),
    improve_rows(
        "No new table — clipboard already persisted via board backup / repertorization payload",
        "If intensity-on-clipboard is client-only, persist it on backup payload",
        "Clipboard UX: confirm delete; intensity edit; count badge",
        "QA: remove, undo-not-required, backup restore still has remaining rubrics",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.8", mid="CLN-08",
         main="Audio case taking — record and receive suggested rubrics (keep engine; do not rebuild)",
         module="Audio AI", pdf="§5.3", feat="Audio case taking",
         deliv="Record the consultation and receive suggested rubrics automatically",
         surface="Doctor Web Portal", as_is="Whisper + rubric engines + CaseTaking UI live on .NET 8",
         to_be="Keep; follow existing AUDIO_CASE_TAKING_* docs; accuracy is the next task",
         wt="Existing", prio="P1", ac="Record/upload still produces suggested rubrics for approve/reject"),
    exist_keep(
        "Regression: record, poll, approve/reject suggested rubrics",
        "Consent: audio consent continues to write AudioCaseConsentLog (later also ConsentRecord)",
        "Confirm audio queue/worker still healthy",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.9", mid="CLN-09",
         main="Improved rubric accuracy — enhanced accuracy rules for audio-based rubric suggestions",
         module="Audio AI", pdf="§5.3", feat="Improved rubric accuracy",
         deliv="Enhanced accuracy rules for audio-based rubric suggestions",
         surface="Doctor Web Portal + Admin intelligence", as_is="Engine + admin metaphors/aliases exist; accuracy rules still in progress (~50%)",
         to_be="Complete accuracy rules per AUDIO_CASE_TAKING_AI_ENGINE docs; metaphor/alias QA",
         wt="Existing Improvement", prio="P1", ac="Benchmark set meets agreed accuracy gate before calling this done"),
    [
        {"sub": "Follow AUDIO_CASE_TAKING_* docs: accuracy rules, aliases, metaphors, benchmark",
         "layer": "Backend", "be": Y, "eff": "L", "wt": "Existing Improvement"},
        {"sub": "Admin: metaphor/alias approve-reject remaining UX",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "QA: golden transcripts → expected rubric IDs within tolerance",
         "layer": "QA", "be": Y, "eff": "M", "wt": "Regression"},
    ],
)

add(
    dict(phase=2, pn=PN2, wbs="2.10", mid="CLN-10",
         main="Materia medica reference during consultation (remaining ~10%)",
         module="Patient Board", pdf="§5.3", feat="Materia medica reference",
         deliv="Browse materia medica during the consultation",
         surface="Doctor Web Portal", as_is="MM tab + repertorize differential accordion ~90%",
         to_be="Faster head navigation during case; do not rebuild MM masters",
         wt="Existing Improvement", prio="P2", ac="Doctor can open a remedy MM without leaving the board"),
    improve_rows(
        "No schema change",
        "Existing MM APIs — add a lightweight get-by-remedy if chatty",
        "MM tab: faster head navigation / keep place in case",
        "QA: open MM for a repertorized remedy",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.11", mid="CLN-11",
         main="Allopathic drug side-effect lookup (keep)",
         module="Patient Board", pdf="§5.3", feat="Allopathic drug side-effect lookup",
         deliv="Check side effects of conventional medicines",
         surface="Doctor Web Portal", as_is="Adverse Effect tab live",
         to_be="Regression",
         wt="Existing", prio="P2", ac="Lookup still shows serious/other/adverse"),
    exist_keep(
        "Regression: search drug → side effects",
        "ACL doctor-only",
        "API host unchanged",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.12", mid="CLN-12",
         main="Repertorization & elimination (remaining ~10%)",
         module="Patient Board", pdf="§5.3", feat="Repertorization & elimination",
         deliv="Full repertorization with elimination",
         surface="Doctor Web Portal", as_is="Common/uncommon/elimination live ~90%",
         to_be="Stability polish; Center of Gravity is a NEW panel on the same rubrics (next task)",
         wt="Existing Improvement", prio="P1", ac="Elimination still ranks remedies from selected rubrics"),
    improve_rows(
        "No schema change",
        "Keep existing repertorization APIs; fix any known scoring edge cases",
        "Repertorize tab stability; then COG sits beside it",
        "QA: known case → expected top remedies",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.13", mid="CLN-13",
         main="Center of Gravity — new analysis identifying the central remedy of the case, with reasoning shown",
         module="Repertorization", pdf="§5.3", feat="Center of Gravity",
         deliv="New analysis identifying the central remedy of the case, with reasoning shown",
         surface="Doctor Web Portal", as_is="Does not exist",
         to_be="POST /api/Repertorization/CenterOfGravity from selected rubric ids + intensities; UI panel with explainability",
         wt="New", prio="P1", ac="COG returns ranked central remedies plus contribution reasons the doctor can read"),
    new_full(
        "No mandatory new table (algorithm on selected rubrics); optional CogRun log for support",
        "POST /api/Repertorization/CenterOfGravity — input rubric ids + intensities; output ranked remedies + reasons",
        "Patient Board → Repertorize tab → Center of Gravity sub-panel (toggle vs classic elimination)",
        None,
        "QA: same clipboard as classic repertorize; COG reasons cite contributing rubrics; empty clipboard → clear empty state",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.14", mid="CLN-14",
         main="Lab orders & lab entries (keep; later also sit on eRx)",
         module="Patient Board", pdf="§5.3 + §11", feat="Lab orders & lab entries",
         deliv="Order tests and record results / Tests ordered alongside the prescription",
         surface="Doctor Web Portal", as_is="Lab order + result entry on prescription modal live",
         to_be="Keep; when eRx splits notes from Rx, labs stay on the prescription/eRx side",
         wt="Existing", prio="P2", ac="Order + result still save against the appointment"),
    exist_keep(
        "Regression: order lab from catalogue; enter result",
        "ACL doctor-only",
        "API: classic PatientLab still authoritative until ported",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.15", mid="CLN-15",
         main="Appointment history notes (keep; split from eRx in Phase 10)",
         module="Patient Board", pdf="§5.3", feat="Appointment history notes",
         deliv="Notes recorded against each visit",
         surface="Doctor Web Portal", as_is="AppointmentHistoryNote APIs live (classic); currently mixed into Rx modal tabs",
         to_be="Keep save/get; Phase 10 moves UX outside Rx modal and adds NoteType / IsErxExcluded",
         wt="Existing", prio="P1", ac="Notes still save against PatientAppId"),
    exist_keep(
        "Regression: add/list notes on a visit",
        "ACL doctor-only",
        "Do not dual-write; classic notes stay until Phase 10 extends them",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.16", mid="CLN-16",
         main="Complaints & case details — wire Patient Board to existing SaveComplaints / SaveCaseDetails (~20% today)",
         module="Patient Board", pdf="§5.3", feat="Complaints & case details",
         deliv="Capture and store complaints and case details",
         surface="Doctor Web Portal", as_is="Classic SaveComplaints / SaveCaseDetails APIs exist; Patient Board does not call them",
         to_be="Call the existing APIs from the board; Reception case-paper (Phase 5) uses a subset",
         wt="Existing Improvement", prio="P1", ac="Complaints and case details persist and reload on the board"),
    improve_rows(
        "Use existing tables — no new schema unless fields are missing vs PDF",
        "Keep classic SaveComplaints / SaveCaseDetails; add GET if missing",
        "Patient Board: complaints + case details form that actually POSTs",
        "QA: save, reload, second doctor cannot read",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.17", mid="CLN-17",
         main="Patient back history — complete history of visits, notes, prescriptions and payments",
         module="Patient Board", pdf="§5.3", feat="Patient back history",
         deliv="Complete history of visits, notes, prescriptions and payments",
         surface="Doctor Web Portal", as_is="History panel ~70%; no payment line (payments do not exist yet)",
         to_be="Timeline: visits, notes, eRx, labs; payment row added when Phase 6 lands (placeholder now)",
         wt="Existing Improvement", prio="P1", ac="Doctor sees chronological visits with notes and Rx; payment appears after Phase 6"),
    improve_rows(
        "No new table — compose from Appointment + notes + Rx + (later) ConsultPayment",
        "GET patient timeline DTO including payment status when column exists",
        "History panel: date-ordered visits; open past eRx; payment badge placeholder",
        "QA: multi-visit patient; empty new patient",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.18", mid="CLN-18",
         main="Export case data — download the full case in Excel or PDF",
         module="Patient Board", pdf="§5.3", feat="Export case data",
         deliv="Download the full case in Excel or PDF",
         surface="Doctor Web Portal", as_is="Newer ExportCasesToExcel exists; no Patient Board export UI; no PDF",
         to_be="Toolbar Excel + PDF using existing Excel API + new PDF renderer; exclude private notes from patient-facing PDF (eRx print is Phase 10)",
         wt="Existing Improvement", prio="P1", ac="Doctor downloads Excel and PDF of the case they own"),
    [
        {"sub": "Reuse ExportCasesToExcel; add PDF generation of the clinical case (doctor copy, includes notes)",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Patient Board toolbar: Export Excel / Export PDF",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: export matches on-screen rubrics + Rx; unauthorised doctor 403",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=2, pn=PN2, wbs="2.19", mid="CLN-19",
         main="Session restore — continue exactly where the doctor left off, across multiple patients (keep)",
         module="Patient Board", pdf="§5.3", feat="Session restore",
         deliv="Continue exactly where the doctor left off, across multiple patients",
         surface="Doctor Web Portal", as_is="Up to 5 in-memory sessions + DoctorPatientBoardBackup live",
         to_be="Regression; header stack remains",
         wt="Existing", prio="P2", ac="Kill browser, login, restore last work"),
    exist_keep(
        "Regression: backup, restore, 5-session stack",
        "ACL: backup scoped to DoctorId",
        "API: PatientBoardBackup on .NET 8 unchanged",
    ),
)

add(
    dict(phase=2, pn=PN2, wbs="2.20", mid="CLN-20",
         main="3D anatomy viewer — interactive 3D body model linked to rubric search (remaining ~50%)",
         module="Anatomy", pdf="§5.3", feat="3D anatomy viewer",
         deliv="Interactive 3D body model linked to rubric search",
         surface="Doctor Web Portal", as_is="three.js viewer + hotspot→search live; incomplete mesh coverage / UX",
         to_be="Complete GLB/hotspot coverage; mobile fallback later; admin 3D masters already in Phase 1",
         wt="Existing Improvement", prio="P2", ac="Every admin hotspot on the shipped models opens the correct rubric search"),
    improve_rows(
        "No new tables — Mesh/Section/Hotspot masters already exist",
        "Fix hotspot mapping gaps on newer threeDBodyPart APIs",
        "Viewer UX: incomplete regions, loading, fallback message",
        "QA: click hotspot → subsection search results",
    ),
)


# =============================================================================
# PHASE 3 — Doctor daily operations (PDF §5.1)
# =============================================================================
PN3 = "Phase 3 — Doctor dashboard & daily operations (PDF §5.1)"

add(
    dict(phase=3, pn=PN3, wbs="3.1", mid="DOC-01",
         main="Doctor home dashboard — complete overview of the working day (remaining ~30%)",
         module="Doctor Dashboard", pdf="§5.1", feat="Doctor home dashboard",
         deliv="Complete overview of the working day",
         surface="Doctor Web Portal", as_is="Widgets + BestSellingProducts ~70%; missing online toggle, tele queue, payment badges, follow-up due",
         to_be="Extend, do not replace; later phases light up tele/payment chips",
         wt="Existing Improvement", prio="P1", ac="Doctor sees today’s buckets plus placeholders for tele/payment that become live in Phases 6 and 11"),
    improve_rows(
        "No new dashboard table",
        "Extend dashboard DTOs with paymentStatus / tele flags when columns exist (nullable until Phase 6/11)",
        "Dashboard chrome: keep counts; add slots for Online toggle, tele queue, unpaid badge",
        "QA: existing buckets still populate; reception still shares layout until Phase 5 reception chrome",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.2", mid="DOC-02",
         main="Patient counts & waiting list — live counts by status with quick actions (keep; extend paid/unpaid)",
         module="Doctor Dashboard", pdf="§5.1", feat="Patient counts & waiting list",
         deliv="Live counts by status with quick actions",
         surface="Doctor Web Portal", as_is="Status buckets WAITING, WALK-IN, NOT ARRIVED, E-CONSULT, REMAINING, COMPLETED live",
         to_be="Keep; add CANCELLED in Phase 4; paid/unpaid badge in Phase 6",
         wt="Existing", prio="P1", ac="Counts still match appointment list"),
    exist_keep(
        "Regression: each bucket count and click-through",
        "Reception sharing this view is OK until Phase 5 splits chrome",
        "API: existing dashboard count endpoints",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.3", mid="DOC-03",
         main="Patient search, list & open case (remaining ~10%)",
         module="Doctor Dashboard", pdf="§5.1", feat="Patient search, list & open case",
         deliv="Find any patient and open their case instantly",
         surface="Doctor Web Portal", as_is="Search/list/open ~90%; missing last visit, unpaid, family filters",
         to_be="Add last-visit column; unpaid/family after those modules exist",
         wt="Existing Improvement", prio="P1", ac="Search still opens Patient Board; last visit visible"),
    improve_rows(
        "Use existing Patient + Appointment — add DTO fields",
        "Search DTO includes lastVisitAt",
        "List columns: last visit; later payment/family",
        "QA: search by name/mobile; open case",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.4", mid="DOC-04",
         main="Add new patient — register a patient in seconds (keep)",
         module="Patients", pdf="§5.1", feat="Add new patient",
         deliv="Register a patient in seconds",
         surface="Doctor Web Portal", as_is="Create patient live",
         to_be="Regression; public self-registration is a different Patient identity in Phase 9",
         wt="Existing", prio="P1", ac="Doctor can still create a clinic patient"),
    exist_keep(
        "Regression: add patient validation and save",
        "ACL: doctor/reception of that clinic only",
        "API unchanged",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.5", mid="DOC-05",
         main="Import patients — bring existing patient records in bulk (keep)",
         module="Patients", pdf="§5.1", feat="Import patients",
         deliv="Bring existing patient records in bulk",
         surface="Doctor Web Portal", as_is="Import live",
         to_be="Regression",
         wt="Existing", prio="P2", ac="Import still creates patients for this doctor"),
    exist_keep(
        "Regression: template + import + error rows",
        "ACL doctor-only (or reception if already allowed)",
        "API unchanged",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.6", mid="DOC-06",
         main="Export patients — download patient data in Excel, CSV or PDF (keep)",
         module="Patients", pdf="§5.1", feat="Export patients",
         deliv="Download patient data in Excel, CSV or PDF",
         surface="Doctor Web Portal", as_is="Export live (confirm PDF vs Excel/CSV in QA)",
         to_be="If PDF is missing, add it as Existing Improvement; do not skip the PDF wording in the PDF",
         wt="Existing Improvement", prio="P2", ac="Excel, CSV and PDF each download this doctor’s patients only"),
    [
        {"sub": "Audit current export formats; add PDF if only Excel/CSV exist today",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "UI: Excel / CSV / PDF actions labelled as in the delivery document",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: another doctor cannot export this list",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=3, pn=PN3, wbs="3.7", mid="DOC-07",
         main="Recent activity panel (keep)",
         module="Doctor Dashboard", pdf="§5.1", feat="Recent activity panel",
         deliv="Everything that happened recently, at a glance",
         surface="Doctor Web Portal", as_is="Recent activity live",
         to_be="Regression; later include payment/tele events when those exist",
         wt="Existing", prio="P2", ac="Panel still lists recent actions"),
    exist_keep(
        "Regression: activity loads",
        "ACL scoped to doctor",
        "API unchanged",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.8", mid="DOC-08",
         main="Patient statistics charts (keep; foundation for clinic performance in Phase 18)",
         module="Doctor Dashboard", pdf="§5.1", feat="Patient statistics charts",
         deliv="Visual practice statistics",
         surface="Doctor Web Portal", as_is="GetPatientStats / GetPatientStatsCharts + ApexCharts live",
         to_be="Keep; clinic performance report extends this in Phase 18",
         wt="Existing", prio="P2", ac="Charts still render for the doctor"),
    exist_keep(
        "Regression: stats charts date range",
        "ACL doctor-only",
        "API unchanged",
    ),
)

add(
    dict(phase=3, pn=PN3, wbs="3.9", mid="DOC-09",
         main="Manage reception staff — add, edit and remove reception users (backend done, UI absent)",
         module="Staff", pdf="§5.1", feat="Manage reception staff",
         deliv="Add, edit and remove reception users",
         surface="Doctor Web Portal", as_is=".NET 8 api/ReceptionStaff CRUD exists; no SPA page",
         to_be="List/add/edit/disable at /doctor/reception-staff; doctor owns the staff record",
         wt="Existing Improvement", prio="P1", ac="Doctor can create a reception user who logs in to that DoctorId"),
    [
        {"sub": "No new API — authorise that the doctor owns the staff record on existing CRUD",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "Build /doctor/reception-staff list, add, edit, disable pages; menu entry in doctor chrome",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "QA: reception user cannot manage other doctors’ staff; disable prevents login",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=3, pn=PN3, wbs="3.10", mid="DOC-10",
         main="Doctor profile — clinic details, consultation fee, photo, qualifications, working hours, bank details",
         module="Doctor Profile", pdf="§5.1", feat="Doctor profile",
         deliv="Clinic details, consultation fee, photo, qualifications, working hours, bank details",
         surface="Doctor Web Portal", as_is="/profile is Velzon first_name only (~20%); credentials sit on Doctor entity; no fee/KYC UI",
         to_be="Real doctor profile: clinic, photo, qualifications, consult fee (feeds Phase 6), hours (feeds schedule), bank KYC (feeds Account payees)",
         wt="Existing Improvement", prio="P0", ac="Doctor can save clinic, photo, qualifications, fee, hours, bank details; verification status visible"),
    improve_rows(
        "Extend Doctor / add DoctorPayeeKyc fields (bank, PAN) used later by Account; ConsultFeeConfig may wait until Phase 6 — store fee on profile now if simpler",
        "GET/PUT /api/Profile/Me (doctor DTO); POST photo multipart; do not allow role escalation",
        "Replace Velzon /profile for doctor role with clinic/fee/photo/qualifications/hours/bank tabs",
        "QA: reception cannot edit doctor bank details; photo displays on public profile later",
    ),
)


# =============================================================================
# PHASE 4 — Appointments & scheduling (PDF §5.2)
# =============================================================================
PN4 = "Phase 4 — Appointments & scheduling (PDF §5.2) — product completeness before public booking"

add(
    dict(phase=4, pn=PN4, wbs="4.1", mid="APT-01",
         main="Appointment list & waiting patients — full day view with live status (keep; later payment/tele)",
         module="Appointments", pdf="§5.2", feat="Appointment list & waiting patients",
         deliv="Full day view with live status",
         surface="Doctor Web Portal", as_is="Day list + status live",
         to_be="Keep; reschedule/cancel/payment attach to this list",
         wt="Existing", prio="P1", ac="Day view still lists today’s appointments"),
    exist_keep(
        "Regression: day list, status colour, open patient",
        "ACL doctor/reception of that DoctorId",
        "API unchanged",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.2", mid="APT-02",
         main="Create new appointment — persist visit type, consultation mode and fee",
         module="Appointments", pdf="§5.2", feat="Create new appointment",
         deliv="Book with visit type, consultation mode and fee",
         surface="Doctor Web Portal", as_is="Create appointment ~90%; Formik modal has no VisitType / ConsultMode / fee",
         to_be="VisitType First/Follow-up; ConsultMode In-clinic / E-Consult; fee display from doctor profile / ConsultFeeConfig",
         wt="Existing Improvement", prio="P0", ac="New appointments persist visit type, mode and fee; filters/analytics can use them"),
    improve_rows(
        "DB: PatientAppointment.VisitType, ConsultMode; migrate existing rows VisitType=First (do not infer retrospectively)",
        "API: create/update appointment accepts the new fields; PATCH VisitType",
        "Doctor/Reception new-appointment modal: visit type, consult mode, fee (placeholder until Phase 6 config)",
        "QA: staff-created appointments carry the same fields public booking will use in Phase 9",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.3", mid="APT-03",
         main="Appointment status update — move patients through the visit stages (keep)",
         module="Appointments", pdf="§5.2", feat="Appointment status update",
         deliv="Move patients through the visit stages",
         surface="Doctor Web Portal", as_is="Status transitions live",
         to_be="Keep; add CANCELLED in APT-06; do not reuse DeleteStatus as cancel",
         wt="Existing", prio="P1", ac="Status still updates and buckets refresh"),
    exist_keep(
        "Regression: each status transition",
        "ACL doctor/reception",
        "API unchanged",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.4", mid="APT-04",
         main="Appointment time update — adjust timings as the day runs (keep; distinct from formal reschedule)",
         module="Appointments", pdf="§5.2", feat="Appointment time update",
         deliv="Adjust timings as the day runs",
         surface="Doctor Web Portal", as_is="UpdateAppointmentTime live — changes a time only; no audit/notify",
         to_be="Keep same-day time tweak; formal reschedule (old→new, notify, conflict) is APT-05",
         wt="Existing", prio="P1", ac="Same-day time change still works; does not silently replace Reschedule"),
    exist_keep(
        "Regression: UpdateAppointmentTime on dashboard",
        "Document in UI that Reschedule is the patient-notified path",
        "API: keep UpdateAppointmentTime; Reschedule wraps it",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.5", mid="APT-05",
         main="Reschedule appointment — move to a new slot with automatic patient notification (NEW — not a rename of UpdateAppointmentTime)",
         module="Appointments", pdf="§5.2 + §2.3 + §6", feat="Reschedule appointment",
         deliv="Move to a new slot with automatic patient notification / Move to another slot without calling the clinic",
         surface="Doctor Web + Reception + later Patient", as_is="Time can be edited; no formal Reschedule (audit, notify, old→new confirm)",
         to_be="POST RescheduleAppointment; AppointmentChangeLog; conflict 409 with alternatives; SMS/WhatsApp fan-out (notify failure must not roll back)",
         wt="New", prio="P0", ac="Every reschedule is in the change log; patient notified; old slot released; reception uses the same API"),
    new_full(
        "DB: AppointmentChangeLog (PatientAppId, Action, OldValue, NewValue, ByUserId, ByRole, At) + indexes",
        "API: POST /api/PatientAppointment/RescheduleAppointment — validate schedule, bookings, not in the past; 409 with alternatives; wrap UpdateAppointmentTime; GET ChangeLog/{id}",
        "UI: RescheduleModal shared by Doctor and Reception — old date/time + new slot picker; reason optional",
        "Notifications: SMS + WhatsApp + (later) push — old→new; must not block the transaction if notify fails",
        "QA: slot taken → 409; patient already booked elsewhere; across a schedule change; reception authorised for own DoctorId only",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.6", mid="APT-06",
         main="Cancel appointment — reason, free the slot, trigger refund where applicable (NEW CANCELLED status — do NOT reuse DeleteStatus)",
         module="Appointments", pdf="§5.2 + §2.3 + §6", feat="Cancel appointment",
         deliv="Cancel with reason, free the slot, trigger refund where applicable / Cancel with reason, refund handled per policy",
         surface="Doctor Web + Reception + later Patient", as_is="No CANCELLED; DeleteStatus is soft-delete only",
         to_be="CANCELLED + reason fields; free slot; waitlist hook; if PAID evaluate refund policy (Phase 6/7)",
         wt="New", prio="P0", ac="Cancelled appointments leave active buckets; slot immediately bookable; refund path triggered rather than silently keeping money"),
    new_full(
        "DB: add CANCELLED to status set; CancelReasonCode, CancelReasonText, CancelledBy, CancelledAt; DeleteStatus remains administrative",
        "API: POST /api/PatientAppointment/CancelAppointment — reason required (enum; text if Other); free slot; offer waitlist stub; if PAID enqueue refund policy",
        "UI: CancelAppointmentModal shared by Doctor and Reception; dashboard Cancelled filter",
        "Waitlist worker stub: on cancel, offer freed slot (full waitlist capture is Phase 9 GST waitlist)",
        "QA: reason mandatory; all three roles (later Patient) cancel with correct authorisation; freed slot bookable",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.7", mid="APT-07",
         main="Daily schedule setup — define working days, hours and break times (promote modal to first-class screen)",
         module="Scheduling", pdf="§5.2", feat="Daily schedule setup",
         deliv="Define working days, hours and break times",
         surface="Doctor Web Portal", as_is="DailyScheduleSetupModal + GetDailySchedule / SaveDailySchedule exist; not a first-class screen (draft marked 0%)",
         to_be="Promote to /doctor/schedule with week grid and copy-day; reuse existing APIs — do not rebuild",
         wt="Existing Improvement", prio="P0", ac="Public booking slots (Phase 9) respect the saved schedule"),
    [
        {"sub": "Promote DailyScheduleSetupModal to /doctor/schedule; week grid; copy-day; reuse Get/SaveDailySchedule",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "No new API unless validation (breaks, overlapping hours) is missing — add server-side validation",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: save week; public slot generator (Phase 9) uses this data; reception read-only is Phase 5",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=4, pn=PN4, wbs="4.8", mid="APT-08",
         main="Appointment slot grid — visual slot-by-slot view of the day (keep; public booking reuses this)",
         module="Scheduling", pdf="§5.2", feat="Appointment slot grid",
         deliv="Visual slot-by-slot view of the day",
         surface="Doctor Web Portal", as_is="Slot picker / grid live",
         to_be="Keep; Phase 9 public APIs reuse the same slot engine",
         wt="Existing", prio="P0", ac="Grid matches saved daily schedule"),
    exist_keep(
        "Regression: slot grid vs schedule",
        "ACL doctor/reception",
        "Document GetAppointmentSlots as the public-booking engine — do not fork a second slot calculator",
    ),
)

add(
    dict(phase=4, pn=PN4, wbs="4.9", mid="APT-09",
         main="Consultation payment status on every appointment (schema now; values live in Phase 6)",
         module="Appointments", pdf="§5.2 + §9", feat="Consultation payment status",
         deliv="See paid, unpaid or pay-at-clinic on every appointment / Doctor, reception and patient all see the same payment status",
         surface="Doctor Web Portal", as_is="No PaymentStatus on appointment",
         to_be="Add PaymentStatus UNPAID / PENDING / PAID / FAILED / PAY_AT_CLINIC / REFUNDED; UI badge; source of truth is webhook in Phase 6 not the client screen",
         wt="New", prio="P0", ac="Every appointment row can show a payment badge; until Phase 6 all show UNPAID/PAY_AT_CLINIC as configured"),
    [
        {"sub": "DB: PatientAppointment.PaymentStatus, PaymentMethod, PayAtClinicAllowed; default UNPAID",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API: include PaymentStatus on all appointment DTOs (doctor, reception, later public/patient)",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "UI: Paid / Unpaid / Pay-at-clinic / Failed badges on doctor appointment list",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: same appointment shows the same badge on doctor and reception",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 5 — Reception portal (PDF §6)
# =============================================================================
PN5 = "Phase 5 — Reception portal (PDF §6)"

add(
    dict(phase=5, pn=PN5, wbs="5.1", mid="REC-01",
         main="Reception login — secure staff access (keep)",
         module="Reception", pdf="§6", feat="Reception login",
         deliv="Secure staff access",
         surface="Reception Portal", as_is="Reception logs in via Account; JWT includes DoctorID; uses doctor dashboard layout",
         to_be="Keep login; Phase 0 hashing applies; later dedicated reception chrome",
         wt="Existing", prio="P1", ac="Reception user still logs in and lands on front-desk home"),
    exist_keep(
        "Regression: reception login with DoctorID claim",
        "Disabled staff cannot login (after DOC-09)",
        "Same Login API as doctor",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.2", mid="REC-02",
         main="Reception profile — manage own profile details (0% today)",
         module="Reception", pdf="§6", feat="Reception profile",
         deliv="Manage own profile details",
         surface="Reception Portal", as_is="No reception-specific profile; shared Velzon /profile",
         to_be="Name, mobile, attached doctor (read-only); cannot edit doctor bank/fee",
         wt="New", prio="P1", ac="Reception can update own name/mobile; cannot change DoctorId"),
    new_full(
        "Reuse Profile/Me with role=Reception DTO (no new table if UserMaster suffices)",
        "GET/PUT profile fields allowed for reception; ignore clinic fee/bank",
        "/profile reception variant",
        None,
        "QA: cannot edit doctor qualifications or fee",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.3", mid="REC-03",
         main="Reception home dashboard — purpose-built front-desk dashboard (not only hidden doctor actions)",
         module="Reception", pdf="§6", feat="Reception home dashboard",
         deliv="Purpose-built front-desk dashboard",
         surface="Reception Portal", as_is="Same /doctordashboard with some clinical actions disabled ~90%",
         to_be="Reception-first chrome: queue, collect pay, case paper, no repertory/Rx; optional /reception home route",
         wt="Existing Improvement", prio="P1", ac="Reception never sees repertorize/audio/COG; sees queue and collect-pay"),
    [
        {"sub": "Route /reception (or role-based chrome on existing dashboard) hiding clinical tabs",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Quick actions: new patient, new appointment, collect payment, case paper, schedule",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: deep-link to Patient Board full clinical URL is 403 for reception",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=5, pn=PN5, wbs="5.4", mid="REC-04",
         main="Create patient — register a walk-in patient quickly (keep)",
         module="Reception", pdf="§6", feat="Create patient",
         deliv="Register a walk-in patient quickly",
         surface="Reception Portal", as_is="Create patient live on shared dashboard",
         to_be="Keep on reception chrome",
         wt="Existing", prio="P1", ac="Walk-in patient created under the JWT DoctorId"),
    exist_keep(
        "Regression: reception create patient",
        "ACL: own doctor only",
        "Same create-patient API",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.5", mid="REC-05",
         main="Patient search & list — find any patient instantly (keep)",
         module="Reception", pdf="§6", feat="Patient search & list",
         deliv="Find any patient instantly",
         surface="Reception Portal", as_is="Search live",
         to_be="Keep; no opening full clinical board",
         wt="Existing", prio="P1", ac="Search returns this doctor’s patients only"),
    exist_keep(
        "Regression: search by name/mobile",
        "Opening a row goes to case paper or appointment, not full repertory",
        "API unchanged",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.6", mid="REC-06",
         main="Create & manage appointments — full appointment management (keep; plus reschedule/cancel from Phase 4)",
         module="Reception", pdf="§6", feat="Create & manage appointments",
         deliv="Full appointment management",
         surface="Reception Portal", as_is="Create/status/time live; reschedule/cancel missing until Phase 4 APIs",
         to_be="Enable RescheduleAppointment and CancelAppointment for JWT DoctorId; isReceptionUser must not block",
         wt="Existing Improvement", prio="P0", ac="Reception acts only for its doctor; permission tests pass"),
    [
        {"sub": "Authorise Reception for Reschedule and Cancel for its own doctor's appointments only",
         "layer": "API + SPA", "fe": Y, "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Reception UI uses the same Reschedule/Cancel modals as doctor",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: cannot cancel another clinic’s appointment",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=5, pn=PN5, wbs="5.7", mid="REC-07",
         main="New appointment — book on behalf of the patient (keep; same VisitType/ConsultMode/fee as APT-02)",
         module="Reception", pdf="§6", feat="New appointment",
         deliv="Book on behalf of the patient",
         surface="Reception Portal", as_is="New appointment modal live; fields incomplete until APT-02",
         to_be="Same modal fields as doctor so public booking later matches staff-created rows",
         wt="Existing Improvement", prio="P1", ac="Reception-created appointments have visit type, mode, fee"),
    [
        {"sub": "Share the extended new-appointment modal with reception chrome",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: in-clinic vs E-Consult both creatable from reception",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=5, pn=PN5, wbs="5.8", mid="REC-08",
         main="Waiting queue management — manage the live queue and call the next patient",
         module="Reception", pdf="§6", feat="Waiting queue management",
         deliv="Manage the live queue and call the next patient",
         surface="Reception Portal", as_is="Status buckets ~60%; no dedicated queue UX / call-next",
         to_be="Dedicated queue: walk-in vs tele; payment state; Call next; notify patient (later push)",
         wt="Existing Improvement", prio="P1", ac="Reception can call next; status moves; patient can be notified when Phase 12 exists"),
    improve_rows(
        "No new table required if status+order suffice; optional QueuePosition",
        "Optional CallNext API that sets status and timestamp",
        "Queue panel: order, wait time, paid/unpaid, Call next",
        "QA: two waiting patients; call next picks the right one",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.9", mid="REC-09",
         main="Appointment status update (reception) — update as patients arrive and are seen (keep)",
         module="Reception", pdf="§6", feat="Appointment status update",
         deliv="Update as patients arrive and are seen",
         surface="Reception Portal", as_is="Status update live",
         to_be="Keep on reception chrome",
         wt="Existing", prio="P1", ac="Arrived / seen transitions still work"),
    exist_keep(
        "Regression: reception status update",
        "ACL own doctor",
        "Same status API as doctor",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.10", mid="REC-10",
         main="Appointment time update (reception) — adjust timings through the day (keep)",
         module="Reception", pdf="§6", feat="Appointment time update",
         deliv="Adjust timings through the day",
         surface="Reception Portal", as_is="Time update live",
         to_be="Keep; formal reschedule remains REC-06",
         wt="Existing", prio="P2", ac="Same-day time tweak from reception works"),
    exist_keep(
        "Regression: reception UpdateAppointmentTime",
        "ACL own doctor",
        "Same API",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.11", mid="REC-11",
         main="View doctor schedule — see the doctor’s working schedule (read-only)",
         module="Reception", pdf="§6", feat="View doctor schedule",
         deliv="See the doctor’s working schedule",
         surface="Reception Portal", as_is="GetDailySchedule exists; no reception view",
         to_be="Read-only /reception/schedule using DoctorID from JWT",
         wt="New", prio="P1", ac="Reception views hours; cannot edit clinical schedule unless authorised"),
    [
        {"sub": "Frontend only over existing GetDailySchedule — /reception/schedule read-only",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: edit buttons hidden; doctor can still edit on /doctor/schedule",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=5, pn=PN5, wbs="5.12", mid="REC-12",
         main="Log case paper / chief complaint — record the reason for visit before the consultation",
         module="Reception", pdf="§6", feat="Log case paper / chief complaint",
         deliv="Record the reason for visit before the consultation",
         surface="Reception Portal", as_is="Missing; SaveComplaints exists but board is doctor-only",
         to_be="Restricted case-paper form (chief complaint only — no repertory/Rx); audit reception author",
         wt="New", prio="P1", ac="Chief complaint visible to doctor on Patient Board; reception cannot repertorize"),
    new_full(
        "Reuse complaints table; add CreatedByRole=Reception",
        "POST case-paper subset of SaveComplaints; GET on board",
        "/reception/case-paper or limited board",
        None,
        "QA: doctor sees CC; reception 403 on repertory APIs",
    ),
)

add(
    dict(phase=5, pn=PN5, wbs="5.13", mid="REC-13",
         main="Collect consultation payment — cash, UPI, card or payment link and issue a receipt (API in Phase 6; UI shell now)",
         module="Reception", pdf="§6 + §9", feat="Collect consultation payment",
         deliv="Take cash, UPI, card or payment link and issue a receipt",
         surface="Reception Portal", as_is="Does not exist",
         to_be="UI modal ready; CollectAtReception API lands with payment spine in Phase 6 — do not fake a paid status here",
         wt="New", prio="P0", ac="Modal exists but cannot mark PAID until Phase 6 webhook/ledger path is live — or land UI in Phase 6 with PAY-06. Prefer implement fully in Phase 6; this task is the reception UX that Phase 6 must include",
         notes="Implement with PAY-02 in Phase 6; listed here so PDF §6 is not skipped in the Reception phase map"),
    [
        {"sub": "Specify receipt fields (amount, method, appointment, GST placeholder) for Phase 6 implementation",
         "layer": "Design", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "Do not write PaymentStatus=PAID from the client — Account/webhook is source of truth (Phase 6)",
         "layer": "Architecture", "be": Y, "eff": "S", "wt": "New"},
    ],
)


# =============================================================================
# PHASE 6 — Payment system (PDF §9) — money spine
# =============================================================================
PN6 = "Phase 6 — Payment system (PDF §9) — Homeocentrum holds every rupee"

add(
    dict(phase=6, pn=PN6, wbs="6.1", mid="PAY-01",
         main="Consultation fee configuration — each doctor’s fee set and displayed to patients",
         module="Payments", pdf="§9", feat="Consultation fee configuration",
         deliv="Each doctor’s fee set and displayed to patients",
         surface="Doctor profile + Public + Admin", as_is="No ConsultFeeConfig; doctor profile work in DOC-10 may have stored a fee field",
         to_be="ConsultFeeConfig (InClinic, EConsult, optional Instant surcharge, PayAtClinicEnabled, GST flags)",
         wt="New", prio="P0", ac="Fee shown to patients matches what checkout charges; admin/doctor can update"),
    new_full(
        "DB: ConsultFeeConfig (DoctorId, InClinicFee, TeleFee, InstantSurcharge, Currency, PayAtClinicEnabled, EffectiveFrom)",
        "API: GET public fee; PUT doctor/admin; history of changes in audit",
        "Doctor profile fee tab + later public doctor card",
        None,
        "QA: changing fee does not alter already-created unpaid orders without a rule (document the rule)",
    ),
)

add(
    dict(phase=6, pn=PN6, wbs="6.2", mid="PAY-02",
         main="Automatic payment verification — bank/webhook is source of truth, not the patient’s screen",
         module="Payments", pdf="§9", feat="Automatic payment verification",
         deliv="Payments confirmed by the bank, not by the patient’s screen — nothing is lost if the app or browser closes",
         surface="Shared", as_is="Razorpay checkout for S1 subscription only; no webhook; no signature verify on server",
         to_be="Port Razorpay to .NET 8; keys in config; POST /api/Payments/Webhook/Razorpay anonymous + signature verify; PaymentOrder table",
         wt="New Integration", prio="P0", ac="Kill browser after pay → webhook still marks PAID; client callback alone never marks PAID"),
    [
        {"sub": "DB: PaymentOrder, PaymentEvent (raw webhook), idempotency key on payment_id",
         "layer": "Database", "db": Y, "be": Y, "eff": "M", "wt": "New"},
        {"sub": "Port Order/GenerateOrderId pattern to .NET 8; move keys to configuration; do not commit secrets",
         "layer": "API", "be": Y, "api": Y, "int": Y, "eff": "L", "wt": "New Integration"},
        {"sub": "Webhook handler payment.captured / payment.failed; signature verify; never trust client-only success",
         "layer": "API", "be": Y, "api": Y, "int": Y, "eff": "M", "wt": "New Integration"},
        {"sub": "QA: duplicate webhook; delayed webhook; failed capture; app closed mid-checkout",
         "layer": "QA", "be": Y, "int": Y, "eff": "M", "wt": "Regression"},
    ],
)

add(
    dict(phase=6, pn=PN6, wbs="6.3", mid="PAY-03",
         main="Online consultation payment — patient pays by card, UPI, netbanking or wallet",
         module="Payments", pdf="§9 + §4 + §2.3", feat="Online consultation payment",
         deliv="Patient pays by card, UPI, netbanking or wallet / Pay the consultation fee at the time of booking / Pay the consultation fee inside the app",
         surface="Public web + later Patient app", as_is="Does not exist (only S1 SaaS)",
         to_be="CreateConsultOrder attached to PatientAppId; Razorpay Checkout; webhook sets PAID; do not overload PackageEntryDetail",
         wt="New", prio="P0", ac="Paid appointment unlocks tele queue rules; doctor sees PAID; Account sees ledger row (Phase 7)"),
    new_full(
        "DB: ConsultPayment linked to PaymentOrder + PatientAppId — NOT PackageEntryDetail",
        "API: CreateConsultOrder, Verify (optional), status by appointment; amount from ConsultFeeConfig",
        "Public /book/pay/:bookingId using Razorpay pattern from Widgets.js but consult order API",
        "Razorpay Checkout methods: card, UPI, netbanking, wallet as enabled on the account",
        "QA: success, failure, pending, double-click create order idempotent",
    ),
)

add(
    dict(phase=6, pn=PN6, wbs="6.4", mid="PAY-04",
         main="Reception payment collection — cash, UPI, card or payment link with printed receipt",
         module="Payments", pdf="§9 + §6", feat="Reception payment collection",
         deliv="Cash, UPI, card or payment link with printed receipt",
         surface="Reception Portal", as_is="Does not exist",
         to_be="CollectAtReception writes ledger S3; print receipt; payment-link uses same Razorpay order as online",
         wt="New", prio="P0", ac="Cash still creates an immutable ledger row; receipt printable; status matches doctor list"),
    new_full(
        "Reuse PaymentOrder + ConsultPayment with Method=CASH/UPI_OFFLINE/CARD_POS/PAY_LINK",
        "API: CollectAtReception; generate pay-link; print receipt payload",
        "Reception collect modal: amount, method, GST placeholder, print receipt",
        "Pay-link SMS/WhatsApp optional (Phase 12); must still ledger if notify fails",
        "QA: cash collection visible in clinic collection view after Phase 7; cannot mark PAID without this API",
    ),
)

add(
    dict(phase=6, pn=PN6, wbs="6.5", mid="PAY-05",
         main="Pay-at-clinic option — book now, pay at the clinic where the doctor allows it",
         module="Payments", pdf="§9", feat="Pay-at-clinic option",
         deliv="Book now, pay at the clinic where the doctor allows it",
         surface="Public + Doctor + Reception", as_is="Does not exist",
         to_be="PayAtClinicEnabled on fee config; appointment PaymentStatus=PAY_AT_CLINIC; tele queue may still require PAID (policy)",
         wt="New", prio="P1", ac="If doctor disables pay-at-clinic, public checkout requires online pay"),
    [
        {"sub": "Honor PayAtClinicEnabled on public create-booking and checkout",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "UI: public booking shows Pay at clinic vs Pay now; doctor list badge",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: doctor flag off → pay-at-clinic hidden; reception can still collect",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=6, pn=PN6, wbs="6.6", mid="PAY-06",
         main="Doctor subscription payment — buy and renew subscription packages (complete remaining ~30%)",
         module="Payments", pdf="§9", feat="Doctor subscription payment",
         deliv="Buy and renew subscription packages",
         surface="Doctor Web Portal", as_is="Checkout works ~70%; no webhook, refund, invoice, GST, Account ledger",
         to_be="Same webhook spine; write S1 ledger; invoice; failed pay UX; PackageEntryDetail shape UNCHANGED",
         wt="Existing Improvement", prio="P1", ac="SaaS pay verified by webhook; doctor plan flags update without trusting client callback alone"),
    [
        {"sub": "Reuse PAY-02 webhook for S1 orders; link PackageEntryDetail as S1 only",
         "layer": "API", "be": Y, "int": Y, "eff": "M", "wt": "Existing Improvement"},
        {"sub": "Widgets.js: handle failed pay; show invoice when Phase 7 invoice exists",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: renew, fail, webhook delay, last-5-days warning still works",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=6, pn=PN6, wbs="6.7", mid="PAY-07",
         main="Medicine payment — pay online or cash on delivery (API contract now; HomeoMeds Phase 14 implements orders)",
         module="Payments", pdf="§9 + §12", feat="Medicine payment",
         deliv="Pay online or choose cash on delivery / Both options supported, with status tracking",
         surface="Shared", as_is="Does not exist",
         to_be="PaymentOrder.Stream=MEDICINE; COD_PENDING until pharmacy confirms collection; do not mix with consult money",
         wt="New", prio="P1", ac="A medicine PaymentOrder cannot be summarised as consult GMV"),
    [
        {"sub": "Extend PaymentOrder with Stream=CONSULT|SUBSCRIPTION|MEDICINE so Phase 14 can attach without schema rewrite",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "API stub: CreateMedicineOrder payment + COD status enum — implemented fully in Phase 14",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: stream isolation query (consult vs medicine totals differ)",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression", "prio": "P2"},
    ],
)

add(
    dict(phase=6, pn=PN6, wbs="6.8", mid="PAY-08",
         main="Payment status everywhere — doctor, reception and patient all see the same payment status",
         module="Payments", pdf="§9", feat="Payment status everywhere",
         deliv="Doctor, reception and patient all see the same payment status",
         surface="All", as_is="No shared status",
         to_be="Single PaymentStatus on appointment/order; all DTOs read it; never a second client-only flag",
         wt="New", prio="P0", ac="Doctor list, reception queue, public booking confirmation, later patient app show identical status for the same id"),
    [
        {"sub": "Single source: PaymentOrder/ConsultPayment → Appointment.PaymentStatus updated only by webhook or CollectAtReception",
         "layer": "API", "be": Y, "eff": "M", "wt": "New"},
        {"sub": "Surface badges: doctor, reception; public confirmation page; patient app in Phase 16 consumes same API",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: three roles, one appointment, same enum value",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=6, pn=PN6, wbs="6.9", mid="PAY-09",
         main="Refunds — processed and tracked against the original payment",
         module="Payments", pdf="§9 + §8 + §2.8", feat="Refunds",
         deliv="Processed and tracked against the original payment / Issue and track refunds",
         surface="Account + Patient", as_is="Does not exist",
         to_be="Refund entity against original PaymentOrder; Razorpay refund API; cancel-appointment hook; patient sees status",
         wt="New", prio="P0", ac="Refund cannot exceed captured amount; original payment remains visible; appointment status REFUNDED/partial as designed"),
    new_full(
        "DB: Refund (PaymentOrderId, Amount, Reason, Status, GatewayRefundId, ByUserId, At)",
        "API: create refund (Account); webhook refund.processed; cancel-appointment policy engine (full/partial/none)",
        "Account refund screen in Phase 7; patient payments list in Phase 16",
        "Razorpay refund API + webhook",
        "QA: double refund blocked; cash collection refund is ledger-only per PRE-02 policy",
    ),
)

add(
    dict(phase=6, pn=PN6, wbs="6.10", mid="PAY-10",
         main="Payment receipts & invoices — issued for every transaction",
         module="Payments", pdf="§9 + §8", feat="Payment receipts & invoices",
         deliv="Issued for every transaction / Invoice numbering and tax reporting",
         surface="All", as_is="No invoice entity",
         to_be="Invoice numbering series from PRE-02; PDF receipt; GST fields; email in Phase 12",
         wt="New", prio="P1", ac="Every PAID consult/subscription/medicine has a retrievable receipt number"),
    new_full(
        "DB: Invoice (Number, Series, Stream, PaymentOrderId, GstBreakup, PdfPath)",
        "API: get invoice by payment; generate PDF via SecureDocument",
        "Print/download on reception, doctor (subscription), later patient app",
        None,
        "QA: unique invoice numbers; cancelled invoice not reused",
    ),
)

add(
    dict(phase=6, pn=PN6, wbs="6.11", mid="PAY-11",
         main="Complete payment audit trail — every transaction traceable end to end by Homeocentrum",
         module="Payments", pdf="§9 + §17", feat="Complete audit trail",
         deliv="Every transaction traceable end to end by Homeocentrum",
         surface="Account", as_is="No webhook store",
         to_be="PaymentEvent raw payloads + AuditEvent + OtpAudit for payouts; Account can reconstruct without Razorpay dashboard",
         wt="New", prio="P0", ac="Given a PatientAppId, Account retrieves order, events, invoice, refunds"),
    [
        {"sub": "Persist raw webhook JSON; correlation id on Appointment and Ledger",
         "layer": "Database", "db": Y, "be": Y, "eff": "S", "wt": "New"},
        {"sub": "Account query API: trail by appointment / payment / doctor (Phase 7 UI)",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: missing webhook still shows PENDING not PAID",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 7 — Account & finance (PDF §8)
# =============================================================================
PN7 = "Phase 7 — Account & finance department (PDF §8)"

add(
    dict(phase=7, pn=PN7, wbs="7.1", mid="FIN-01",
         main="Unified transaction ledger — every rupee that moves through the platform, in one record",
         module="Account", pdf="§8", feat="Unified transaction ledger",
         deliv="Every rupee that moves through the platform, in one record",
         surface="Account Portal", as_is="Does not exist",
         to_be="Append-only LedgerEntry for S1–S7; filters; OTP audit link; Admin may view, Account owns movement",
         wt="New", prio="P0", ac="No money movement without a ledger row; rows are immutable"),
    new_full(
        "DB: LedgerEntry (Stream, Direction, Amount, Gst, Commission, EntityType, EntityId, PaymentOrderId, At) immutable",
        "API: GET /account/ledger filters; export; write-only from payment/settlement services",
        "/account/ledger UI — Account layout sidebar Finance",
        None,
        "QA: consult pay writes S2; subscription writes S1; they never share one row",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.2", mid="FIN-02",
         main="Consultation payment reconciliation — online payments and reception collections matched and verified",
         module="Account", pdf="§8 + §7.2 + §16", feat="Consultation payment reconciliation",
         deliv="Online payments and reception collections matched and verified / Match every booking against every payment",
         surface="Account + Admin", as_is="Does not exist",
         to_be="Join booking ↔ payment ↔ appointment; KPI collected/pending/failed/refunded; CSV",
         wt="New", prio="P0", ac="Every paid booking has exactly one captured consult payment or an exception"),
    new_full(
        "Reuse ConsultPayment + Appointment; reconciliation query",
        "GET Admin/Account reconciliation + aggregates",
        "/account/consult-recon and /admin/consult-payments (same data, Account owns, Admin views)",
        None,
        "QA: unmatched payment appears in exception queue not silent",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.3", mid="FIN-03",
         main="Medicine ledger — medicine money tracked separately from consultation money",
         module="Account", pdf="§8 + §12", feat="Medicine ledger",
         deliv="Medicine money tracked separately from consultation money / Seller, platform and delivery shares recorded separately from consultation money",
         surface="Account Portal", as_is="Does not exist",
         to_be="/account/medicine-ledger; Stream=MEDICINE; seller/platform/delivery split columns; filled in Phase 14",
         wt="New", prio="P1", ac="Medicine GMV query cannot include consult fees"),
    new_full(
        "LedgerEntry stream MEDICINE with SplitSeller/SplitPlatform/SplitDelivery",
        "GET /account/medicine-ledger",
        "/account/medicine-ledger UI (empty until HomeoMeds orders exist — still ship the screen)",
        None,
        "QA: S2 consult row never appears on this screen",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.4", mid="FIN-04",
         main="Settlement runs — scheduled settlement of amounts due to doctors and pharmacies",
         module="Account", pdf="§8 + §16", feat="Settlement runs",
         deliv="Scheduled settlement of amounts due to doctors and pharmacies",
         surface="Account Portal", as_is="Does not exist",
         to_be="SettlementRun job T+N hold; net of commission/GST/refunds; pharmacies after Phase 14",
         wt="New", prio="P0", ac="A settlement run is replayable and lists included transactions"),
    new_full(
        "DB: SettlementRun, SettlementLine",
        "API: create run (dry-run + commit); list; detail",
        "/account/settlements",
        None,
        "QA: hold days exclude too-new captures; refunds net out",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.5", mid="FIN-05",
         main="Payout approval — payouts released only after OTP-verified approval",
         module="Account", pdf="§8 + §17", feat="Payout approval",
         deliv="Payouts released only after OTP-verified approval",
         surface="Account Portal", as_is="Does not exist",
         to_be="ApprovePayout requires SEC-07 OTP; Admin cannot approve; NEFT/file or gateway per PRE-01",
         wt="New", prio="P0", ac="Payout without OTP fails; audit shows who approved"),
    new_full(
        "DB: Payout (PayeeType Doctor|Pharmacy, Amount, Status, SettlementRunId)",
        "API: ApprovePayout with OTP; reject",
        "/account/payouts",
        "Bank file / Razorpay payout adapter per PRE-01",
        "QA: Account role only; replay OTP; Admin 403 on approve",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.6", mid="FIN-06",
         main="Refund processing — Account issues and tracks refunds (uses PAY-09)",
         module="Account", pdf="§8", feat="Refund processing",
         deliv="Issue and track refunds",
         surface="Account Portal", as_is="Does not exist",
         to_be="/account/refunds UI on PAY-09 APIs; OTP if policy requires",
         wt="New", prio="P0", ac="Account can issue refund; patient status updates; ledger reverses"),
    [
        {"sub": "UI /account/refunds: original payment, amount, reason, status",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "Ledger reverse entries on successful refund",
         "layer": "API", "be": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: partial refund; full refund; cash-policy path",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=7, pn=PN7, wbs="7.7", mid="FIN-07",
         main="Payment exception handling — investigate and resolve every failed or disputed payment",
         module="Account", pdf="§8 + §7.2", feat="Payment exception handling",
         deliv="Investigate and resolve every failed or disputed payment / Every failed or mismatched payment surfaced for action",
         surface="Account + Admin", as_is="Does not exist",
         to_be="Exception queue Failed/Partial/Webhook-mismatch/Refund-pending; retry/resolve with OTP on manual reconcile",
         wt="New", prio="P0", ac="Nothing failed is silent; resolve writes audit"),
    new_full(
        "DB: PaymentException state machine",
        "API: list/detail/retry/resolve; created by webhook mismatches",
        "/account/exceptions and /admin/payment-exceptions (same queue)",
        None,
        "QA: failed webhook creates exception; manual resolve requires Account role + OTP",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.8", mid="FIN-08",
         main="GST & invoice reports — invoice numbering and tax reporting",
         module="Account", pdf="§8 + §16", feat="GST & invoice reports",
         deliv="Invoice numbering and tax reporting",
         surface="Account Portal", as_is="Does not exist",
         to_be="/account/tax export using Invoice table and PRE-02 GST treatment — configuration not hardcoded",
         wt="New", prio="P1", ac="Date-range GST report matches invoices issued"),
    new_full(
        "Reuse Invoice; tax config table (GstRate, treatment flags)",
        "GET tax report + CSV",
        "/account/tax",
        None,
        "QA: rates come from config; changing rate does not rewrite old invoices",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.9", mid="FIN-09",
         main="Payee records — bank and KYC details for doctors and pharmacies",
         module="Account", pdf="§8", feat="Payee records",
         deliv="Bank and KYC details for doctors and pharmacies",
         surface="Account Portal", as_is="Doctor bank fields added in DOC-10; no Account UI; no pharmacy",
         to_be="/account/payees; OTP on doctor KYC change (fraud); pharmacies after Phase 14",
         wt="New", prio="P1", ac="Payout cannot run to a payee missing KYC"),
    new_full(
        "DB: Payee (DoctorId/PharmacyId, Bank, PAN, KYC status)",
        "API: list/update; OTP on bank change",
        "/account/payees",
        None,
        "QA: missing KYC blocks FIN-05",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.10", mid="FIN-10",
         main="Clinic collection view — cash and UPI collected at reception, per clinic and per day",
         module="Account", pdf="§8", feat="Clinic collection view",
         deliv="Cash and UPI collected at reception, per clinic and per day",
         surface="Account Portal", as_is="Does not exist",
         to_be="/account/clinic-collections from CollectAtReception rows; respect PRE-02 retain vs remit",
         wt="New", prio="P1", ac="Per clinic per day cash/UPI totals match reception receipts"),
    new_full(
        "Query ConsultPayment Method in CASH/UPI_OFFLINE grouped by DoctorId/date",
        "GET clinic-collections",
        "/account/clinic-collections",
        None,
        "QA: online Razorpay consults do not appear as clinic cash",
    ),
)

add(
    dict(phase=7, pn=PN7, wbs="7.11", mid="FIN-11",
         main="Separation of duties — Account controls money; Admin controls platform and clinical data",
         module="Account", pdf="§8", feat="Separation of duties",
         deliv="The Account department controls money. Admin controls the platform and clinical data. Neither can perform the other’s role, which protects both.",
         surface="Admin + Account", as_is="Account role does not exist",
         to_be="Enforce in RBAC: Admin cannot ApprovePayout; Account cannot edit repertory/users’ clinical masters",
         wt="New", prio="P0", ac="Automated tests for both forbidden directions"),
    [
        {"sub": "Seed Account role menus: ledger, recon, medicine ledger, exceptions, settlements, payouts, refunds, tax, payees, clinic collections",
         "layer": "Configuration", "db": Y, "be": Y, "fe": Y, "eff": "M", "wt": "New"},
        {"sub": "Negative tests: Admin POST ApprovePayout → 403; Account POST repertory section → 403",
         "layer": "QA", "be": Y, "api": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 8 — Trust & verification (PDF §13 + §7.2 credentialing) BEFORE public directory
# =============================================================================
PN8 = "Phase 8 — Trust & verification (PDF §13) — gate public doctor discovery"

add(
    dict(phase=8, pn=PN8, wbs="8.1", mid="TRU-01",
         main="Doctor credentialing — doctors upload qualification and registration documents for review",
         module="Trust", pdf="§13 + §7.2", feat="Doctor credentialing",
         deliv="Doctors upload qualification and registration documents for review",
         surface="Doctor Web + Admin", as_is="RegisterDoctor activates immediately; PassingUniversity/CertNo/QualificationId exist; no document store/queue",
         to_be="DoctorCredentialDocument + VerificationStatus Pending; do not auto-activate practice for public directory",
         wt="New", prio="P0", ac="New registrations sit in Pending with documents; existing doctors backfill per PRE-03"),
    new_full(
        "DB: DoctorVerification (Pending/Approved/Rejected/NeedsInfo), DoctorCredentialDocument",
        "API: upload documents; MyStatus; RegisterDoctor modified to VerificationStatus=Pending (not auto-practice)",
        "Doctor profile credentials tab + document upload",
        None,
        "QA: doctor without docs cannot appear in public directory (Phase 9)",
    ),
)

add(
    dict(phase=8, pn=PN8, wbs="8.2", mid="TRU-02",
         main="Verification review — Admin approves, rejects or requests further information",
         module="Trust", pdf="§13 + §7.2", feat="Verification review",
         deliv="Admin approves, rejects or requests more information before they can practise on the platform",
         surface="Admin Portal", as_is="Does not exist",
         to_be="/admin/doctor-credentialing list + detail; email notify doctor",
         wt="New", prio="P0", ac="Approve/reject/request-info all audited; doctor sees status"),
    new_full(
        "Reuse DoctorVerification",
        "API: Queue, {id}, Approve, Reject, RequestInfo",
        "/admin/doctor-credentialing and /:doctorId; verification column on users list",
        "Email notify (Phase 12 SMTP already used for activation)",
        "QA: reject reason required; request-info reopens Pending",
    ),
)

add(
    dict(phase=8, pn=PN8, wbs="8.3", mid="TRU-03",
         main="Verified badge — patients see which doctors have been verified",
         module="Trust", pdf="§13 + §2.2", feat="Verified badge / Credentials & verification",
         deliv="Patients see which doctors have been verified / Verified badge showing the doctor’s credentials are checked",
         surface="Public + Patient app", as_is="Does not exist",
         to_be="Badge on public doctor card and profile from VerificationStatus=Approved",
         wt="New", prio="P0", ac="Unverified doctors never show a verified badge"),
    [
        {"sub": "Public doctor DTO includes isVerified",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Badge component for website (Phase 9) and patient app (Phase 16)",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: pending doctor has no badge",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=8, pn=PN8, wbs="8.4", mid="TRU-04",
         main="Practice gate — only verified doctors appear to patients and take paid consultations",
         module="Trust", pdf="§13", feat="Practice gate",
         deliv="Only verified doctors appear to patients and take paid consultations",
         surface="Public + Payments", as_is="Any activated doctor could theoretically be listed; no public list yet",
         to_be="Public search and CreateConsultOrder require Approved; own-clinic staff booking may still work per PRE-03",
         wt="New", prio="P0", ac="Unverified doctor absent from /book; paid public consult blocked"),
    [
        {"sub": "Public doctor list filter VerificationStatus=Approved",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "CreateConsultOrder rejects unverified doctor for patient-originated bookings",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: reception can still book own clinic patient for unverified doctor if PRE-03 allows",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=8, pn=PN8, wbs="8.5", mid="TRU-05",
         main="Patient reviews — patients rate and review consultations",
         module="Trust", pdf="§13 + §2.8", feat="Patient reviews / Write a review",
         deliv="Patients rate and review consultations / Rate and review the consultation",
         surface="Patient + Public", as_is="Does not exist",
         to_be="Review linked to PatientAppId; one review per visit; moderation flag optional",
         wt="New", prio="P1", ac="Review only after completed consult; doctor cannot write self-reviews"),
    new_full(
        "DB: Review (PatientAppId, Rating, Text, Status, At)",
        "API: create (patient), list by doctor (public approved), doctor list own",
        "Public doctor profile reviews; patient app write flow Phase 16",
        None,
        "QA: duplicate review blocked; incomplete visit blocked",
    ),
)

add(
    dict(phase=8, pn=PN8, wbs="8.6", mid="TRU-06",
         main="Review appeal — doctors can appeal a review; patients manage reviews and raise an appeal",
         module="Trust", pdf="§13 + §2.8", feat="Review appeal / My reviews & appeal",
         deliv="Doctors can appeal a review / Manage reviews and raise an appeal",
         surface="Doctor + Patient + Admin", as_is="Does not exist",
         to_be="Appeal entity; admin resolves; patient my-reviews list",
         wt="New", prio="P2", ac="Appealed review hidden or marked pending per policy; admin decision audited"),
    new_full(
        "DB: ReviewAppeal",
        "API: doctor appeal; patient list/mine; admin resolve",
        "Doctor dashboard my-reviews; admin queue",
        None,
        "QA: only treating doctor of that visit can appeal",
    ),
)

add(
    dict(phase=8, pn=PN8, wbs="8.7", mid="TRU-07",
         main="Ranking transparency — patients are shown why doctors are ranked as they are",
         module="Trust", pdf="§13 + §2.2", feat="Ranking transparency / Ranking explanation",
         deliv="Patients are shown why doctors are ranked as they are / Transparent explanation of why doctors appear in a given order",
         surface="Public + Patient app", as_is="Does not exist",
         to_be="Documented ranking factors (verified, availability, fee, reviews, experience) + explain API/UI",
         wt="New", prio="P1", ac="Each search result has an explain link that matches the sort order"),
    new_full(
        "Ranking config (weights) — not a black box hardcoded sort",
        "GET doctors + GET ranking-explain/{doctorId}",
        "Search results: Why this order? panel",
        None,
        "QA: changing weights changes order and explanation together",
    ),
)


# =============================================================================
# PHASE 9 — Patient website / public (PDF §4)
# =============================================================================
PN9 = "Phase 9 — Patient website (PDF §4) — public discovery, booking, payment"

add(
    dict(phase=9, pn=PN9, wbs="9.1", mid="WEB-01",
         main="Product home page — complete public website presenting the service",
         module="Patient Website", pdf="§4", feat="Product home page",
         deliv="Complete public website presenting the service",
         surface="Patient Website", as_is="HomeoJobLanding ~50%; Velzon leftovers; CTAs incomplete for book/trust/apps",
         to_be="Book CTA, verified doctors, consult+SaaS pricing teaser, app download placeholders",
         wt="Existing Improvement", prio="P1", ac="Unauthenticated user can go Home → Book without calling the clinic"),
    improve_rows(
        "No schema",
        "Reuse public doctor/fee APIs when live",
        "HomeLandingPage: book, verified doctors, pricing teaser, store badges; strip leftover job-board feel",
        "QA: mobile viewport; deep link /book",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.2", mid="WEB-02",
         main="Consultation & subscription pricing — transparent pricing display",
         module="Patient Website", pdf="§4", feat="Consultation & subscription pricing",
         deliv="Transparent pricing display",
         surface="Patient Website", as_is="SaaS package marketing /pricing exists; no consult fee display per doctor",
         to_be="Keep SaaS pricing; add consult fee on doctor cards from ConsultFeeConfig",
         wt="Existing Improvement", prio="P1", ac="Patient sees consult fee before booking; doctor SaaS pricing still shown for doctors buying plans"),
    [
        {"sub": "Public doctor card shows in-clinic and tele fee",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "/pricing remains SaaS packages for doctors; label it clearly vs patient consult fees",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: fee matches checkout amount",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=9, pn=PN9, wbs="9.3", mid="WEB-03",
         main="Doctor discovery — browse and search verified doctors",
         module="Patient Website", pdf="§4 + §2.2", feat="Doctor discovery",
         deliv="Browse and search verified doctors",
         surface="Patient Website", as_is="Does not exist as a patient product (marketing only)",
         to_be="/book search: name, condition, category; filters availability, fee, experience, language, location",
         wt="New", prio="P0", ac="Only verified doctors; ranking explanation available (TRU-07)"),
    new_full(
        "Public doctor search indexes (from Doctor + verification + fee + schedule)",
        "GET /api/Public/Doctors + filters + profile",
        "/book list + filters + doctor profile page",
        None,
        "QA: unverified excluded; filters combine; empty state",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.4", mid="WEB-04",
         main="Self-service booking — book an appointment online without calling the clinic",
         module="Patient Website", pdf="§4 + §2.3", feat="Self-service booking",
         deliv="Book an appointment online without calling the clinic",
         surface="Patient Website", as_is="Staff-only appointments",
         to_be="Funnel /book/:id/slots → review+consent → create Patient+Appointment; OTP; rate-limit; respect schedule + unpaid hold policy PRE-03",
         wt="New", prio="P0", ac="Slot no longer free after confirmed booking; consent written to ConsentRecord"),
    new_full(
        "Reuse Patient + Appointment; BookingToken; PatientAuth OTP",
        "Public: Doctors/{id}/Slots, Create, {bookingToken}; POST PatientAuth RequestOtp/VerifyOtp",
        "/book/:doctorId/slots; /book/confirm; booking success",
        None,
        "QA: double-book same slot; hold expiry; OTP fail; consent required",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.5", mid="WEB-05",
         main="Online consultation payment on website — pay the fee at the time of booking",
         module="Patient Website", pdf="§4 + §9", feat="Online consultation payment",
         deliv="Pay the consultation fee at the time of booking",
         surface="Patient Website", as_is="Does not exist",
         to_be="Checkout uses PAY-03; success/fail pages; pay-at-clinic branch from PAY-05",
         wt="New", prio="P0", ac="Closing the tab does not lose a captured payment (PAY-02)"),
    [
        {"sub": "/book/pay/:bookingId + success/failure pages",
         "layer": "Web SPA", "fe": Y, "api": Y, "int": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: pay-at-clinic skip checkout; failed pay remains PENDING not PAID",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=9, pn=PN9, wbs="9.6", mid="WEB-06",
         main="Contact & enquiry form — enquiries reach the admin team and are tracked",
         module="Patient Website", pdf="§4 + §15", feat="Contact & enquiry form",
         deliv="Enquiries reach the admin team and are tracked",
         surface="Patient Website + Admin", as_is="EnquiryDetail API exists; public form partial; admin inbox screen missing",
         to_be="Confirmation + enquiry id; /admin/enquiries inbox; optional convert to SupportTicket",
         wt="Existing Improvement", prio="P1", ac="Every website enquiry is visible to admin with status"),
    improve_rows(
        "Reuse EnquiryDetail; add Status/AssignedTo if missing",
        "Existing enquiry POST; admin list API if missing",
        "/contact confirmation; /admin/enquiries inbox (not Velzon mailbox)",
        "QA: guest submit → admin sees row",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.7", mid="WEB-07",
         main="Privacy policy — covering data protection, recording and pharmacy consent",
         module="Legal", pdf="§4 + §17", feat="Privacy policy",
         deliv="Covering data protection, recording and pharmacy consent",
         surface="Patient Website", as_is="Privacy page partial; not DPDP/tele/pharmacy complete",
         to_be="Publish PRE-03 legal copy: DPDP, recording, pharmacy, data-rights, family/caregiver",
         wt="Existing Improvement", prio="P0", ac="Versioned policy; booking consent references this version"),
    [
        {"sub": "Replace /privacy content with signed legal copy; store PolicyVersion",
         "layer": "Web SPA", "fe": Y, "db": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: booking consent shows policy version number",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=9, pn=PN9, wbs="9.8", mid="WEB-08",
         main="Terms of service — covering payments, refunds, telemedicine and medicine orders",
         module="Legal", pdf="§4", feat="Terms of service",
         deliv="Covering payments, refunds, telemedicine and medicine orders",
         surface="Patient Website", as_is="Terms partial",
         to_be="Payments, refunds, telemedicine, HomeoMeds, codes-before-names",
         wt="Existing Improvement", prio="P0", ac="Checkout cannot complete without accepted terms version"),
    [
        {"sub": "Replace /terms with signed copy; version stamp on booking/payment",
         "layer": "Web SPA", "fe": Y, "db": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: terms checkbox required on confirm",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=9, pn=PN9, wbs="9.9", mid="WEB-09",
         main="Doctor self-registration — doctors apply to join the platform online",
         module="Patient Website", pdf="§4", feat="Doctor self-registration",
         deliv="Doctors apply to join the platform online",
         surface="Public", as_is="RegisterDoctor ~90%; activates immediately; no document upload",
         to_be="Document upload; pending-verification message; TRU-01 status Pending",
         wt="Existing Improvement", prio="P0", ac="New doctor cannot take paid public consults until Approved"),
    improve_rows(
        "Documents via TRU-01 tables",
        "RegisterDoctor sets Pending; does not auto-activate practice for directory",
        "/register multipart docs; pending message; status page",
        "QA: register → admin queue row",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.10", mid="WEB-10",
         main="Email activation — secure account activation by email link",
         module="Auth", pdf="§4", feat="Email activation",
         deliv="Secure account activation by email link",
         surface="Public", as_is="ActivateUser via /login?UserId= ~90%; align with verification; expiry/resend incomplete",
         to_be="Token expiry, resend, align with credentialing (activation ≠ verified practice)",
         wt="Existing Improvement", prio="P1", ac="Expired link fails; resend works; activated unverified still not in directory"),
    improve_rows(
        "Activation token TTL if missing",
        "Resend activation; do not treat activation as TRU-04 gate",
        "Login activation query param UX + expiry copy",
        "QA: used link, expired link, resend",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.11", mid="WEB-11",
         main="Waitlist offer — get offered a slot when one is freed by a cancellation (patient website + later app)",
         module="Booking", pdf="§2.3", feat="Waitlist offer",
         deliv="Get offered a slot when one is freed by a cancellation",
         surface="Patient Website + App", as_is="Cancel hook stub in Phase 4; no waitlist records",
         to_be="BookingWaitlist; on cancel, offer FIFO; timeout",
         wt="New", prio="P2", ac="Cancelled slot offered to waitlisted patient before returning to open search"),
    new_full(
        "DB: BookingWaitlist",
        "API: join waitlist; worker on CancelAppointment",
        "Public waitlist CTA on full days",
        "Notify SMS/push when Phase 12 exists",
        "QA: two waiters; first gets offer; expiry releases to second",
    ),
)

add(
    dict(phase=9, pn=PN9, wbs="9.12", mid="WEB-12",
         main="Health articles — educational content to build trust and engagement",
         module="Patient Website", pdf="§2.2", feat="Health articles",
         deliv="Educational content to build trust and engagement",
         surface="Patient Website + App", as_is="Blog/news APIs on classic exist for marketing",
         to_be="Reuse blog/news as health articles on public site and later patient app; admin already has news/blog",
         wt="Existing Improvement", prio="P2", ac="Published articles appear on public site; draft does not"),
    [
        {"sub": "Public articles list/detail from existing blog/news APIs",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: unpublished hidden",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 10 — Digital prescription eRx (PDF §11)
# =============================================================================
PN10 = "Phase 10 — Digital prescription / eRx (PDF §11)"

add(
    dict(phase=10, pn=PN10, wbs="10.1", mid="ERX-01",
         main="Potency module — every remedy carries a defined potency from a managed list, no free text",
         module="eRx", pdf="§11", feat="Potency module",
         deliv="Every remedy carries a defined potency, selected from a managed list — no free text",
         surface="Doctor Web + Admin", as_is="Dose saved as empty free-text dose: ''",
         to_be="PotencyMaster CRUD; PotencyId on PrescriptionRemedyDetail; stop empty Dose",
         wt="New", prio="P0", ac="Cannot sign eRx with missing potency"),
    new_full(
        "DB: PotencyMaster; PrescriptionRemedyDetail.PotencyId",
        "API: potency DDL; extend SavePrescriptionDetail",
        "Rx modal potency picker (6C, 30C, 200C, 1M…)",
        None,
        "QA: free text potency rejected",
    ),
)

add(
    dict(phase=10, pn=PN10, wbs="10.2", mid="ERX-02",
         main="Dosage details — frequency, duration and instructions recorded clearly",
         module="eRx", pdf="§11", feat="Dosage details",
         deliv="Frequency, duration and instructions recorded clearly",
         surface="Doctor Web Portal", as_is="Unstructured/empty dose",
         to_be="Frequency, Duration, Instructions fields on each remedy line",
         wt="New", prio="P0", ac="Print/PDF shows frequency, duration, instructions"),
    new_full(
        "DB: Frequency, Duration, Instructions on remedy line",
        "Extend save/get Rx APIs",
        "Rx row UI for the three fields",
        None,
        "QA: all three persist and print",
    ),
)

add(
    dict(phase=10, pn=PN10, wbs="10.3", mid="ERX-03",
         main="Visit notes kept separate from the prescription — private clinical notes never appear on the prescription given to the patient or pharmacy",
         module="eRx", pdf="§11 + §5.3", feat="Visit notes kept separate from the prescription",
         deliv="The doctor’s private clinical notes never appear on the prescription given to the patient or pharmacy",
         surface="Doctor Web Portal", as_is="History notes mixed into Rx modal tabs",
         to_be="Visit Notes panel outside Rx; NoteType; IsErxExcluded; eRx print has remedies+labs only",
         wt="New", prio="P0", ac="Patient PDF and pharmacy payload contain zero private note text"),
    new_full(
        "Extend AppointmentHistoryNote with NoteType, IsErxExcluded",
        "GET /api/Erx/ByAppointment/{id} returns snapshot without notes",
        "Move History Notes out of Rx modal; note types Chief complaint / Follow-up / General",
        None,
        "QA: print eRx vs export case PDF (doctor copy still has notes)",
    ),
)

add(
    dict(phase=10, pn=PN10, wbs="10.4", mid="ERX-04",
         main="Signed digital prescription — issued and locked once signed, so it cannot be altered afterwards",
         module="eRx", pdf="§11 + §2.5 + §5.3", feat="Signed digital prescription",
         deliv="Issued and locked once signed, so it cannot be altered afterwards / Official prescription issued by the doctor",
         surface="Doctor Web", as_is="SavePrescriptionDetail overwritable; no lock/sign",
         to_be="ErxSnapshot immutable on sign; subsequent edits create a new unsigned draft not a silent mutate",
         wt="New", prio="P0", ac="Signed snapshot hash stable; UI cannot edit locked lines"),
    new_full(
        "DB: ErxSnapshot, ErxSnapshotItem (codes+names stored; patient API returns codes until pharmacy accept)",
        "API: SignErx; GetErx; reject mutate of signed snapshot",
        "Sign button; locked badge on board header",
        None,
        "QA: POST save after sign does not change snapshot",
    ),
)

add(
    dict(phase=10, pn=PN10, wbs="10.5", mid="ERX-05",
         main="Prescription history — every prescription stored against the patient and visit",
         module="eRx", pdf="§11", feat="Prescription history",
         deliv="Every prescription stored against the patient and visit",
         surface="Doctor + Patient", as_is="Rx saved per appointment but no signed history product",
         to_be="List snapshots by PatientId; open past visit eRx",
         wt="New", prio="P1", ac="All signed eRx retrievable by patient and treating doctor"),
    [
        {"sub": "GET history by patient; by appointment",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Doctor history panel lists signed eRx; patient timeline in Phase 16",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: another doctor 403",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=10, pn=PN10, wbs="10.6", mid="ERX-06",
         main="Patient view of prescription in app — patient sees the prescription immediately (codes until pharmacy accept)",
         module="eRx", pdf="§11 + §2.5", feat="Patient view in app",
         deliv="Patient sees the prescription immediately in the app",
         surface="Patient App + Website", as_is="No patient surface",
         to_be="Patient GET eRx by appointment; names redacted until Phase 14 accept; website optional view",
         wt="New", prio="P0", ac="Immediately after sign, patient API returns codes (or names if no HomeoMeds order required for in-clinic dispense policy — document)"),
    [
        {"sub": "Patient-authorised GET Erx; apply code protection ERX-08",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Web confirmation/records page can show the same payload before the app exists",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: doctor names vs patient codes differ until pharmacy OTP accept",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=10, pn=PN10, wbs="10.7", mid="ERX-07",
         main="Print & PDF — printable and downloadable prescription",
         module="eRx", pdf="§11", feat="Print & PDF",
         deliv="Printable and downloadable prescription",
         surface="Doctor + Patient", as_is="No eRx PDF",
         to_be="Patient copy without notes; doctor copy may include letterhead; SecureDocument store",
         wt="New", prio="P1", ac="PDF matches signed snapshot; notes absent on patient copy"),
    [
        {"sub": "PDF renderer from ErxSnapshot; store via SEC-09",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "Print/Download on doctor board and later patient app",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: notes leak test on patient PDF",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=10, pn=PN10, wbs="10.8", mid="ERX-08",
         main="Remedy code protection — patient sees remedy codes rather than names until a licensed pharmacy accepts the order",
         module="eRx", pdf="§11", feat="Remedy code protection",
         deliv="The patient sees remedy codes rather than remedy names until a licensed pharmacy accepts the order",
         surface="Patient", as_is="Does not exist",
         to_be="Code map on snapshot items; patient serializer strips names; doctor always sees names",
         wt="New", prio="P0", ac="Patient JSON has no remedy name fields until PharmacyAccepted"),
    [
        {"sub": "Assign stable remedy codes on snapshot items",
         "layer": "API", "be": Y, "db": Y, "eff": "M", "wt": "New"},
        {"sub": "Patient serializer tests (contract tests)",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=10, pn=PN10, wbs="10.9", mid="ERX-09",
         main="Reveal on pharmacy acceptance — names and price revealed; payment QR or link generated (implemented with HomeoMeds; contract now)",
         module="eRx", pdf="§11 + §12", feat="Reveal on pharmacy acceptance",
         deliv="Once a pharmacy accepts and verifies by OTP, the remedy names and price are revealed and a payment QR or link is generated",
         surface="Patient + Pharmacy", as_is="Does not exist",
         to_be="Flag PharmacyAccepted on order; patient GET starts returning names; CreateMedicine payment link; full UI in Phase 14",
         wt="New", prio="P0", ac="Names appear only after OTP accept; QR/link only after quote"),
    [
        {"sub": "API contract: AcceptOrder(OTP) → reveal names; Quote → PayQR",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Implement fully with MED-* in Phase 14 — do not reveal names from doctor sign alone",
         "layer": "Architecture", "be": Y, "eff": "S", "wt": "New"},
    ],
)

add(
    dict(phase=10, pn=PN10, wbs="10.10", mid="ERX-10",
         main="Repeat prescription / refill — patient requests; doctor approves or rejects from the mobile app (inbox on web too)",
         module="eRx", pdf="§11 + §3 + §2.7", feat="Repeat prescription / refill",
         deliv="Patient requests a refill; doctor approves or rejects from the mobile app",
         surface="Patient + Doctor mobile + Doctor web", as_is="Does not exist",
         to_be="RefillRequest against ErxSnapshot; doctor inbox; no case-taking on mobile",
         wt="New", prio="P1", ac="Reject/approve audited; approved refill creates a new order draft not an editable old snapshot"),
    new_full(
        "DB: RefillRequest",
        "API: patient request; doctor inbox; approve/reject",
        "Doctor web inbox; doctor app in Phase 17; patient app in Phase 16",
        None,
        "QA: cannot refill unsigned/locked-expired per policy",
    ),
)

add(
    dict(phase=10, pn=PN10, wbs="10.11", mid="ERX-11",
         main="Lab orders alongside the prescription (keep existing labs on eRx snapshot)",
         module="eRx", pdf="§11", feat="Lab orders",
         deliv="Tests ordered alongside the prescription",
         surface="Doctor Web", as_is="Labs on Rx modal live",
         to_be="Include lab orders in signed snapshot; still not on private notes",
         wt="Existing Improvement", prio="P1", ac="Signed eRx PDF lists labs; results stay clinical-side"),
    [
        {"sub": "Copy lab orders into ErxSnapshot on sign",
         "layer": "API", "be": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: labs visible on eRx; result values not required on patient copy",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 11 — Telemedicine (PDF §10 + §2.4 + §3 + §5)
# =============================================================================
PN11 = "Phase 11 — Telemedicine / video consultation (PDF §10)"

add(
    dict(phase=11, pn=PN11, wbs="11.1", mid="TEL-01",
         main="Doctor availability — online or offline status and working hours",
         module="Telemedicine", pdf="§10 + §3", feat="Doctor availability",
         deliv="Doctor sets online or offline status and working hours",
         surface="Doctor Web + Doctor App", as_is="E-CONSULT status + WhatsApp video stub only; no availability",
         to_be="TeleAvailability + heartbeat; working hours reuse daily schedule",
         wt="New", prio="P0", ac="Offline doctor cannot be offered instant consult"),
    new_full(
        "DB: TeleAvailability (Online/Offline, LastHeartbeat)",
        "API: SET/GET availability; public GET by doctorId",
        "Dashboard toggle Online for teleconsult",
        None,
        "QA: heartbeat expiry sets offline",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.2", mid="TEL-02",
         main="Telemedicine waiting queue — who is waiting, how long, whether they have paid",
         module="Telemedicine", pdf="§10 + §3", feat="Telemedicine waiting queue",
         deliv="Doctor sees who is waiting, for how long, and whether they have paid",
         surface="Doctor Web + App", as_is="E-CONSULT list only",
         to_be="Queue by doctor + Online + paid/confirmed; wait time; Join",
         wt="New", prio="P0", ac="Unpaid patients do not join if policy requires PAID"),
    new_full(
        "Queue query on E-CONSULT + PaymentStatus + join time",
        "GET Tele/Queue; POST start session",
        "Dedicated Tele queue on dashboard (beyond E-CONSULT list)",
        "SignalR/poll",
        "QA: wait time increases; paid badge matches PAY-08",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.3", mid="TEL-03",
         main="Video consultation in browser — doctor consults directly from the web portal",
         module="Telemedicine", pdf="§10", feat="Video consultation in browser",
         deliv="Doctor consults directly from the web portal",
         surface="Doctor Web Portal", as_is="SweetAlert WhatsApp stub",
         to_be="/teleconsult/:sessionId in-browser video; replace stub; vendor from PRE-03",
         wt="New Integration", prio="P0", ac="Doctor and patient in same room linked to PatientAppId"),
    new_full(
        "DB: TeleSession (PatientAppId, RoomId, Status)",
        "API: create/end session; vendor tokens",
        "Full-page video room mute/end",
        "Video SDK (Agora/Twilio/Daily)",
        "QA: replace WhatsApp alert; session ends write timestamp",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.4", mid="TEL-04",
         main="Video consultation on mobile — doctor and patient both consult from the app (apps in 16–17; same session tokens)",
         module="Telemedicine", pdf="§10 + §2.4 + §3", feat="Video consultation on mobile",
         deliv="Doctor and patient both consult from the app / Face-to-face consultation from home / Start the video or audio consultation from the phone",
         surface="Patient App + Doctor App", as_is="No apps",
         to_be="Same TeleSession tokens; mobile join implemented in Phases 16–17 against this API",
         wt="New", prio="P0", ac="Web doctor + mobile patient in one session"),
    [
        {"sub": "Token API must be client-agnostic (web and mobile)",
         "layer": "API", "be": Y, "api": Y, "mob": Y, "eff": "S", "wt": "New"},
        {"sub": "Contract tests: token TTL, room id = session id",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=11, pn=PN11, wbs="11.5", mid="TEL-05",
         main="Device check — camera, microphone and connection tested before joining",
         module="Telemedicine", pdf="§10 + §2.4", feat="Device check",
         deliv="Camera, microphone and connection tested before joining",
         surface="Doctor Web + Patient", as_is="Does not exist",
         to_be="Pre-join device/network check UI; block join on hard fail with guidance",
         wt="New", prio="P1", ac="User sees camera/mic/network result before entering the room"),
    [
        {"sub": "Shared device-check component (web); mobile equivalent in Phase 16",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: deny camera permission → clear guidance",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=11, pn=PN11, wbs="11.6", mid="TEL-06",
         main="Waiting room — patient waits with live status until the doctor joins",
         module="Telemedicine", pdf="§10 + §2.4", feat="Waiting room",
         deliv="Patient waits with live status / Wait with live status until the doctor joins",
         surface="Patient", as_is="Does not exist",
         to_be="Waiting room page: position, doctor online, paid, estimated wait",
         wt="New", prio="P1", ac="Status updates without refresh (poll/SignalR)"),
    new_full(
        "Session status events",
        "GET session status for patient token",
        "Waiting room UI on web (and app later)",
        None,
        "QA: doctor join transitions patient into the call",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.7", mid="TEL-07",
         main="Recording consent — consent captured before the consultation; no recording without it",
         module="Telemedicine", pdf="§10 + §2.4 + §17", feat="Recording consent",
         deliv="Consent captured before the consultation; no recording without it / Explicit consent requested before any recording",
         surface="Patient + Doctor", as_is="AudioCaseConsentLog only for audio case taking",
         to_be="ConsentRecord type TeleRecording; block record if declined; both sides see status",
         wt="New", prio="P0", ac="Recording pipeline cannot start if consent denied"),
    new_full(
        "Write ConsentRecord TeleRecording; TeleConsentLog events",
        "API: capture consent; session.recordAllowed",
        "Pre-join modal both sides",
        None,
        "QA: decline → no record; audit row still written",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.8", mid="TEL-08",
         main="Rejoin after dropped call — either side can rejoin the same consultation",
         module="Telemedicine", pdf="§10 + §2.4", feat="Rejoin after dropped call",
         deliv="Either side can rejoin the same consultation / Rejoin the same consultation if the connection drops",
         surface="Doctor + Patient", as_is="Does not exist",
         to_be="Session Active until End; rejoin tokens; dashboard Rejoin",
         wt="New", prio="P0", ac="Drop and rejoin same PatientAppId session; End prevents rejoin"),
    [
        {"sub": "API: Rejoin issues token if Status=Active",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "UI Rejoin on doctor dashboard and patient waiting/call fail",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: after End, rejoin 409",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=11, pn=PN11, wbs="11.9", mid="TEL-09",
         main="Join failure handling — clear fallback if the call cannot connect, with options to retry",
         module="Telemedicine", pdf="§10 + §2.4", feat="Join failure handling / Fallback & join failure handling",
         deliv="Clear fallback if the call cannot connect / Clear guidance if the call fails, with options to retry",
         surface="Doctor + Patient", as_is="Does not exist",
         to_be="Error codes; retry; support ticket deep link later; do not silently fail",
         wt="New", prio="P1", ac="Every join failure has a user-visible next action"),
    [
        {"sub": "Standardise join error codes; log TeleSessionEvent",
         "layer": "API", "be": Y, "eff": "S", "wt": "New"},
        {"sub": "Fallback UI: retry, rejoin, contact support",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: simulated token fail",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=11, pn=PN11, wbs="11.10", mid="TEL-10",
         main="Case-linked chat — chat attached to the consultation record, not a loose message thread",
         module="Telemedicine", pdf="§10 + §2.4", feat="Case-linked chat",
         deliv="Chat attached to the consultation record / Chat tied to the consultation, not a loose message thread",
         surface="Doctor + Patient", as_is="Does not exist",
         to_be="Messages keyed by TeleSession/PatientAppId; retained with the case",
         wt="New", prio="P1", ac="Chat not visible on a different appointment"),
    new_full(
        "DB: TeleChatMessage (SessionId, PatientAppId, Body, At)",
        "API: post/list; doctor and patient only",
        "In-room chat panel",
        None,
        "QA: IDOR other session 403",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.11", mid="TEL-11",
         main="Consultation summary — issued to the patient at the end of the call",
         module="Telemedicine", pdf="§10 + §2.4", feat="Consultation summary",
         deliv="Issued to the patient at the end of the call / Written summary available immediately after the call",
         surface="Patient + Doctor", as_is="Does not exist",
         to_be="Doctor submits summary (not private notes); patient sees immediately; may share eRx id",
         wt="New", prio="P1", ac="Summary available without waiting for signed eRx if doctor issues it first"),
    new_full(
        "DB: ConsultationSummary (PatientAppId, Text, At)",
        "API: doctor PUT; patient GET",
        "Post-call summary form; patient view",
        None,
        "QA: private notes not copied into summary by default",
    ),
)

add(
    dict(phase=11, pn=PN11, wbs="11.12", mid="TEL-12",
         main="Instant consultation — patient requests an immediate consultation with an available doctor",
         module="Telemedicine", pdf="§10 + §2.3", feat="Instant consultation / Instant consult request / Queue & doctor offer",
         deliv="Patient can request an immediate consultation with an available doctor / See position in queue and accept an offered doctor",
         surface="Patient + Doctor", as_is="Does not exist",
         to_be="Instant request pool; offer doctor; patient accepts; optional surcharge from fee config",
         wt="New", prio="P1", ac="Offline doctors never offered; patient can decline an offered doctor"),
    new_full(
        "DB: InstantConsultRequest, DoctorOffer",
        "API: request, match, offer, accept, queue position",
        "Patient instant CTA; doctor offer banner",
        None,
        "QA: no online doctor → empty state; surcharge on checkout",
    ),
)


# =============================================================================
# PHASE 12 — Communication & notifications (PDF §14)
# =============================================================================
PN12 = "Phase 12 — Communication & notifications (PDF §14)"

add(
    dict(phase=12, pn=PN12, wbs="12.1", mid="COM-01",
         main="SMS — appointment confirmation, registration confirmation, doctor-unavailable, OTP, reschedule and cancellation",
         module="Notifications", pdf="§14", feat="SMS",
         deliv="Appointment confirmation, registration confirmation, doctor-unavailable notice, OTP, reschedule and cancellation alerts",
         surface="Shared", as_is="No SMS provider; WhatsApp exists separately",
         to_be="SmsProvider + templates + DLT; event hooks; DND/opt-out on Patient",
         wt="New Integration", prio="P0", ac="Each listed event can send SMS; notify failure does not roll back the business transaction"),
    new_full(
        "DB: SmsTemplate, SmsMessageLog; Patient.SmsOptOut",
        "API: Templates CRUD; Send; History; event worker",
        "Doctor /doctor/sms enable-per-event; admin templates",
        "MSG91/Twilio/etc. from PRE-03",
        "QA: OTP SMS; reschedule old→new; cancel; doctor unavailable; opt-out honored",
    ),
)

add(
    dict(phase=12, pn=PN12, wbs="12.2", mid="COM-02",
         main="WhatsApp — templates, offers, health tips and bulk messaging with delivery reports and opt-out (complete remaining ~50%)",
         module="Notifications", pdf="§14", feat="WhatsApp",
         deliv="Templates, offers, health tips and bulk messaging to patients, with delivery reports and opt-out",
         surface="Doctor Web", as_is="Meta Cloud send/templates/campaigns live; bulk/opt-in/analytics incomplete",
         to_be="Finish bulk queue UX, template approval states, delivery reports, opt-out",
         wt="Existing Improvement", prio="P1", ac="Bulk campaign shows delivered/failed; opted-out patients excluded"),
    improve_rows(
        "Extend WhatsApp log if delivery receipts not stored",
        "Finish bulk worker + receipt webhook",
        "WhatsAppModal: campaign status, failures, opt-out",
        "QA: opt-in required; template rejected state",
    ),
)

add(
    dict(phase=12, pn=PN12, wbs="12.3", mid="COM-03",
         main="Push notifications — booking confirmed, payment received, doctor ready, prescription issued, order updates, refill requests",
         module="Notifications", pdf="§14 + §2.9 + §3", feat="Push notifications",
         deliv="Booking confirmed, payment received, doctor ready, prescription issued, order updates, refill requests / Patient waiting, new booking, reschedule, cancellation, refill request, payment received",
         surface="Patient App + Doctor App", as_is="Does not exist",
         to_be="DeviceToken + FCM/APNs; event fan-out; apps register tokens in Phases 16–17",
         wt="New Integration", prio="P0", ac="Each listed event has a template; unregister on logout"),
    new_full(
        "DB: DeviceToken, AppNotification",
        "API: Devices/Register; Notifications/Send; list",
        "Server event hooks (apps consume in 16–17)",
        "FCM + APNs",
        "QA: doctor and patient payloads isolated; token refresh",
    ),
)

add(
    dict(phase=12, pn=PN12, wbs="12.4", mid="COM-04",
         main="Email — registration, activation, password reset, invoices and receipts",
         module="Notifications", pdf="§14", feat="Email",
         deliv="Registration, activation, password reset, invoices and receipts",
         surface="Shared", as_is="SMTP used for activation/forgot (plaintext issue fixed in Phase 0); no invoice email",
         to_be="Template emails for listed events; invoice PDF attach via SEC-09",
         wt="Existing Improvement", prio="P1", ac="Each listed event has an email; bounce logged"),
    improve_rows(
        "EmailMessageLog if missing",
        "Send invoice/receipt email on PAID; keep activation/reset",
        "No new public page — transactional only",
        "QA: reset still has no plaintext password; invoice attached",
    ),
)

add(
    dict(phase=12, pn=PN12, wbs="12.5", mid="COM-05",
         main="In-app notification centre — all alerts collected in one place for patients and doctors",
         module="Notifications", pdf="§14 + §2.8", feat="In-app notification centre / Notifications centre",
         deliv="All alerts collected in one place for patients and doctors / All alerts in one place",
         surface="Patient App + Doctor Web/App", as_is="Does not exist (Velzon notification demo is not product)",
         to_be="AppNotification inbox; mark read; do not use Velzon demo notifications",
         wt="New", prio="P1", ac="Same events as push appear in-app if push is denied"),
    new_full(
        "Reuse AppNotification",
        "GET/PATCH notifications",
        "Doctor web bell; patient/doctor app centres in 16–17",
        None,
        "QA: mark read; empty state; not Velzon demo route",
    ),
)


# =============================================================================
# PHASE 13 — Support & help desk (PDF §15 + §7.2 tickets)
# =============================================================================
PN13 = "Phase 13 — Support & help desk (PDF §15)"

add(
    dict(phase=13, pn=PN13, wbs="13.1", mid="SUP-01",
         main="Patient support tickets — patients raise issues from the app and track them",
         module="Support", pdf="§15 + §2.9", feat="Patient support tickets / Help centre & support tickets",
         deliv="Patients raise issues from the app and track them to resolution",
         surface="Patient App + Admin", as_is="Does not exist; EnquiryDetail is not tickets; Velzon tickets demo ignore",
         to_be="SupportTicket reporter=Patient; status track; not Enquiry, not Velzon /apps-tickets-*",
         wt="New", prio="P1", ac="Patient sees status changes; cannot see other patients’ tickets"),
    new_full(
        "DB: SupportTicket, SupportTicketMessage, SupportTicketAttachment",
        "API: create/list/mine for patient",
        "Patient app screens in Phase 16; web help optional",
        None,
        "QA: IDOR; attachment via SEC-09",
    ),
)

add(
    dict(phase=13, pn=PN13, wbs="13.2", mid="SUP-02",
         main="Doctor support tickets — doctors raise issues from the portal",
         module="Support", pdf="§15", feat="Doctor support tickets",
         deliv="Doctors raise issues from the portal",
         surface="Doctor Web + Admin", as_is="Does not exist",
         to_be="Same ticket tables; reporter=Doctor",
         wt="New", prio="P1", ac="Doctor portal has Raise issue; tracks to resolution"),
    [
        {"sub": "Reuse SupportTicket with ReporterRole=Doctor",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "/doctor/support list/create/detail",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: doctor cannot see another doctor’s tickets",
         "layer": "QA", "fe": Y, "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=13, pn=PN13, wbs="13.3", mid="SUP-03",
         main="Admin issue queue — all issues in one queue with priority, assignment and status",
         module="Support", pdf="§15 + §7.2", feat="Admin issue queue / Support ticket queue",
         deliv="All issues in one queue with priority, assignment and status / All patient and doctor issues in one queue with assignment and resolution",
         surface="Admin Portal", as_is="Does not exist",
         to_be="/admin/support-tickets; SLA timestamps; assign",
         wt="New", prio="P1", ac="Patient and doctor tickets in one queue; assignable"),
    new_full(
        "Extend ticket with Priority, Assignee, SLA",
        "Admin list/filter/assign/status",
        "/admin/support-tickets and /:id",
        None,
        "QA: not using Velzon ticket pages",
    ),
)

add(
    dict(phase=13, pn=PN13, wbs="13.4", mid="SUP-04",
         main="Ticket conversation — full message thread with attachments",
         module="Support", pdf="§15", feat="Ticket conversation",
         deliv="Full message thread with attachments",
         surface="Admin + Doctor + Patient", as_is="Does not exist",
         to_be="Threaded messages + SEC-09 attachments",
         wt="New", prio="P1", ac="Both sides see the same thread order"),
    [
        {"sub": "API: messages CRUD + attachment",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "Detail UI thread",
         "layer": "Web SPA", "fe": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: attachment unauthorised download 403",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=13, pn=PN13, wbs="13.5", mid="SUP-05",
         main="Enquiry management — website enquiries tracked (ties WEB-06)",
         module="Support", pdf="§15 + §4", feat="Enquiry management",
         deliv="Website enquiries reach the admin team and are tracked",
         surface="Admin Portal", as_is="API exists; inbox missing",
         to_be="Complete WEB-06 admin inbox if not already; convert enquiry → ticket optional",
         wt="Existing Improvement", prio="P1", ac="No enquiry is only in email"),
    [
        {"sub": "Admin enquiry inbox SLA fields",
         "layer": "Web SPA", "fe": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: convert to ticket keeps original enquiry id",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=13, pn=PN13, wbs="13.6", mid="SUP-06",
         main="Help centre — self-service help content for patients",
         module="Support", pdf="§15 + §2.9", feat="Help centre",
         deliv="Self-service help content for patients",
         surface="Patient Website + App", as_is="Does not exist",
         to_be="CMS articles or static help; searchable",
         wt="New", prio="P2", ac="Patient can open help without raising a ticket"),
    new_full(
        "HelpArticle table or reuse blog category=help",
        "Public GET help articles",
        "Help centre list/detail",
        None,
        "QA: unpublished hidden",
    ),
)

add(
    dict(phase=13, pn=PN13, wbs="13.7", mid="SUP-07",
         main="Assisted booking — help for patients who need support booking",
         module="Support", pdf="§15 + §2.9", feat="Assisted booking",
         deliv="Support for patients who need help booking / Help for patients who need support booking",
         surface="Support + Reception + Patient", as_is="Does not exist",
         to_be="Staff (admin/support/reception) can complete public booking on behalf of a patient with audit; patient app CTA to request assistance",
         wt="New", prio="P2", ac="Assisted booking appears in audit as staff-created; patient still owns the appointment"),
    new_full(
        "Flag BookingChannel=Assisted; AssistedRequest ticket type",
        "API: request assistance; staff completes WEB-04 create on behalf",
        "Patient CTA; admin/reception assisted-book wizard",
        None,
        "QA: consent still captured; payment still Homeocentrum",
    ),
)


# =============================================================================
# PHASE 14 — HomeoMeds (PDF §12 + pharmacy admin)
# =============================================================================
PN14 = "Phase 14 — HomeoMeds online medicine delivery (PDF §12)"

add(
    dict(phase=14, pn=PN14, wbs="14.1", mid="MED-01",
         main="Pharmacy partner onboarding — licensed premises, licence records, activation checklist",
         module="HomeoMeds", pdf="§12 + §7.2", feat="Pharmacy partner onboarding / Pharmacy partner management",
         deliv="Licensed premises registered with licence records and an activation checklist / Onboard and activate licensed pharmacies",
         surface="Pharmacy + Admin", as_is="Does not exist",
         to_be="PharmacyPartner + licences; admin activation; pharmacy console login",
         wt="New", prio="P0", ac="Unactivated pharmacy cannot receive orders"),
    new_full(
        "DB: PharmacyPartner, PharmacyLicence, activation checklist fields",
        "API: onboard, admin activate, pharmacy login role",
        "/pharmacy/onboard; /admin pharmacy partners",
        None,
        "QA: checklist incomplete blocks activate",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.2", mid="MED-02",
         main="Licence gating — expired or suspended licences blocked automatically from receiving orders",
         module="HomeoMeds", pdf="§12", feat="Licence gating",
         deliv="Pharmacies with expired or suspended licences are blocked automatically from receiving orders",
         surface="Backend", as_is="Does not exist",
         to_be="Scheduler; routing skip; banner on pharmacy console",
         wt="New", prio="P0", ac="Expired licence receives zero new orders the next minute"),
    [
        {"sub": "Job: mark licence expired/suspended; exclude from routing",
         "layer": "Backend", "be": Y, "db": Y, "eff": "M", "wt": "New"},
        {"sub": "Pharmacy console status banner",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: expire a licence in test → order not offered",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.3", mid="MED-03",
         main="Fulfilment consent & pharmacy selection — patient chooses the seller and consents before the prescription is shared",
         module="HomeoMeds", pdf="§12 + §2.7", feat="Fulfilment consent & pharmacy selection",
         deliv="The patient chooses the seller and consents before the prescription is shared",
         surface="Patient", as_is="Does not exist",
         to_be="ConsentRecord PharmacyShare; patient picks from routed sellers; share only after consent",
         wt="New", prio="P0", ac="No pharmacy sees names until patient consents and seller is selected (codes still protected until accept+OTP)"),
    new_full(
        "Consent + selected PharmacyPartnerId on MedicineOrder",
        "API: list sellers; select; grant consent",
        "Patient pharmacy pick UI (web optional + app Phase 16)",
        None,
        "QA: without consent, handoff API 403",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.4", mid="MED-04",
         main="Prescription handoff — signed prescription becomes the order directly — no re-entry, no transcription errors",
         module="HomeoMeds", pdf="§12 + §2.7", feat="Prescription handoff / Order start",
         deliv="The signed prescription becomes the order directly — no re-entry, no transcription errors / Start an order directly from the prescription",
         surface="Patient + Pharmacy", as_is="Does not exist",
         to_be="MedicineOrder copied from ErxSnapshot items (codes); pharmacy never types remedies",
         wt="New", prio="P0", ac="Order lines == snapshot lines; pharmacy cannot add a different remedy"),
    new_full(
        "DB: MedicineOrder, MedicineOrderItem from ErxSnapshot",
        "API: POST order from ErxId",
        "Order start from patient eRx screen",
        None,
        "QA: line count mismatch fails sign-off",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.5", mid="MED-05",
         main="Rule-based seller routing — licence validity, service area, working hours and capacity",
         module="HomeoMeds", pdf="§12", feat="Rule-based seller routing",
         deliv="Orders routed by licence validity, service area, working hours and capacity",
         surface="Backend", as_is="Does not exist",
         to_be="SellerRoutingRule engine",
         wt="New", prio="P0", ac="Out-of-area pharmacy never listed"),
    new_full(
        "DB: SellerRoutingRule (area, hours, capacity)",
        "Routing service used by seller list API",
        "Admin/pharmacy hours and area config UI",
        None,
        "QA: capacity full excluded; hours closed excluded",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.6", mid="MED-06",
         main="Manual stock confirmation — pharmacy confirms availability — no live inventory integration required",
         module="HomeoMeds", pdf="§12", feat="Manual stock confirmation",
         deliv="The pharmacy confirms availability — no live inventory integration required",
         surface="Pharmacy console", as_is="Does not exist",
         to_be="Accept/reject with stock flag; do not build inventory master",
         wt="New", prio="P0", ac="Reject for out-of-stock routes to next seller or exception queue"),
    [
        {"sub": "Accept payload includes stock confirmation; reject reason OUT_OF_STOCK",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Pharmacy UI confirm availability",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: no inventory SKU table required",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.7", mid="MED-07",
         main="Pharmacy console — accept or reject, enter price quote, mark ready and dispatched",
         module="HomeoMeds", pdf="§12", feat="Pharmacy console",
         deliv="Accept or reject orders, enter the price quote, mark ready and dispatched",
         surface="Pharmacy console", as_is="Does not exist",
         to_be="/pharmacy/orders list+detail; OTP accept (SEC-07); quote; ready; dispatched",
         wt="New", prio="P0", ac="OTP required on accept; status visible to patient"),
    new_full(
        "Order status machine: Offered/Accepted/Rejected/Quoted/Paid/Ready/Dispatched/Delivered",
        "API: accept(OTP), reject, quote, ready, dispatch",
        "/pharmacy/orders and detail",
        "OTP on accept",
        "QA: accept without OTP fails; names still hidden to patient until accept succeeds",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.8", mid="MED-08",
         main="Quote before payment — the patient always sees the price before paying",
         module="HomeoMeds", pdf="§12 + §2.7", feat="Quote before payment / Pharmacy selection & quote",
         deliv="The patient always sees the price before paying / Choose a licensed pharmacy and see the price before paying",
         surface="Patient", as_is="Does not exist",
         to_be="MedicineQuote; patient must confirm quote; then PAY-07",
         wt="New", prio="P0", ac="Payment API rejects if quote not accepted by patient"),
    new_full(
        "DB: MedicineQuote",
        "API: quote, patient accept quote, then create medicine payment",
        "Patient quote screen",
        None,
        "QA: price change after quote invalidates old quote",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.9", mid="MED-09",
         main="Payment or cash on delivery — both options, with status tracking",
         module="HomeoMeds", pdf="§12 + §9", feat="Payment or cash on delivery",
         deliv="Both options supported, with status tracking",
         surface="Patient", as_is="Stream MEDICINE stub in Phase 6",
         to_be="Online pay via PAY-02; COD_PENDING until pharmacy confirms collection",
         wt="New", prio="P0", ac="COD does not mark Stream MEDICINE as captured until collection confirmed"),
    [
        {"sub": "Wire PAY-07 CreateMedicineOrder; COD status enum",
         "layer": "API", "be": Y, "api": Y, "int": Y, "eff": "M", "wt": "New"},
        {"sub": "Patient pay vs COD UI",
         "layer": "Web SPA", "fe": Y, "mob": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: COD vs prepaid ledger split",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.10", mid="MED-10",
         main="Order tracking — live status for the patient at every stage",
         module="HomeoMeds", pdf="§12 + §2.7", feat="Order tracking",
         deliv="Live status for the patient at every stage / Live status from accepted to delivered",
         surface="Patient", as_is="Does not exist",
         to_be="Status timeline events; push updates COM-03",
         wt="New", prio="P1", ac="Every status change visible; never silent skip"),
    new_full(
        "DB: MedicineOrderEvent",
        "GET tracking by order id (patient)",
        "Tracking UI",
        "Push on status change",
        "QA: dispatched without ready is invalid transition",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.11", mid="MED-11",
         main="Medicine ledger shares — seller, platform and delivery recorded (FIN-03 live fill)",
         module="HomeoMeds", pdf="§12 + §8", feat="Medicine ledger",
         deliv="Seller, platform and delivery shares recorded separately from consultation money",
         surface="Account", as_is="Empty medicine ledger screen from Phase 7",
         to_be="On pay/COD collect, write SplitSeller/Platform/Delivery from config",
         wt="New", prio="P0", ac="Account medicine ledger shows real rows after first order"),
    [
        {"sub": "On captured medicine payment write three-share ledger lines",
         "layer": "API", "be": Y, "db": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: consult GMV unchanged by a medicine order",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.12", mid="MED-12",
         main="Order exception queue — stuck, rejected or failed orders always surfaced — never silent",
         module="HomeoMeds", pdf="§12", feat="Order exception queue",
         deliv="Stuck, rejected or failed orders are always surfaced to the operations team — never silent",
         surface="Admin + Account", as_is="Does not exist",
         to_be="HomeoMeds exception queue; SLA; never drop",
         wt="New", prio="P0", ac="Rejected-by-all-sellers appears in queue within worker SLA"),
    new_full(
        "Exception rows for medicine orders",
        "Admin/ops list/detail/retry re-route",
        "/admin/homemeds-exceptions",
        None,
        "QA: silent discard impossible (metric + queue)",
    ),
)

add(
    dict(phase=14, pn=PN14, wbs="14.13", mid="MED-13",
         main="OTP-based control — key actions verified by OTP so Homeocentrum can monitor and trace every transaction",
         module="HomeoMeds", pdf="§12 + §17", feat="OTP-based control",
         deliv="Key actions verified by OTP so Homeocentrum can monitor and trace every transaction",
         surface="Pharmacy + Account", as_is="SEC-07 infrastructure exists",
         to_be="OTP on pharmacy accept (and other high-risk); OtpAudit queryable",
         wt="New", prio="P0", ac="Accept without OTP fails; audit lists the action"),
    [
        {"sub": "Bind OTP actions: PharmacyAccept, (optional) QuoteLock, Payout already FIN-05",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: Account OTP audit shows pharmacy accept",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.14", mid="MED-14",
         main="Refill orders — repeat orders raised from an existing prescription",
         module="HomeoMeds", pdf="§12 + §2.7 + §11", feat="Refill orders / Order detail & refill",
         deliv="Repeat orders raised from an existing prescription / Full order history and one-tap refill request",
         surface="Patient + Doctor", as_is="ERX-10 refill request exists",
         to_be="Approved refill creates a new MedicineOrder from the same snapshot; one-tap patient",
         wt="New", prio="P1", ac="Refill does not mutate the original signed snapshot"),
    [
        {"sub": "On doctor approve, clone order from snapshot; patient one-tap",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "Patient order detail + refill CTA",
         "layer": "Web SPA", "fe": Y, "mob": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: original eRx unchanged",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.15", mid="MED-15",
         main="Order review & consent (patient) — confirm and consent before the prescription is shared",
         module="HomeoMeds", pdf="§2.7", feat="Order review & consent",
         deliv="Confirm and consent before the prescription is shared",
         surface="Patient", as_is="Covered by MED-03; explicit review screen still required",
         to_be="Review screen listing codes (not names) + pharmacy + consent checkbox",
         wt="New", prio="P0", ac="Share does not occur until review confirmed"),
    [
        {"sub": "Patient review screen before handoff",
         "layer": "Web SPA", "fe": Y, "mob": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: back button does not share",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=14, pn=PN14, wbs="14.16", mid="MED-16",
         main="Medicines tab — all medicine orders in one place (API now; app UI in Phase 16)",
         module="HomeoMeds", pdf="§2.7", feat="Medicines tab",
         deliv="All medicine orders in one place",
         surface="Patient", as_is="Does not exist",
         to_be="GET my orders; website optional list; app tab Phase 16",
         wt="New", prio="P1", ac="Patient sees all orders for their PatientId only"),
    [
        {"sub": "GET /api/Patient/MedicineOrders",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Optional web list; mobile tab later",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: IDOR other patient 403",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 15 — Continuity APIs (PDF §2.5–2.8) before mobile
# =============================================================================
PN15 = "Phase 15 — Patient continuity APIs (PDF §2.5–2.8) — build before mobile UI"

add(
    dict(phase=15, pn=PN15, wbs="15.1", mid="CON-01",
         main="Family members — add family members under one account",
         module="Patient identity", pdf="§2.1", feat="Family members",
         deliv="Add family members under one account",
         surface="Patient App (API now)", as_is="Does not exist",
         to_be="FamilyMember under Patient account; bookings can be for a member",
         wt="New", prio="P0", ac="Member has own PatientId or MemberId linked; records isolated correctly"),
    new_full(
        "DB: FamilyMember",
        "API: CRUD family; book-as-member",
        "Consumed in Phase 16",
        None,
        "QA: adult consent rules for minors documented",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.2", mid="CON-02",
         main="Caregiver authorization — let a family member book and manage on their behalf",
         module="Patient identity", pdf="§2.1", feat="Caregiver authorization",
         deliv="Let a family member book and manage on their behalf",
         surface="Patient App (API now)", as_is="Does not exist",
         to_be="CaregiverAuth with scopes; OTP; withdrawable via consent centre",
         wt="New", prio="P0", ac="Caregiver cannot exceed granted scopes; withdrawal immediate"),
    new_full(
        "DB: CaregiverAuth",
        "API: grant/revoke/list; booking authorisation check",
        "Consumed in Phase 16",
        "OTP on grant",
        "QA: revoked token cannot book",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.3", mid="CON-03",
         main="Records timeline + consultation note + document upload APIs",
         module="Patient records", pdf="§2.5", feat="Records timeline / Consultation note / Document upload",
         deliv="Complete visit history in date order / Doctor’s notes for each visit / Upload reports, scans and previous prescriptions",
         surface="Patient App (API now)", as_is="Doctor history exists; no patient timeline; no patient upload",
         to_be="Patient timeline DTO (visits, summary, eRx codes, labs); consultation note = TEL-11 summary not private notes; uploads via SEC-09",
         wt="New", prio="P0", ac="Private doctor notes never in patient timeline; uploads listed"),
    new_full(
        "Reuse appointments + summary + eRx + SecureDocument",
        "GET timeline; GET consultation note (summary); POST documents",
        "Consumed in Phase 16",
        None,
        "QA: notes leak test on timeline JSON",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.4", mid="CON-04",
         main="Follow-up plan & tasks — clear next steps after each consultation",
         module="Continuity", pdf="§2.6", feat="Follow-up plan & tasks",
         deliv="Clear next steps after each consultation",
         surface="Doctor Web + Patient", as_is="Does not exist (follow-up option in notes does not save)",
         to_be="FollowUpTask created by doctor; patient sees tasks",
         wt="New", prio="P1", ac="Tasks persist against PatientAppId; due dates drive Phase 18 follow-up analysis"),
    new_full(
        "DB: FollowUpPlan, FollowUpTask",
        "API: doctor set; patient list/complete",
        "Doctor board follow-up panel; app in Phase 16",
        None,
        "QA: overdue flag",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.5", mid="CON-05",
         main="Symptom diary — record symptoms between visits",
         module="Continuity", pdf="§2.6", feat="Symptom diary",
         deliv="Record symptoms between visits",
         surface="Patient App (API now)", as_is="Does not exist",
         to_be="SymptomDiary entries; doctor can read on board",
         wt="New", prio="P1", ac="Entries dated; treating doctor can read; other doctors 403"),
    new_full(
        "DB: SymptomDiary",
        "API: patient CRUD; doctor GET by patient",
        "Doctor board read-only; app write Phase 16",
        None,
        "QA: IDOR",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.6", mid="CON-06",
         main="Progress tracking — see improvement over time across visits",
         module="Continuity", pdf="§2.6", feat="Progress tracking",
         deliv="See improvement over time across visits",
         surface="Patient + Doctor", as_is="Does not exist",
         to_be="Progress series from diary + visit outcomes; not a fake chart",
         wt="New", prio="P2", ac="Empty state for new patients; series uses real diary/visit data"),
    new_full(
        "Derived API from diary + appointments",
        "GET progress series",
        "Consumed in Phase 16",
        None,
        "QA: no dummy numbers",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.7", mid="CON-07",
         main="Consent centre & data rights — view, manage and withdraw consents; request data",
         module="Privacy", pdf="§17 + §2.8", feat="Consent centre & data rights",
         deliv="Patients can view, manage and withdraw consent / View, manage and withdraw consents; request data",
         surface="Patient App (API now)", as_is="ConsentRecord from Phase 0",
         to_be="Patient list/withdraw; data export request (DPDP); deletion request workflow",
         wt="New", prio="P0", ac="Withdraw tele-recording prevents future recording; export request ticketed"),
    new_full(
        "Reuse ConsentRecord; DataRequest table",
        "API: list/withdraw; request export/delete",
        "Consumed in Phase 16",
        None,
        "QA: withdraw marketing stops WhatsApp/SMS offers",
    ),
)

add(
    dict(phase=15, pn=PN15, wbs="15.8", mid="CON-08",
         main="Patient payments & refunds read model",
         module="Patient money", pdf="§2.8", feat="Payments & refunds",
         deliv="Full payment history and refund status",
         surface="Patient App (API now)", as_is="Ledger exists for Account",
         to_be="Patient-safe DTO of own consult+medicine payments and refunds — no other patients, no commission internals",
         wt="New", prio="P1", ac="Patient sees status matching PAY-08; no Account-only fields leaked"),
    [
        {"sub": "GET /api/Patient/Payments",
         "layer": "API", "be": Y, "api": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: commission/GST internal lines hidden",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=15, pn=PN15, wbs="15.9", mid="CON-09",
         main="Physical clinic appointment — book an in-person visit as well as online",
         module="Booking", pdf="§2.9 + §5.2", feat="Physical clinic appointment",
         deliv="Book an in-person visit as well as online",
         surface="Patient", as_is="ConsultMode added in APT-02; public booking in WEB-04",
         to_be="Ensure public/patient booking supports In-clinic as first-class, not tele-only",
         wt="Existing Improvement", prio="P1", ac="Patient can book in-clinic slot where doctor offers it"),
    [
        {"sub": "Public/patient create booking accepts ConsultMode=In-clinic; uses in-clinic fee",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "Existing Improvement"},
        {"sub": "QA: in-clinic does not require device check/video",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=15, pn=PN15, wbs="15.10", mid="CON-10",
         main="Patient profile setup API — name, age, gender, contact and health basics",
         module="Patient identity", pdf="§2.1", feat="Profile setup",
         deliv="Name, age, gender, contact and health basics",
         surface="Patient App (API now)", as_is="Clinic-created Patient exists; no self-serve profile product",
         to_be="Patient profile fields + health basics; used by OTP-verified patient token",
         wt="New", prio="P0", ac="Profile required before first paid booking (or soft-gate documented)"),
    new_full(
        "Extend Patient with health-basics JSON/columns",
        "GET/PUT /api/Patient/Profile",
        "Consumed in Phase 16",
        None,
        "QA: cannot escalate to doctor role via profile PUT",
    ),
)


# =============================================================================
# PHASE 16 — Patient mobile app (PDF §2) iOS & Android
# =============================================================================
PN16 = "Phase 16 — Patient mobile app iOS & Android (PDF §2) — consumes Phases 8–15 APIs"

def pat(wbs, mid, feat, deliv, main, pdf="§2", extra_dep="—"):
    add(
        dict(phase=16, pn=PN16, wbs=wbs, mid=mid, main=main, module="Patient Mobile",
             pdf=pdf, feat=feat, deliv=deliv, surface="Patient Mobile App",
             as_is="No patient app in repos", to_be="Native/RN screens on existing APIs; iOS and Android parity",
             wt="New", prio="P1", dep=extra_dep,
             ac=f"iOS and Android: {deliv}"),
        mobile_consume(
            f"UI: {feat} screens (empty, loading, error, offline)",
            f"Wire to the Phase 8–15 API for: {feat}",
            "Low-data assets where applicable; accessibility labels",
            f"QA both platforms: {feat} matches website/API status; no local fake paid/signed state",
        ),
    )

pat("16.1", "PAT-01", "Language selection", "Choose preferred language on first open",
    "Patient app — language selection on first open", "§2.1")
pat("16.2", "PAT-02", "Welcome & introduction", "Guided introduction to the app",
    "Patient app — welcome and introduction", "§2.1")
pat("16.3", "PAT-03", "Mobile number + OTP login", "Secure sign-in with no password to remember",
    "Patient app — mobile number + OTP login (PatientAuth)", "§2.1", "WEB-04")
pat("16.4", "PAT-04", "Profile setup", "Name, age, gender, contact and health basics",
    "Patient app — profile setup", "§2.1", "CON-10")
pat("16.5", "PAT-05", "Privacy consent", "Clear consent before any personal data is stored",
    "Patient app — privacy consent before data stored", "§2.1", "SEC-06")
pat("16.6", "PAT-06", "Family members", "Add family members under one account",
    "Patient app — family members", "§2.1", "CON-01")
pat("16.7", "PAT-07", "Caregiver authorization", "Let a family member book and manage on their behalf",
    "Patient app — caregiver authorization", "§2.1", "CON-02")
pat("16.8", "PAT-08", "Home dashboard", "Personalised home with upcoming visits and quick actions",
    "Patient app — home dashboard", "§2.2")
pat("16.9", "PAT-09", "Search & care categories", "Search by name, condition or category of care",
    "Patient app — search and care categories", "§2.2", "WEB-03")
pat("16.10", "PAT-10", "Search results", "Clear list of available doctors",
    "Patient app — search results", "§2.2", "WEB-03")
pat("16.11", "PAT-11", "Filters", "Narrow by availability, fee, experience, language and location",
    "Patient app — discovery filters", "§2.2", "WEB-03")
pat("16.12", "PAT-12", "Doctor profile", "Full profile with qualifications, experience and consultation fee",
    "Patient app — doctor profile", "§2.2", "WEB-03")
pat("16.13", "PAT-13", "Credentials & verification", "Verified badge showing the doctor’s credentials are checked",
    "Patient app — verified badge on doctor", "§2.2", "TRU-03")
pat("16.14", "PAT-14", "Ranking explanation", "Transparent explanation of why doctors appear in a given order",
    "Patient app — ranking explanation", "§2.2", "TRU-07")
pat("16.15", "PAT-15", "Health articles", "Educational content to build trust and engagement",
    "Patient app — health articles", "§2.2", "WEB-12")
pat("16.16", "PAT-16", "Slot picker", "See real available time slots and choose one",
    "Patient app — slot picker", "§2.3", "APT-08")
pat("16.17", "PAT-17", "Booking review & consent", "Confirm details and consent before booking",
    "Patient app — booking review and consent", "§2.3", "WEB-04")
pat("16.18", "PAT-18", "Checkout", "Pay the consultation fee inside the app",
    "Patient app — checkout", "§2.3", "PAY-03")
pat("16.19", "PAT-19", "Payment status", "Immediate confirmation of paid, pending or failed",
    "Patient app — payment status", "§2.3", "PAY-08")
pat("16.20", "PAT-20", "Appointment detail", "Full booking details in one place",
    "Patient app — appointment detail", "§2.3")
pat("16.21", "PAT-21", "Reschedule appointment", "Move to another slot without calling the clinic",
    "Patient app — reschedule", "§2.3", "APT-05")
pat("16.22", "PAT-22", "Cancel appointment", "Cancel with reason, refund handled per policy",
    "Patient app — cancel", "§2.3", "APT-06")
pat("16.23", "PAT-23", "Waitlist offer", "Get offered a slot when one is freed by a cancellation",
    "Patient app — waitlist offer", "§2.3", "WEB-11")
pat("16.24", "PAT-24", "Instant consult request", "Request an immediate consultation with an available doctor",
    "Patient app — instant consult request", "§2.3", "TEL-12")
pat("16.25", "PAT-25", "Queue & doctor offer", "See position in queue and accept an offered doctor",
    "Patient app — queue and doctor offer", "§2.3", "TEL-12")
pat("16.26", "PAT-26", "Device check", "Camera, microphone and network checked before the call",
    "Patient app — device check", "§2.4", "TEL-05")
pat("16.27", "PAT-27", "Waiting room", "Wait with live status until the doctor joins",
    "Patient app — waiting room", "§2.4", "TEL-06")
pat("16.28", "PAT-28", "Live video consultation", "Face-to-face consultation from home",
    "Patient app — live video", "§2.4", "TEL-04")
pat("16.29", "PAT-29", "Recording consent", "Explicit consent requested before any recording",
    "Patient app — recording consent", "§2.4", "TEL-07")
pat("16.30", "PAT-30", "Fallback & join failure handling", "Clear guidance if the call fails, with options to retry",
    "Patient app — join failure fallback", "§2.4", "TEL-09")
pat("16.31", "PAT-31", "Case-linked chat", "Chat tied to the consultation, not a loose message thread",
    "Patient app — case-linked chat", "§2.4", "TEL-10")
pat("16.32", "PAT-32", "Consultation summary", "Written summary available immediately after the call",
    "Patient app — consultation summary", "§2.4", "TEL-11")
pat("16.33", "PAT-33", "Rejoin after dropped call", "Rejoin the same consultation if the connection drops",
    "Patient app — rejoin", "§2.4", "TEL-08")
pat("16.34", "PAT-34", "Records timeline", "Complete visit history in date order",
    "Patient app — records timeline", "§2.5", "CON-03")
pat("16.35", "PAT-35", "Consultation note", "Doctor’s notes for each visit",
    "Patient app — consultation note (summary, not private notes)", "§2.5", "CON-03")
pat("16.36", "PAT-36", "Signed digital prescription", "Official prescription issued by the doctor",
    "Patient app — signed eRx (codes until pharmacy accept)", "§2.5", "ERX-06")
pat("16.37", "PAT-37", "Document upload", "Upload reports, scans and previous prescriptions",
    "Patient app — document upload", "§2.5", "CON-03")
pat("16.38", "PAT-38", "Follow-up plan & tasks", "Clear next steps after each consultation",
    "Patient app — follow-up plan and tasks", "§2.6", "CON-04")
pat("16.39", "PAT-39", "Symptom diary", "Record symptoms between visits",
    "Patient app — symptom diary", "§2.6", "CON-05")
pat("16.40", "PAT-40", "Progress tracking", "See improvement over time across visits",
    "Patient app — progress tracking", "§2.6", "CON-06")
pat("16.41", "PAT-41", "Medicines tab", "All medicine orders in one place",
    "Patient app — medicines tab", "§2.7", "MED-16")
pat("16.42", "PAT-42", "Order start", "Start an order directly from the prescription",
    "Patient app — start HomeoMeds order from eRx", "§2.7", "MED-04")
pat("16.43", "PAT-43", "Pharmacy selection & quote", "Choose a licensed pharmacy and see the price before paying",
    "Patient app — pharmacy selection and quote", "§2.7", "MED-08")
pat("16.44", "PAT-44", "Order review & consent", "Confirm and consent before the prescription is shared",
    "Patient app — order review and consent", "§2.7", "MED-15")
pat("16.45", "PAT-45", "Order tracking", "Live status from accepted to delivered",
    "Patient app — order tracking", "§2.7", "MED-10")
pat("16.46", "PAT-46", "Order detail & refill", "Full order history and one-tap refill request",
    "Patient app — order detail and refill", "§2.7", "MED-14")
pat("16.47", "PAT-47", "Write a review", "Rate and review the consultation",
    "Patient app — write a review", "§2.8", "TRU-05")
pat("16.48", "PAT-48", "My reviews & appeal", "Manage reviews and raise an appeal",
    "Patient app — my reviews and appeal", "§2.8", "TRU-06")
pat("16.49", "PAT-49", "Payments & refunds", "Full payment history and refund status",
    "Patient app — payments and refunds", "§2.8", "CON-08")
pat("16.50", "PAT-50", "Profile & settings", "Manage personal details and preferences",
    "Patient app — profile and settings", "§2.8", "CON-10")
pat("16.51", "PAT-51", "Consent centre & data rights", "View, manage and withdraw consents; request data",
    "Patient app — consent centre and data rights", "§2.8", "CON-07")
pat("16.52", "PAT-52", "Notifications centre", "All alerts in one place",
    "Patient app — notifications centre", "§2.8", "COM-05")
pat("16.53", "PAT-53", "Help centre & support tickets", "Raise an issue and track it to resolution",
    "Patient app — help centre and tickets", "§2.9", "SUP-01")
pat("16.54", "PAT-54", "Physical clinic appointment", "Book an in-person visit as well as online",
    "Patient app — physical clinic booking", "§2.9", "CON-09")
pat("16.55", "PAT-55", "Assisted booking", "Help for patients who need support booking",
    "Patient app — assisted booking request", "§2.9", "SUP-07")
pat("16.56", "PAT-56", "Low-data mode", "Works on slow or limited connections",
    "Patient app — low-data mode", "§2.9")
pat("16.57", "PAT-57", "Push notifications",
    "Booking confirmations, reminders, doctor ready, prescription issued, order updates",
    "Patient app — push notifications", "§2.9", "COM-03")

add(
    dict(phase=16, pn=PN16, wbs="16.58", mid="PAT-58",
         main="Patient app — iOS and Android store release, device matrix, and parity QA",
         module="Patient Mobile", pdf="§2 + §18", feat="Patient mobile app iOS and Android",
         deliv="iOS and Android — onboarding, discovery, booking, payment, video consultation, records, prescriptions, follow-up, medicines, reviews, support, notifications",
         surface="Patient Mobile App", as_is="No app",
         to_be="Store builds; crash-free join; feature parity checklist against every §2 row",
         wt="New", prio="P0", ac="Parity matrix signed: every §2 feature works on both OS"),
    [
        {"sub": "CI: iOS + Android release builds; env pointing to .NET 8",
         "layer": "Mobile", "mob": Y, "eff": "L", "wt": "New"},
        {"sub": "Parity QA spreadsheet: each PAT-01..57 on both OS",
         "layer": "QA", "mob": Y, "eff": "L", "wt": "Regression"},
        {"sub": "Crashlytics/Sentry equivalent for patient app",
         "layer": "Mobile", "mob": Y, "int": Y, "eff": "S", "wt": "New Integration"},
    ],
)


# =============================================================================
# PHASE 17 — Doctor mobile app (PDF §3)
# =============================================================================
PN17 = "Phase 17 — Doctor mobile app iOS & Android (PDF §3) — no case-taking on phone"

add(
    dict(phase=17, pn=PN17, wbs="17.1", mid="DMO-01",
         main="Doctor app — login & identity (same login and same account as the web portal)",
         module="Doctor Mobile", pdf="§3", feat="Login & identity",
         deliv="Same login and same account as the web portal",
         surface="Doctor Mobile App", as_is="No doctor app",
         to_be="Same Account login + DoctorId; bind device; no separate doctor identity",
         wt="New", prio="P0", ac="Web and app show the same queue for the same DoctorId"),
    mobile_consume(
        "Login screen (password or OTP-to-registered mobile)",
        "Call same Login API; store JWT securely (Keychain/Keystore)",
        "Expired session → re-login; no case-taking entry point",
        "QA: same DoctorId as web; reception accounts must not open doctor clinical actions",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.2", mid="DMO-02",
         main="Doctor app — first-run onboarding (guided setup, number confirmation and permissions)",
         module="Doctor Mobile", pdf="§3", feat="First-run onboarding",
         deliv="Guided setup, number confirmation and permissions",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Confirm registered mobile; permission primer",
         wt="New", prio="P1", ac="Onboarding cannot be skipped past permission explanation for mic/camera/push"),
    mobile_consume(
        "Onboarding screens",
        "Confirm number against profile",
        "Skip vs required steps documented",
        "QA: number mismatch blocked",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.3", mid="DMO-03",
         main="Doctor app — notification permissions (push, microphone and camera handled properly)",
         module="Doctor Mobile", pdf="§3", feat="Notification permissions",
         deliv="Push, microphone and camera permissions handled properly",
         surface="Doctor Mobile App", as_is="No app",
         to_be="OS permission flows; settings deep link if denied",
         wt="New", prio="P0", ac="Denied camera still allows queue viewing; join explains why it cannot start"),
    mobile_consume(
        "Permission primer + settings CTA",
        "Register FCM/APNs token (COM-03)",
        "Denied-state UX",
        "QA: deny each permission independently",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.4", mid="DMO-04",
         main="Doctor app — today’s queue & schedule (the full day at a glance)",
         module="Doctor Mobile", pdf="§3", feat="Today’s queue & schedule",
         deliv="The full day at a glance",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Appointment list APIs; payment + tele flags; no Patient Board",
         wt="New", prio="P0", ac="Matches web dashboard buckets for today"),
    mobile_consume(
        "TodayQueue screen",
        "Existing appointment list APIs",
        "Pull-to-refresh; empty Sunday",
        "QA: parity with web counts",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.5", mid="DMO-05",
         main="Doctor app — availability toggle (go online or offline, set working hours)",
         module="Doctor Mobile", pdf="§3", feat="Availability toggle",
         deliv="Go online or offline, set working hours",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Same TeleAvailability + schedule APIs as web",
         wt="New", prio="P0", ac="Toggle on app updates web and patient instant matching"),
    mobile_consume(
        "Availability + hours screens",
        "TEL-01 + APT-07 APIs (hours read/edit policy: edit hours may stay web-only if product decides — PDF says set working hours so include)",
        "Offline banner",
        "QA: web and app status match within heartbeat",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.6", mid="DMO-06",
         main="Doctor app — push notifications (waiting, booking, reschedule, cancellation, refill, payment received)",
         module="Doctor Mobile", pdf="§3", feat="Push notifications",
         deliv="Patient waiting, new booking, reschedule, cancellation, refill request, payment received",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Consume COM-03 doctor templates",
         wt="New", prio="P0", ac="Each listed event delivers a push when permission granted"),
    mobile_consume(
        "Notification handling + deep links into queue/refill/join",
        "Device register",
        "Foreground in-app banner",
        "QA: each event type deep link",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.7", mid="DMO-07",
         main="Doctor app — patient context card (name, age, chief complaint and last visit) — NO case-taking",
         module="Doctor Mobile", pdf="§3", feat="Patient context card",
         deliv="Name, age, chief complaint and last visit before joining",
         surface="Doctor Mobile App", as_is="No app; PDF: detailed case-taking stays on the web portal",
         to_be="Read-only card; explicitly no repertory/audio/Rx compose",
         wt="New", prio="P0", ac="No path to Patient Board clinical tools from the app"),
    mobile_consume(
        "PatientCard read-only",
        "Context API (name, age, CC, last visit, payment, tele)",
        "Do not ship case-taking tabs",
        "QA: repertory endpoints never called from app",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.8", mid="DMO-08",
         main="Doctor app — join consultation (video or audio from the phone)",
         module="Doctor Mobile", pdf="§3", feat="Join consultation",
         deliv="Start the video or audio consultation from the phone",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Same session tokens as TEL-03/04",
         wt="New", prio="P0", ac="Join same room as web patient or mobile patient"),
    mobile_consume(
        "VideoRoom",
        "Tele session tokens",
        "Audio-only fallback",
        "QA: web patient + app doctor",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.9", mid="DMO-09",
         main="Doctor app — refill approval (approve or reject repeat prescription requests)",
         module="Doctor Mobile", pdf="§3 + §11", feat="Refill approval",
         deliv="Approve or reject repeat prescription requests",
         surface="Doctor Mobile App", as_is="No app",
         to_be="ERX-10 inbox; approve only — Patient Board remains web",
         wt="New", prio="P0", ac="Approve/reject from phone writes the same RefillRequest as web"),
    mobile_consume(
        "RefillInbox + RefillDetail",
        "Approve/Reject APIs",
        "Show codes vs names per policy (doctor sees names)",
        "QA: reject reason; cannot edit snapshot",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.10", mid="DMO-10",
         main="Doctor app — earnings summary (today, this week, this month, with payout status)",
         module="Doctor Mobile", pdf="§3 + §16", feat="Earnings summary",
         deliv="Today, this week and this month, with payout status",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Account ledger read model for this DoctorId — not a second money system",
         wt="New", prio="P1", ac="Figures match Account settlement for that doctor"),
    mobile_consume(
        "Earnings screens",
        "GET Earnings/Summary",
        "Empty first week",
        "QA: matches /account figures ± rounding",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.11", mid="DMO-11",
         main="Doctor app — offline handling (weak connection or expired session)",
         module="Doctor Mobile", pdf="§3", feat="Offline handling",
         deliv="Clear behaviour on weak connection or expired session",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Queued actions policy (do not fake join); session expiry uses SEC-03",
         wt="New", prio="P1", ac="Expired JWT cannot join video; user is told to re-login"),
    mobile_consume(
        "Offline/error screens",
        "401 interceptor → logout API",
        "No silent retry of payments",
        "QA: airplane mode on join",
    ),
)

add(
    dict(phase=17, pn=PN17, wbs="17.12", mid="DMO-12",
         main="Doctor app — reliability monitoring (failures reported automatically)",
         module="Doctor Mobile", pdf="§3", feat="Reliability monitoring",
         deliv="Failures are reported automatically, not left to a phone call",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Crash/error SDK; join failures also as TeleSessionEvent",
         wt="New Integration", prio="P1", ac="A forced crash appears in the monitoring tool within SLA"),
    [
        {"sub": "Integrate Crashlytics/Sentry; map to doctorId hashed",
         "layer": "Mobile", "mob": Y, "int": Y, "eff": "S", "wt": "New Integration"},
        {"sub": "QA: simulated crash + join fail event",
         "layer": "QA", "mob": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=17, pn=PN17, wbs="17.13", mid="DMO-13",
         main="Doctor app — iOS and Android parity QA (queue, availability, notifications, context, video, refill, earnings)",
         module="Doctor Mobile", pdf="§3 + §18", feat="Doctor mobile app iOS and Android",
         deliv="iOS and Android — queue, availability, notifications, patient context, video consultation, refill approval, earnings",
         surface="Doctor Mobile App", as_is="No app",
         to_be="Parity checklist for every §3 row on both OS",
         wt="New", prio="P0", ac="Parity matrix signed"),
    [
        {"sub": "Release builds both OS",
         "layer": "Mobile", "mob": Y, "eff": "L", "wt": "New"},
        {"sub": "Parity QA DMO-01..12 both OS",
         "layer": "QA", "mob": Y, "eff": "L", "wt": "Regression"},
    ],
)


# =============================================================================
# PHASE 18 — Reports & analytics (PDF §16 + admin dashboard)
# =============================================================================
PN18 = "Phase 18 — Reports & analytics (PDF §16) — after data exists"

add(
    dict(phase=18, pn=PN18, wbs="18.1", mid="RPT-01",
         main="Admin overview dashboard — platform-wide activity at a glance (replace Velzon ecommerce)",
         module="Reports", pdf="§16 + §7.2", feat="Admin overview dashboard / Admin dashboard",
         deliv="Platform-wide activity at a glance / Live overview of the whole platform",
         surface="Admin Portal", as_is="Velzon ecommerce dashboard 0% product; login lands /dashboard",
         to_be="True /admin/dashboard KPIs: doctors, credentialing, bookings, GMV, failed payments, tickets, HomeoMeds exceptions; redirect /dashboard away from StoreVisits",
         wt="Existing Improvement", prio="P0", ac="Admin home is product KPIs, not Velzon Revenue/StoreVisits"),
    improve_rows(
        "No ecommerce tables",
        "GET /api/AdminDashboard/Overview",
        "Replace pages/Admin/Dashboard; getHomeDashboardPath → /admin/dashboard",
        "QA: login as admin never shows fake store visits",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.2", mid="RPT-02",
         main="Doctor practice statistics — patient and consultation trends (extend existing charts)",
         module="Reports", pdf="§16 + §5.1", feat="Doctor practice statistics",
         deliv="Patient and consultation trends",
         surface="Doctor Web", as_is="Patient stats charts live",
         to_be="Keep as foundation; ensure visit type/mode included after APT-02",
         wt="Existing Improvement", prio="P1", ac="Trends include first vs follow-up when VisitType exists"),
    improve_rows(
        "No new warehouse required for v1",
        "Extend GetPatientStatsCharts with VisitType/ConsultMode",
        "Dashboard charts consume new series",
        "QA: empty new clinic",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.3", mid="RPT-03",
         main="Follow-up analysis — due/overdue patients and follow-up conversion",
         module="Reports", pdf="§16", feat="Follow-up analysis",
         deliv="Which patients are due or overdue for follow-up, and follow-up conversion",
         surface="Doctor Web", as_is="Does not exist; follow-up option in notes does not save",
         to_be="/doctor/follow-up-analysis from VisitType + CON-04 tasks",
         wt="New", prio="P1", ac="Due list and conversion first-visit → follow-up are real counts"),
    new_full(
        "Queries on VisitType + FollowUpTask",
        "GET FollowUpDue, FollowUpSummary",
        "/doctor/follow-up-analysis",
        None,
        "QA: overdue vs due buckets",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.4", mid="RPT-04",
         main="Clinic performance analysis — appointments, slot utilisation, no-shows, cancellations, waiting time and revenue",
         module="Reports", pdf="§16", feat="Clinic performance analysis",
         deliv="Appointments, slot utilisation, no-shows, cancellations, waiting time and revenue",
         surface="Doctor Web", as_is="Does not exist",
         to_be="/doctor/clinic-performance including payment collected and tele duration if timed",
         wt="New", prio="P1", ac="Each KPI has a definition note on the screen (no silent formula)"),
    new_full(
        "Aggregate APIs beyond GetPatientStatsCharts",
        "GET ClinicPerformance",
        "/doctor/clinic-performance",
        None,
        "QA: cancelled excluded from completed; revenue uses ledger not client",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.5", mid="RPT-05",
         main="Payment reconciliation report — bookings matched against payments",
         module="Reports", pdf="§16 + §7.2", feat="Payment reconciliation report / Consult booking & payment reconciliation",
         deliv="Bookings matched against payments / Match every booking against every payment",
         surface="Admin + Account", as_is="FIN-02 built the queue; this is the report/export",
         to_be="Exportable report from the same reconciliation query",
         wt="New", prio="P1", ac="CSV totals match on-screen KPIs"),
    [
        {"sub": "CSV/PDF export of FIN-02 query",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Report UI on admin/account recon screens",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: unmatched rows included",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=18, pn=PN18, wbs="18.6", mid="RPT-06",
         main="Settlement & payout reports — amounts due and paid to doctors and pharmacies",
         module="Reports", pdf="§16 + §8", feat="Settlement & payout reports",
         deliv="Amounts due and paid to doctors and pharmacies",
         surface="Account", as_is="FIN-04/05 exist as ops screens",
         to_be="Formal report/export of runs and payouts",
         wt="New", prio="P1", ac="Paid vs due columns reconcile to ledger"),
    [
        {"sub": "Export settlement runs + payout register",
         "layer": "API", "be": Y, "api": Y, "eff": "S", "wt": "New"},
        {"sub": "Report UI",
         "layer": "Web SPA", "fe": Y, "eff": "S", "wt": "New"},
        {"sub": "QA: pharmacy lines appear only after HomeoMeds",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=18, pn=PN18, wbs="18.7", mid="RPT-07",
         main="GST & invoice reports (reporting surface of FIN-08)",
         module="Reports", pdf="§16 + §8", feat="GST & invoice reports",
         deliv="Tax reporting",
         surface="Account", as_is="FIN-08 built tax screen",
         to_be="Confirm report pack (GSTR-style export as agreed with CA) — configuration",
         wt="New", prio="P1", ac="Invoice list export matches FIN-08"),
    [
        {"sub": "Confirm CA format; implement export if not already in FIN-08",
         "layer": "API", "be": Y, "eff": "M", "wt": "New"},
        {"sub": "QA: number series gaps flagged",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=18, pn=PN18, wbs="18.8", mid="RPT-08",
         main="Doctor earnings summary — daily, weekly and monthly earnings with payout status (web + same API as DMO-10)",
         module="Reports", pdf="§16 + §3", feat="Doctor earnings summary",
         deliv="Daily, weekly and monthly earnings with payout status",
         surface="Doctor Web + App", as_is="Does not exist",
         to_be="Web earnings page using same API as doctor app",
         wt="New", prio="P1", ac="Web and app match"),
    new_full(
        "Reuse earnings API",
        "GET Earnings/Summary already for app — add if missing date buckets",
        "/doctor/earnings",
        None,
        "QA: payout pending vs paid",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.9", mid="RPT-09",
         main="Medicine order reports — order volumes, fulfilment and exceptions",
         module="Reports", pdf="§16 + §12", feat="Medicine order reports",
         deliv="Order volumes, fulfilment and exceptions",
         surface="Admin + Account", as_is="Does not exist",
         to_be="Volumes, SLA, exception counts from HomeoMeds",
         wt="New", prio="P1", ac="Exception count matches MED-12 queue"),
    new_full(
        "Aggregates on MedicineOrder",
        "GET medicine order report",
        "Admin/account report page",
        None,
        "QA: cancelled orders not counted as delivered",
    ),
)

add(
    dict(phase=18, pn=PN18, wbs="18.10", mid="RPT-10",
         main="Platform users remaining 20% — import, export and activation control",
         module="Admin", pdf="§7.2", feat="Platform users",
         deliv="Manage all users with import, export and activation control",
         surface="Admin Portal", as_is="List/Add/Edit live; Import/Export buttons dead; no verification column",
         to_be="Wire Import/Export; verification column; activate/lock; link credentialing",
         wt="Existing Improvement", prio="P1", ac="Import/Export work; unverified visible; lock prevents login"),
    improve_rows(
        "VerificationStatus on user/doctor list DTO",
        "User import/export APIs if missing",
        "ListUser.js wire dead buttons; verification column",
        "QA: export contains only permitted fields",
    ),
)


# =============================================================================
# PHASE 19 — Cross-surface consistency, NFR, summary gate (PDF §1 + §18)
# =============================================================================
PN19 = "Phase 19 — Cross-surface QA, NFR, delivery summary gate (PDF §1 + §18)"

add(
    dict(phase=19, pn=PN19, wbs="19.1", mid="QA-01",
         main="One ecosystem proof — patient, appointment, prescription and payment created anywhere are visible everywhere",
         module="QA", pdf="§1 + §18", feat="One connected ecosystem",
         deliv="A patient, an appointment, a prescription and a payment created anywhere in the system are visible everywhere else",
         surface="All", as_is="Surfaces were built in phases",
         to_be="End-to-end scripts covering all six user types",
         wt="Regression", prio="P0", ac="Signed E2E pack: 6 roles, no dual ledgers, no dual appointments"),
    [
        {"sub": "E2E: Reception creates patient+appt → Doctor board → sign eRx → Patient app sees codes → Pharmacy accept OTP → patient sees names+pay → Account ledger S2 and S5",
         "layer": "QA", "fe": Y, "be": Y, "mob": Y, "eff": "L", "wt": "Regression"},
        {"sub": "E2E: Patient app books+pays tele → doctor app join → summary+eRx → refill on doctor app",
         "layer": "QA", "mob": Y, "eff": "L", "wt": "Regression"},
        {"sub": "Negative: Admin cannot payout; Account cannot edit repertory",
         "layer": "QA", "be": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=19, pn=PN19, wbs="19.2", mid="QA-02",
         main="Delivery summary gate — every area in PDF §18 signed off",
         module="QA", pdf="§18", feat="Summary of delivery",
         deliv="All features listed in this document are included in the delivery scope.",
         surface="Programme", as_is="This register is the scope",
         to_be="Walk the coverage sheet; zero PDF rows with Status≠Done unless formally deferred with client sign",
         wt="Regression", prio="P0", ac="Coverage sheet 100% mapped; deferred items have written client sign"),
    [
        {"sub": "Traceability review: 01_PDF_Coverage every row has Main Task ID and final Status",
         "layer": "QA", "eff": "M", "wt": "Regression"},
        {"sub": "Do not ship Velzon demo tickets/KYC/ecommerce as product",
         "layer": "QA", "fe": Y, "eff": "S", "wt": "Regression"},
    ],
)

add(
    dict(phase=19, pn=PN19, wbs="19.3", mid="NFR-01",
         main="NFR — performance, low-data, reliability, secrets, dual-API freeze",
         module="NFR", pdf="§2.9 + §3 + §17", feat="Low-data mode / Reliability monitoring / Secure login",
         deliv="Works on slow connections; failures reported automatically; credentials protected",
         surface="All", as_is="Known risks: plaintext passwords (fixed P0), secrets in appsettings, dual API drift",
         to_be="Rotate secrets; no new third API; patient low-data verified; doctor reliability SDK live",
         wt="Existing Improvement", prio="P0", ac="Load test on booking+webhook; secrets not in git; dual-API map documented"),
    [
        {"sub": "Rotate API secrets; keys in config/KeyVault; directory browsing remains off",
         "layer": "Security", "be": Y, "eff": "M", "wt": "Existing Modification"},
        {"sub": "Performance: slot lock under concurrent public booking",
         "layer": "QA", "be": Y, "eff": "M", "wt": "Regression"},
        {"sub": "Patient low-data mode QA on 2G profile",
         "layer": "QA", "mob": Y, "eff": "S", "wt": "Regression"},
    ],
)


# =============================================================================
# EXCEL WRITER
# =============================================================================
STATUS_VALUES = "Not Started,In Progress,Blocked,Done,Deferred"
WT_VALUES = "Existing,Existing Improvement,Existing Modification,New,New Integration,Configuration,Data Migration,Client Decision,Regression"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_header_row(ws, row, fill, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN


def write_cover(wb):
    ws = wb.active
    ws.title = "00_Cover"
    ws.sheet_properties.tabColor = "1B3A4B"
    ws.merge_cells("B2:G2")
    ws["B2"] = "HOMEOCENTRUM — FEATURE DELIVERY TASK REGISTER"
    ws["B2"].font = FONT_TITLE
    ws.merge_cells("B3:G3")
    ws["B3"] = "Source: Homeocentrum_Feature_Delivery_Document.pdf  ·  Prepared for Homeocentrum / NIGA  ·  29 August 2026"
    ws["B3"].font = FONT_S
    ws.merge_cells("B5:G5")
    ws["B5"] = "How to use this workbook for development"
    ws["B5"].font = FONT_WHITE_MD
    ws["B5"].fill = FILLS["header"]
    rules = [
        "1. Develop strictly by Seq on sheet 03_Sequential_Register. Do not skip a Seq because a later PDF section looks more exciting (Patient App is Phase 16 because it consumes APIs built in Phases 0–15).",
        "2. One PDF feature = one Main Task. Sub Tasks are the actual development steps (DB → API → UI/Mobile → Integration → QA).",
        "3. Work Type tells you whether to KEEP, IMPROVE, CHANGE, or BUILD NEW. Existing = do not rebuild. Existing Improvement = finish what is stub/partial. Existing Modification = behaviour must change. New = not in the three repos today.",
        "4. Depends On must be Done (or explicitly waived) before starting a row. Status dropdown: Not Started / In Progress / Blocked / Done / Deferred.",
        "5. Sheet 01_PDF_Coverage is the line-by-line trace to the PDF. If a PDF feature is missing there, it is a defect in this register — add it before coding.",
        "6. Shared APIs are built once on the earliest surface (usually Doctor/Admin/Account). Later Patient Website and Mobile rows are clients of those APIs, not a second money or appointment system.",
        "7. Five web portals in the PDF: Patient Website, Doctor Web, Reception, Admin, Account. Two mobile apps: Patient, Doctor. Pharmacy console is part of HomeoMeds (PDF §12), not a sixth portal in the §1 count.",
        "8. All new domain modules go on NigaHomeopathy-API (.NET 8). Do not create a third API. Classic API stays only where login / Rx write / Razorpay already live until ported (Phase 6 ports payments).",
        "9. Account controls money. Admin controls platform and clinical data. Neither performs the other’s role (PDF §8).",
        "10. Webhook/bank is payment source of truth — never the patient’s screen (PDF §9).",
    ]
    r = 6
    for line in rules:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(r, 2, line).alignment = WRAP
        ws.cell(r, 2).font = FONT_N
        ws.row_dimensions[r].height = 36
        r += 1
    r += 1
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, "Development sequence (why this order)").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    phases_why = [
        ("PRE", "Client gates", "Settlement model, cash policy, GST, vendors, legal copy — block money and telemedicine if unsigned."),
        ("0", "Foundation", "Shared keys, six roles, login/reset/session/RBAC, consent, OTP, audit, secure documents (PDF §1 + §17)."),
        ("1", "Admin clinical masters", "Keep live repertory/MM/diagnosis/drugs/questions/3D (PDF §7.1) — do not rebuild."),
        ("2", "Doctor clinical workspace", "Patient Board remaining work + Center of Gravity (PDF §5.3)."),
        ("3", "Doctor daily ops", "Dashboard, patients, reception staff UI, real doctor profile (PDF §5.1)."),
        ("4", "Appointments", "Visit type/mode, formal reschedule/cancel, schedule screen, payment status column (PDF §5.2)."),
        ("5", "Reception portal", "Front-desk chrome, queue, case paper, schedule view (PDF §6)."),
        ("6", "Payment system", "Homeocentrum merchant of record, webhook, consult pay, reception collect, refunds, invoices (PDF §9)."),
        ("7", "Account department", "Ledger, recon, settlements, OTP payouts, SoD (PDF §8)."),
        ("8", "Trust", "Credentialing and practice gate BEFORE public directory (PDF §13)."),
        ("9", "Patient website", "Discovery, self-booking, public pay, legal, waitlist (PDF §4)."),
        ("10", "Digital prescription", "Potency, signed lock, notes≠eRx, codes (PDF §11)."),
        ("11", "Telemedicine", "Availability, video, consent, rejoin, instant (PDF §10)."),
        ("12", "Communications", "SMS, WhatsApp remaining, push, email, in-app (PDF §14)."),
        ("13", "Support", "Tickets, enquiry inbox, help, assisted booking (PDF §15)."),
        ("14", "HomeoMeds", "Pharmacy network, routing, quotes, COD, exceptions (PDF §12)."),
        ("15", "Patient continuity APIs", "Family, caregiver, diary, follow-up, consent centre — before the app."),
        ("16", "Patient mobile", "Every PDF §2 row as a screen on those APIs."),
        ("17", "Doctor mobile", "Every PDF §3 row — no case-taking on phone."),
        ("18", "Reports", "Every PDF §16 report after data exists."),
        ("19", "Delivery gate", "E2E ecosystem proof + §18 summary sign-off."),
    ]
    ws.cell(r, 2, "Phase").font = FONT_WHITE
    ws.cell(r, 3, "Name").font = FONT_WHITE
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
    ws.cell(r, 4, "Why this phase sits here").font = FONT_WHITE
    for c in range(2, 8):
        ws.cell(r, c).fill = FILLS["header"]
        ws.cell(r, c).border = THIN
    r += 1
    for ph, name, why in phases_why:
        ws.cell(r, 2, ph).alignment = CENTER
        ws.cell(r, 2).fill = PHASE_FILL.get(str(ph), FILLS["section"])
        ws.cell(r, 2).font = FONT_WHITE
        ws.cell(r, 3, name).font = FONT_BOLD
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7)
        ws.cell(r, 4, why).alignment = WRAP
        ws.row_dimensions[r].height = 32
        for c in range(2, 8):
            ws.cell(r, c).border = THIN
        r += 1
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, "Work Type legend").font = FONT_WHITE
    ws.cell(r, 2).fill = FILLS["header"]
    r += 1
    legends = [
        ("Existing", "Live in Admin/Doctor/Reception today. Keep. Regression + ACL only — do not rebuild."),
        ("Existing Improvement", "Screen or API exists but is stub, fake-wired, partial, or missing a role surface."),
        ("Existing Modification", "Exists but behaviour must change (plaintext passwords, auto-activate doctors, client-only Razorpay success)."),
        ("New", "No product screen/API in the three repos."),
        ("New Integration", "Vendor/gateway (Razorpay webhook, video SDK, SMS, FCM/APNs)."),
        ("Configuration", "Rates, flags, templates — not hardcoded."),
        ("Data Migration", "Additive schema / password hash backfill."),
        ("Client Decision", "Blocks engineering until signed (PRE phase)."),
        ("Regression", "Prove an Existing or earlier-phase surface still works."),
    ]
    for wt, txt in legends:
        ws.cell(r, 2, wt).fill = TYPE_FILL.get(wt, FILLS["white"])
        ws.cell(r, 2).font = FONT_BOLD
        ws.cell(r, 2).border = THIN
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(r, 3, txt).alignment = WRAP
        ws.cell(r, 3).border = THIN
        ws.row_dimensions[r].height = 28
        r += 1
    r += 2
    n_sub = len(tasks)
    n_main = len({t["main_id"] for t in tasks})
    n_cov = len(coverage)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, f"Generated counts — Main tasks: {n_main}  ·  Sub tasks: {n_sub}  ·  PDF feature lines mapped: {n_cov}").font = FONT_H
    set_widths(ws, [4, 28, 28, 22, 22, 22, 22])
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "B5"
    ws.print_title_rows = "1:4"
    return ws


def write_coverage(wb):
    ws = wb.create_sheet("01_PDF_Coverage")
    ws.sheet_properties.tabColor = "B9770E"
    headers = [
        "Seq", "PDF Section", "PDF Feature (exact)", "PDF Delivery Text",
        "Work Type", "Phase", "Main Task ID", "AS-IS in codebase", "Status",
    ]
    ws.merge_cells("A1:I1")
    ws["A1"] = "LINE-BY-LINE TRACEABILITY — every feature row in Homeocentrum_Feature_Delivery_Document.pdf. Do not start Phase 19 until every row is Done or Deferred-with-sign."
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws["A1"].alignment = WRAP
    ws.row_dimensions[1].height = 36
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["section"]
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    for i, row in enumerate(coverage, 1):
        vals = [
            i, row["section"], row["feature"], row["delivered"],
            row["work_type"], row["phase"], row["main_id"], row["as_is"], "Not Started",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, c, v)
            cell.font = FONT_N
            cell.alignment = WRAP
            cell.border = THIN
            if c == 5:
                cell.fill = TYPE_FILL.get(v, FILLS["white"])
            elif i % 2 == 0:
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i + 2].height = 48
    last = len(coverage) + 2
    dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"I3:I{last}")
    set_widths(ws, [8, 16, 42, 55, 24, 10, 14, 55, 14])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:I{last}"
    return ws


def write_mains(wb):
    ws = wb.create_sheet("02_Main_Tasks")
    ws.sheet_properties.tabColor = "1A5276"
    headers = [
        "Seq", "Phase", "Phase Name", "WBS", "Main Task ID", "Main Task",
        "Work Type", "Module", "PDF Section", "PDF Feature", "Surface",
        "Sub-task count", "Priority", "AS-IS", "TO-BE", "Status",
    ]
    ws.merge_cells("A1:P1")
    ws["A1"] = "MAIN TASK INDEX — one row per PDF feature. Open 03_Sequential_Register and filter Main Task ID to execute sub-tasks in order."
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["section"]
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    grouped = []
    seen = {}
    for t in tasks:
        mid = t["main_id"]
        if mid not in seen:
            seen[mid] = {**t, "n": 0}
            grouped.append(seen[mid])
        seen[mid]["n"] += 1
    for i, t in enumerate(grouped, 1):
        vals = [
            i, t["phase"], t["phase_name"], t["wbs"], t["main_id"], t["main"],
            t["work_type"], t["module"], t["pdf_section"], t["pdf_feature"], t["surface"],
            t["n"], t["prio"], t["existing"], t["gap"], "Not Started",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, c, v)
            cell.font = FONT_N if c != 6 else FONT_BOLD
            cell.alignment = WRAP
            cell.border = THIN
            if c == 2:
                cell.fill = PHASE_FILL.get(str(v), FILLS["white"])
                cell.font = FONT_WHITE
            elif c == 7:
                cell.fill = TYPE_FILL.get(v, FILLS["white"])
            elif i % 2 == 0:
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i + 2].height = 42
    last = len(grouped) + 2
    dv = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"P3:P{last}")
    set_widths(ws, [8, 10, 42, 10, 14, 62, 22, 22, 16, 36, 28, 12, 10, 42, 42, 14])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:P{last}"
    return ws


def write_register(wb):
    ws = wb.create_sheet("03_Sequential_Register")
    ws.sheet_properties.tabColor = "117A65"
    ws.merge_cells("A1:AD1")
    ws["A1"] = "SEQUENTIAL DEVELOPMENT REGISTER — execute top to bottom by Seq. Filter Phase / Work Type / Surface as needed. Status is the only column developers should edit daily."
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(2, i, h)
        cell.fill = FILLS["section"]
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
    for i, t in enumerate(tasks, 1):
        vals = [
            i, t["phase"], t["phase_name"], t["wbs"], t["main_id"], t["main"],
            t["sub_id"], t["sub"], t["work_type"], t["module"], t["pdf_section"],
            t["pdf_feature"], t["pdf_delivery"], t["surface"], t["layer"],
            t["fe"], t["be"], t["db"], t["api"], t["mobile"], t["integ"],
            t["prio"], t["effort"], t["depends"], t["existing"], t["gap"],
            t["accept"], "Not Started", "", t["notes"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 2, c, v)
            cell.font = FONT_N
            cell.alignment = WRAP if c in (3, 6, 8, 12, 13, 25, 26, 27, 30) else CENTER if c in (1, 2, 16, 17, 18, 19, 20, 21, 22, 23, 28) else WRAP
            cell.border = THIN
            if c == 2:
                cell.fill = PHASE_FILL.get(str(t["phase"]), FILLS["white"])
                cell.font = FONT_WHITE
            elif c == 9:
                cell.fill = TYPE_FILL.get(t["work_type"], FILLS["white"])
            elif i % 2 == 0 and c not in (2, 9):
                cell.fill = FILLS["alt"]
        ws.row_dimensions[i + 2].height = 48
    last = len(tasks) + 2
    dv_s = DataValidation(type="list", formula1=f'"{STATUS_VALUES}"', allow_blank=False)
    dv_w = DataValidation(type="list", formula1=f'"{WT_VALUES}"', allow_blank=False)
    dv_p = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=False)
    dv_e = DataValidation(type="list", formula1='"S,M,L,XL"', allow_blank=False)
    ws.add_data_validation(dv_s)
    ws.add_data_validation(dv_w)
    ws.add_data_validation(dv_p)
    ws.add_data_validation(dv_e)
    dv_s.add(f"AB3:AB{last}")
    dv_w.add(f"I3:I{last}")
    dv_p.add(f"V3:V{last}")
    dv_e.add(f"W3:W{last}")
    widths = [8, 8, 42, 10, 14, 55, 16, 62, 22, 20, 14, 36, 48, 28, 16,
              10, 10, 10, 10, 10, 12, 10, 10, 22, 42, 42, 42, 14, 16, 28]
    set_widths(ws, widths)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:AD{last}"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}{last}"
    return ws


def write_dashboard(wb):
    ws = wb.create_sheet("04_Dashboard")
    ws.sheet_properties.tabColor = "6C3483"
    ws.merge_cells("A1:F1")
    ws["A1"] = "COUNTS FOR PLANNING — source = generated tasks (not live Status). Refresh by re-running the generator after edits to the Python source."
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    # by phase
    ws["A3"] = "By phase"
    ws["A3"].font = FONT_WHITE
    ws["A3"].fill = FILLS["section"]
    ws.merge_cells("A3:C3")
    ws["A4"] = "Phase"
    ws["B4"] = "Phase name"
    ws["C4"] = "Sub tasks"
    for c in range(1, 4):
        ws.cell(4, c).fill = FILLS["header"]
        ws.cell(4, c).font = FONT_WHITE
        ws.cell(4, c).border = THIN
    by_phase = []
    seen_ph = []
    counts_ph = Counter(t["phase"] for t in tasks)
    names_ph = {}
    for t in tasks:
        names_ph.setdefault(t["phase"], t["phase_name"])
        if t["phase"] not in seen_ph:
            seen_ph.append(t["phase"])
    r = 5
    for ph in seen_ph:
        ws.cell(r, 1, ph).fill = PHASE_FILL.get(str(ph), FILLS["white"])
        ws.cell(r, 1).font = FONT_WHITE
        ws.cell(r, 1).border = THIN
        ws.cell(r, 2, names_ph[ph]).alignment = WRAP
        ws.cell(r, 2).border = THIN
        ws.cell(r, 3, counts_ph[ph]).border = THIN
        ws.cell(r, 3).alignment = CENTER
        r += 1
    phase_end = r - 1
    # by work type
    ws["E3"] = "By work type"
    ws["E3"].font = FONT_WHITE
    ws["E3"].fill = FILLS["section"]
    ws.merge_cells("E3:F3")
    ws["E4"] = "Work Type"
    ws["F4"] = "Sub tasks"
    ws["E4"].fill = FILLS["header"]
    ws["F4"].fill = FILLS["header"]
    ws["E4"].font = FONT_WHITE
    ws["F4"].font = FONT_WHITE
    wt_counts = Counter(t["work_type"] for t in tasks)
    r = 5
    for wt, n in wt_counts.most_common():
        ws.cell(r, 5, wt).fill = TYPE_FILL.get(wt, FILLS["white"])
        ws.cell(r, 5).border = THIN
        ws.cell(r, 6, n).border = THIN
        ws.cell(r, 6).alignment = CENTER
        r += 1
    wt_end = r - 1
    pie = PieChart()
    pie.title = "Sub-tasks by work type"
    labels = Reference(ws, min_col=5, min_row=5, max_row=wt_end)
    data = Reference(ws, min_col=6, min_row=4, max_row=wt_end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False
    pie.width = 14
    pie.height = 8
    ws.add_chart(pie, "H3")
    bar = BarChart()
    bar.type = "col"
    bar.title = "Sub-tasks by phase"
    bar.y_axis.title = "Sub tasks"
    bar.x_axis.title = "Phase"
    data_b = Reference(ws, min_col=3, min_row=4, max_row=phase_end)
    cats_b = Reference(ws, min_col=1, min_row=5, max_row=phase_end)
    bar.add_data(data_b, titles_from_data=True)
    bar.set_categories(cats_b)
    bar.shape = 4
    bar.width = 18
    bar.height = 8
    ws.add_chart(bar, "H20")
    # by surface
    r = wt_end + 2
    ws.cell(r, 5, "By application surface").font = FONT_WHITE
    ws.cell(r, 5).fill = FILLS["section"]
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    r += 1
    ws.cell(r, 5, "Surface").font = FONT_WHITE
    ws.cell(r, 6, "Sub tasks").font = FONT_WHITE
    ws.cell(r, 5).fill = FILLS["header"]
    ws.cell(r, 6).fill = FILLS["header"]
    r += 1
    for s, n in Counter(t["surface"] for t in tasks).most_common():
        ws.cell(r, 5, s).border = THIN
        ws.cell(r, 6, n).border = THIN
        ws.cell(r, 6).alignment = CENTER
        r += 1
    set_widths(ws, [12, 55, 14, 4, 28, 14])
    return ws


def write_index(wb):
    ws = wb.create_sheet("05_Screen_API_DB")
    ws.sheet_properties.tabColor = "1F618D"
    ws.merge_cells("A1:D1")
    ws["A1"] = "Implementation index — screens, APIs and tables the sequential register implements. Design sign-off still required on exact paths."
    ws["A1"].font = FONT_WHITE_MD
    ws["A1"].fill = FILLS["header"]
    ws.row_dimensions[1].height = 28
    blocks = [
        ("NEW / CHANGED SCREENS", [
            ("Admin", "KPI home · Credentialing queue/detail · Users import/export · Consult recon · Payment exceptions · Support tickets · Enquiries inbox · Pharmacy activation · HomeoMeds exceptions · Help articles"),
            ("Account", "Ledger · Consult recon · Medicine ledger · Exceptions · Settlements · Payouts · Refunds · Tax · Payees · Clinic collections"),
            ("Doctor web", "Real profile · Schedule · Reception staff · Reschedule/Cancel modals · Follow-up analysis · Clinic performance · Earnings · SMS templates · Tele room · COG panel · eRx sign · Visit notes split · Support"),
            ("Reception", "Reception home/queue · Profile · Read-only schedule · Case paper · Collect payment"),
            ("Public website", "Home CTAs · /book · slots · confirm · pay · doctor profile · articles · privacy/terms · register docs"),
            ("Pharmacy", "Onboard · Order inbox · Order detail (OTP accept, quote, dispatch)"),
            ("Patient app", "Every PAT-01..57 screen in PDF §2"),
            ("Doctor app", "Every DMO-01..12 screen in PDF §3 — no case-taking"),
        ]),
        ("PROPOSED APIs (target contracts)", [
            ("Auth / profile", "Logout · Forgot/Reset/ChangePassword · Profile/Me · Profile/Photo · PatientAuth RequestOtp/VerifyOtp"),
            ("Credentialing", "Queue · {id} · Approve · Reject · RequestInfo · Documents · MyStatus · RegisterDoctor (modify)"),
            ("Payments", "Orders · Webhook/Razorpay · ConsultOrder · CollectAtReception · Exceptions · Refunds · Invoices · Earnings/Summary"),
            ("Ledger", "Ledger · MedicineLedger · Settlements · ApprovePayout (OTP) · Payees · ClinicCollections"),
            ("Public booking", "Doctors · Profile · Slots · Create · Waitlist · InstantConsult"),
            ("Appointments", "RescheduleAppointment · CancelAppointment · ChangeLog · VisitType PATCH"),
            ("eRx", "PotencyMaster · SignErx · Erx/ByAppointment · Export PDF · RefillInbox · Approve/Reject"),
            ("Telemedicine", "Availability · Queue · Sessions · End · Rejoin · Consent · Chat · Summary"),
            ("HomeoMeds", "Onboard · Licence · Sellers · Orders · Accept OTP · Quote · Pay · Status · Exceptions · Refill"),
            ("Support / comms", "SupportTicket · SMS Templates/Send · Devices/Register · Notifications"),
            ("Continuity", "Family · Caregiver · Timeline · Diary · FollowUp · Consents · Patient Payments"),
        ]),
        ("NEW / MODIFIED TABLES", [
            ("Modified existing", "UserMaster (hash) · Doctor (fee/KYC/verification) · Patient · PatientAppointment (VisitType, ConsultMode, PaymentStatus, CANCELLED) · AppointmentHistoryNote (NoteType) · PrescriptionRemedyDetail (PotencyId) · Role/Menu · EnquiryDetail · UserLoginStatus · PackageEntryDetail UNCHANGED shape"),
            ("Trust / identity", "DoctorVerification · DoctorCredentialDocument · FamilyMember · CaregiverAuth · ConsentRecord · Review · ReviewAppeal"),
            ("Money", "PaymentOrder · PaymentEvent · LedgerEntry · ConsultPayment · ConsultFeeConfig · PaymentException · SettlementRun · Payout · Refund · Invoice · Payee"),
            ("Control", "OtpChallenge · OtpAuditLog · PasswordResetToken · AuditEvent · SecureDocument"),
            ("Clinical / eRx / tele", "PotencyMaster · ErxSnapshot · ErxSnapshotItem · RefillRequest · TeleAvailability · TeleSession · TeleChatMessage · ConsultationSummary · FollowUpTask · SymptomDiary"),
            ("HomeoMeds", "PharmacyPartner · PharmacyLicence · SellerRoutingRule · MedicineOrder · MedicineOrderItem · MedicineQuote · MedicineOrderEvent"),
            ("Messaging / support", "SmsTemplate · SmsMessageLog · DeviceToken · AppNotification · SupportTicket · SupportTicketMessage · HelpArticle"),
        ]),
    ]
    r = 3
    for title, rows in blocks:
        ws.cell(r, 1, title).font = FONT_WHITE
        ws.cell(r, 1).fill = FILLS["section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        ws.cell(r, 1, "Area").font = FONT_WHITE
        ws.cell(r, 1).fill = FILLS["header"]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(r, 2, "Items").font = FONT_WHITE
        ws.cell(r, 2).fill = FILLS["header"]
        r += 1
        for area, items in rows:
            ws.cell(r, 1, area).font = FONT_BOLD
            ws.cell(r, 1).border = THIN
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
            ws.cell(r, 2, items).alignment = WRAP
            ws.cell(r, 2).border = THIN
            ws.row_dimensions[r].height = 40
            r += 1
        r += 1
    set_widths(ws, [28, 40, 20, 40])
    ws.freeze_panes = "A3"
    return ws


def main():
    assert tasks, "No tasks generated"
    assert coverage, "No PDF coverage rows"
    wb = Workbook()
    write_cover(wb)
    write_coverage(wb)
    write_mains(wb)
    wsM = write_register(wb)
    write_dashboard(wb)
    write_index(wb)
    for sheet in wb.worksheets:
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_setup.paperSize = wsM.PAPERSIZE_A3
        sheet.page_setup.leftMargin = 0.4
        sheet.page_setup.rightMargin = 0.4
        sheet.sheet_view.zoomScale = 90 if sheet.title.startswith("03") else 100
    wb.save(OUT)
    wt = Counter(t["work_type"] for t in tasks)
    print(f"Wrote {len(tasks)} sub-tasks  |  {len({t['main_id'] for t in tasks})} main tasks  |  {len(coverage)} PDF lines")
    print("Work types:", dict(wt))
    print("→", OUT)


if __name__ == "__main__":
    main()

